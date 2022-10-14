import json
from unittest import skip
import numpy as np
import os

from openpyxl import load_workbook

from datetime import datetime
from dateutil.relativedelta import relativedelta
from unittest.mock import Mock, patch

from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.contrib import auth
from django.conf import settings
from django.contrib.auth.models import Group
from django.contrib.gis.geos import Point
from django.utils import timezone
from django.http import QueryDict

from rest_framework import status
from rest_framework.authtoken.models import Token
from rest_framework.test import APITestCase
from waffle.testutils import override_switch, override_flag

from api.constants import (ProcessingAction,
                           UpdateLocationParams,
                           FeatureGroups)
from api.models import (Facility, FacilityList, FacilityListItem,
                        FacilityClaim, FacilityClaimReviewNote,
                        FacilityMatch, FacilityAlias, Contributor, User,
                        RequestLog, FacilityLocation, Source,
                        ApiLimit, ApiBlock, ContributorNotifications,
                        EmbedConfig, EmbedField, NonstandardField,
                        FacilityActivityReport, ExtendedField, FacilityIndex,
                        index_custom_text)

from api.oar_id import make_oar_id, validate_oar_id
from api.matching import (match_facility_list_items, GazetteerCache,
                          sort_exact_matches)
from api.processing import (parse_facility_list_item,
                            geocode_facility_list_item,
                            reduce_matches, is_string_match,
                            save_match_details)
from api.geocoding import (create_geocoding_params,
                           format_geocoded_address_data,
                           geocode_address)
from api.test_data import parsed_city_hall_data
from api.permissions import referring_host_is_allowed, referring_host
from api.serializers import (ApprovedFacilityClaimSerializer,
                             FacilityCreateBodySerializer,
                             FacilityListSerializer)
from api.limits import check_api_limits, get_end_of_year
from api.close_list import close_list
from api.facility_type_processing_type import (
    FACILITY_TYPE, PROCESSING_TYPE, EXACT_MATCH, ALIAS_MATCH,
    FUZZY_MATCH, ALL_PROCESSING_TYPES, ALL_FACILITY_TYPES,
    PROCESSING_TYPES_TO_FACILITY_TYPES, ASSEMBLY,
    RAW_MATERIAL_PROCESSING_OR_PRODUCTION,
    WET_ROLLER_PRINTING,
    get_facility_and_processing_type
)
from api.extended_fields import MAX_PRODUCT_TYPE_COUNT


class FacilityListCreateTest(APITestCase):
    def setUp(self):
        self.email = 'test@example.com'
        self.password = 'password'
        self.name = 'Test User'
        self.user = User(email=self.email)
        self.user.set_password(self.password)
        self.user.save()

        self.client.post('/user-login/',
                         {"email": self.email,
                          "password": self.password},
                         format="json")

        self.contributor = Contributor(name=self.name, admin=self.user)
        self.contributor.save()
        self.test_csv_rows = [
            'country,name,address,extra_1',
            'US,Somewhere,999 Park St',
            'US,Someplace Else,1234 Main St',
        ]
        self.test_file = SimpleUploadedFile(
            'facilities.csv',
            b'\n'.join([s.encode() for s in self.test_csv_rows]),
            content_type='text/csv')
        self.test_file_with_bom = SimpleUploadedFile(
            'facilities_with_bom.csv',
            b'\n'.join([self.test_csv_rows[0].encode('utf-8-sig')] +
                       [s.encode() for s in self.test_csv_rows[1:]]),
            content_type='text/csv')

        lists_dir = '/usr/local/src/api/management/commands/facility_lists/'

        with open(os.path.join(lists_dir, '12.xlsx'), 'rb') as xlsx:
            self.test_file_xlsx = SimpleUploadedFile(
                '12.xlsx',
                xlsx.read(),
                content_type='application/vnd.openxmlformats-'
                             'officedocument.spreadsheetml.sheet')

    def post_header_only_file(self, **kwargs):
        if kwargs is None:
            kwargs = {}
        csv_file = SimpleUploadedFile('facilities.csv',
                                      b'country,name,address\n',
                                      content_type='text/csv')
        return self.client.post(reverse('facility-list-list'),
                                {'file': csv_file, **kwargs},
                                format='multipart')

    def test_can_post_file_with_bom(self):
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': self.test_file_with_bom},
                                    format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        new_list = FacilityList.objects.last()
        self.assertEqual(self.test_csv_rows[0], new_list.header)

    def test_creates_list_and_items(self):
        previous_list_count = FacilityList.objects.all().count()
        previous_item_count = FacilityListItem.objects.all().count()
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': self.test_file},
                                    format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(FacilityList.objects.all().count(),
                         previous_list_count + 1)
        self.assertEqual(FacilityListItem.objects.all().count(),
                         previous_item_count + len(self.test_csv_rows) - 1)

    @skip('DB is read-only')
    def test_creates_nonstandard_fields(self):
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': self.test_file},
                                    format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        fields = NonstandardField.objects.filter(
            contributor=self.user.contributor).values_list(
            'column_name', flat=True)
        self.assertEquals(1, len(fields))
        self.assertIn('extra_1', fields)

    def test_creates_source(self):
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': self.test_file},
                                    format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        new_list = FacilityList.objects.last()
        self.assertTrue(Source.objects.filter(facility_list=new_list).exists())
        source = Source.objects.filter(facility_list=new_list).first()
        items = list(FacilityListItem.objects.all())
        self.assertEqual(items[0].raw_data, self.test_csv_rows[1])
        for item in items:
            self.assertEqual(source, item.source)

    def test_creates_list_and_items_xlsx(self):
        previous_list_count = FacilityList.objects.all().count()
        previous_item_count = FacilityListItem.objects.all().count()
        response = self.client.post(reverse('facility-list-list'),
                                    {'name': 'creates_list_and_items_xlsx',
                                     'file': self.test_file_xlsx},
                                    format='multipart')
        self.test_file_xlsx.seek(0)
        wb = load_workbook(filename=self.test_file_xlsx)
        ws = wb[wb.sheetnames[0]]
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(FacilityList.objects.all().count(),
                         previous_list_count + 1)
        self.assertEqual(FacilityListItem.objects.all().count(),
                         previous_item_count + ws.max_row - 1)
        items = list(FacilityListItem.objects.all().order_by('row_index'))
        self.assertEqual(items[0].raw_data, '"{}"'.format(
            '","'.join([cell.value for cell in ws[2]])))

    def test_file_required(self):
        response = self.client.post(reverse('facility-list-list'))
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(json.loads(response.content), ['No file specified.'])

    def test_empty_file_is_invalid(self):
        csv_file = SimpleUploadedFile('facilities.csv', b'',
                                      content_type='text/csv')
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': csv_file},
                                    format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(json.loads(response.content),
                         ['Header cannot be blank.'])

    def test_file_has_invalid_header(self):
        previous_list_count = FacilityList.objects.all().count()
        csv_file = SimpleUploadedFile('facilities.csv', b'foo,bar,baz\n',
                                      content_type='text/csv')
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': csv_file},
                                    format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            json.loads(response.content),
            ['Header must contain country, name, and address fields.'])
        self.assertEqual(FacilityList.objects.all().count(),
                         previous_list_count)

    def test_file_has_missing_header_field(self):
        previous_list_count = FacilityList.objects.all().count()
        csv_file = SimpleUploadedFile('facilities.csv', b'country,address\n',
                                      content_type='text/csv')
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': csv_file},
                                    format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            json.loads(response.content),
            ['Header must contain country, name, and address fields.'])
        self.assertEqual(FacilityList.objects.all().count(),
                         previous_list_count)

    def test_file_and_name_specified(self):
        name = 'A list of facilities'
        response = self.post_header_only_file(name=name)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        response_json = json.loads(response.content)
        new_list = FacilityList.objects.get(id=response_json['id'])
        self.assertEqual(new_list.name, name)

    def test_replaces_must_be_numeric(self):
        previous_list_count = FacilityList.objects.all().count()
        response = self.post_header_only_file(replaces='BAD')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            json.loads(response.content),
            ['"replaces" must be an integer ID.'])
        self.assertEqual(FacilityList.objects.all().count(),
                         previous_list_count)

    def test_replaces_must_be_a_list_id(self):
        previous_list_count = FacilityList.objects.all().count()
        response = self.post_header_only_file(replaces=0)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            json.loads(response.content),
            ['0 is not a valid FacilityList ID.'])
        self.assertEqual(FacilityList.objects.all().count(),
                         previous_list_count)

    def test_replaces(self):
        previous_list_count = FacilityList.objects.all().count()

        # Upload original
        response = self.post_header_only_file()
        self.assertEqual(response.status_code,
                         status.HTTP_200_OK)
        response_json = json.loads(response.content)
        original_list = FacilityList.objects.get(pk=response_json['id'])

        # Upload replacement
        response = self.post_header_only_file(replaces=response_json['id'])
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        response_json = json.loads(response.content)
        new_list = FacilityList.objects.get(pk=response_json['id'])

        self.assertEqual(FacilityList.objects.all().count(),
                         previous_list_count + 2)
        self.assertEqual(new_list.replaces.id, original_list.id)

    def test_cant_replace_twice(self):
        previous_list_count = FacilityList.objects.all().count()

        # Upload original
        response = self.post_header_only_file()
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        response_json = json.loads(response.content)

        # Upload replacement
        response = self.post_header_only_file(replaces=response_json['id'])
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Attempt second replacement
        response = self.post_header_only_file(replaces=response_json['id'])
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            json.loads(response.content),
            ['FacilityList {} has already been replaced.'
                .format(response_json['id'])])

        self.assertEqual(FacilityList.objects.all().count(),
                         previous_list_count + 2)

    def test_user_must_be_authenticated(self):
        self.client.post('/user-logout/')
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': self.test_file},
                                    format='multipart')
        self.assertEqual(response.status_code, 401)

    def test_upload_with_authentication_token_succeeds(self):
        token = Token.objects.create(user=self.user)
        self.client.post('/user-logout/')
        header = {'HTTP_AUTHORIZATION': "Token {0}".format(token)}
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': self.test_file},
                                    format='multipart',
                                    **header)
        self.assertEqual(response.status_code, 200)

    def test_get_request_for_user_with_no_lists_returns_empty_array(self):
        response = self.client.get(reverse('facility-list-list'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), [])
        self.assertEqual(len(response.json()), 0)

    def test_get_request_for_user_with_test_file_list_returns_items(self):
        self.client.post(reverse('facility-list-list'),
                         {'file': self.test_file},
                         format='multipart')
        response = self.client.get(reverse('facility-list-list'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 1)

    def test_get_request_for_unauthenticated_user_returns_401(self):
        self.client.post('/user-logout/')
        response = self.client.get(reverse('facility-list-list'))
        self.assertEqual(response.status_code, 401)

    def test_get_with_authentication_token_returns_items(self):
        token = Token.objects.create(user=self.user)
        self.client.post('/user-logout/')
        header = {'HTTP_AUTHORIZATION': "Token {0}".format(token)}
        self.client.post(reverse('facility-list-list'),
                         {'file': self.test_file},
                         format='multipart',
                         **header)
        response = self.client.get(reverse('facility-list-list'),
                                   **header)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 1)

    def test_upload_by_user_with_no_contributor_returns_402(self):
        Contributor.objects.all().delete()
        token = Token.objects.create(user=self.user)
        self.client.post('/user-logout/')
        header = {'HTTP_AUTHORIZATION': "Token {0}".format(token)}
        response = self.client.post(reverse('facility-list-list'),
                                    {'file': self.test_file},
                                    format='multipart',
                                    **header)
        self.assertEqual(response.status_code, 402)

    def test_list_request_by_user_with_no_contributor_returns_400(self):
        Contributor.objects.all().delete()
        response = self.client.get(reverse('facility-list-list'))
        self.assertEqual(response.status_code, 400)


class ProcessingTestCase(TestCase):
    def get_first_status(self, item, action):
        return next(
            r for r in item.processing_results if r['action'] == action)

    def assert_status(self, item, action):
        self.assertIsNotNone(self.get_first_status(item, action))


class FacilityListItemParseTest(ProcessingTestCase):
    def assert_successful_parse_results(self, item):
        self.assertEqual(FacilityListItem.PARSED, item.status)
        self.assert_status(item, ProcessingAction.PARSE)
        results = self.get_first_status(item, ProcessingAction.PARSE)
        self.assertTrue('error' in results)
        self.assertFalse(results['error'])
        self.assertTrue('started_at' in results)
        self.assertTrue('finished_at' in results)
        self.assertTrue(results['finished_at'] > results['started_at'])

    def assert_failed_parse_results(self, item, message=None):
        self.assertEqual(FacilityListItem.ERROR_PARSING, item.status)
        self.assert_status(item, ProcessingAction.PARSE)
        results = self.get_first_status(item, ProcessingAction.PARSE)
        self.assertTrue('error' in results)
        self.assertTrue(results['error'])
        self.assertTrue('message' in results)
        if message is not None:
            self.assertEqual(message, results['message'])
        self.assertTrue('started_at' in results)
        self.assertTrue('finished_at' in results)
        self.assertTrue(results['finished_at'] > results['started_at'])

    def test_expects_facility_list_item(self):
        self.assertRaises(ValueError, parse_facility_list_item, 'bad')

    def test_raises_if_item_is_not_uploaded(self):
        item = FacilityListItem(status=FacilityListItem.ERROR)
        self.assertRaises(ValueError, parse_facility_list_item, item)

    def test_parses_using_header(self):
        facility_list = FacilityList.objects.create(
            header='address,country,name')
        source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=facility_list)
        item = FacilityListItem(raw_data='1234 main st,de,Shirts!',
                                source=source)
        parse_facility_list_item(item)
        self.assert_successful_parse_results(item)
        self.assertEqual('DE', item.country_code)
        self.assertEqual('Shirts!', item.name)
        self.assertEqual('1234 main st', item.address)

    def test_converts_country_name_to_code(self):
        facility_list = FacilityList.objects.create(
            header='address,country,name')
        source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=facility_list)
        item = FacilityListItem(raw_data='1234 main st,ChInA,Shirts!',
                                source=source)
        parse_facility_list_item(item)
        self.assert_successful_parse_results(item)
        self.assertEqual('CN', item.country_code)
        self.assertEqual('Shirts!', item.name)
        self.assertEqual('1234 main st', item.address)

    def test_error_status_if_country_is_unknown(self):
        facility_list = FacilityList.objects.create(
            header='address,country,name')
        source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=facility_list)
        item = FacilityListItem(raw_data='1234 main st,Unknownistan,Shirts!',
                                source=source)
        parse_facility_list_item(item)
        self.assert_failed_parse_results(
            item, 'Could not find a country code for "Unknownistan".')

    def test_ppe_field_parsing(self):
        facility_list = FacilityList.objects.create(
            header=('address,country,name,ppe_product_types,ppe_contact_phone,'
                    'ppe_contact_email,ppe_website'))
        source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=facility_list)
        item = FacilityListItem(
            raw_data=('1234 main st,de,Shirts!, Mask | Gloves ,123-456-7890,'
                      'ppe@example.com,https://example.com/ppe'),
            source=source)
        parse_facility_list_item(item)
        self.assert_successful_parse_results(item)
        self.assertEqual(['Mask', 'Gloves'], item.ppe_product_types)
        self.assertEqual('123-456-7890', item.ppe_contact_phone)
        self.assertEqual('ppe@example.com', item.ppe_contact_email)
        self.assertEqual('https://example.com/ppe', item.ppe_website)

    def test_ppe_product_type_empty_values(self):
        facility_list = FacilityList.objects.create(
            header='address,country,name,ppe_product_types')
        source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=facility_list)
        # The trailing space is important as we are testing a literally
        # non-empty but logically empty value
        item = FacilityListItem(
            raw_data='1234 main st,de,Shirts!,| ',
            source=source)
        parse_facility_list_item(item)
        self.assert_successful_parse_results(item)
        self.assertEqual([], item.ppe_product_types)


class UserTokenGenerationTest(TestCase):
    def setUp(self):
        self.email = "test@example.com"
        self.password = "example123"
        self.user = User.objects.create(email=self.email)
        self.user.set_password(self.password)
        self.user.save()

    def test_user_does_not_have_a_token_created_on_login(self):
        login_response = self.client.post('/user-login/',
                                          {"email": self.email,
                                           "password": self.password},
                                          format="json")
        self.assertEqual(login_response.status_code, 200)
        user = auth.get_user(self.client)
        self.assertTrue(user.is_authenticated)
        token = Token.objects.filter(user=self.user)
        self.assertEqual(token.count(), 0)

    def test_generated_token_is_not_deleted_on_logout(self):
        self.client.login(email=self.email, password=self.password)
        Token.objects.create(user=self.user)
        logout_response = self.client.post('/user-logout/')
        self.assertEqual(logout_response.status_code, 204)
        user = auth.get_user(self.client)
        self.assertFalse(user.is_authenticated)
        token = Token.objects.filter(user=self.user)
        self.assertEqual(token.count(), 1)


class GeocodingUtilsTest(TestCase):
    def setUp(self):
        settings.GOOGLE_SERVER_SIDE_API_KEY = "world"

    def test_geocoding_params_are_created_correctly(self):
        self.assertEqual(
            create_geocoding_params("hello", "US"),
            {
                'components': 'country:US',
                'address': 'hello',
                'key': 'world',
            }
        )

    def test_geocoded_address_data_is_formatted_correctly(self):
        data = parsed_city_hall_data['full_response']
        formatted_data = format_geocoded_address_data(
            data, data['results'][0])
        self.assertEqual(formatted_data, parsed_city_hall_data)


geocoding_data = {'results': [{'address_components': [
    {'long_name': '990', 'short_name': '990', 'types': ['street_number']},
    {'long_name': 'Spring Garden Street', 'short_name': 'Spring Garden St',
     'types': ['route']},
    {'long_name': 'Center City', 'short_name': 'Center City',
     'types': ['neighborhood', 'political']},
    {'long_name': 'Philadelphia', 'short_name': 'Philadelphia',
     'types': ['locality', 'political']},
    {'long_name': 'Philadelphia County', 'short_name': 'Philadelphia County',
     'types': ['administrative_area_level_2', 'political']},
    {'long_name': 'Pennsylvania', 'short_name': 'PA',
     'types': ['administrative_area_level_1', 'political']},
    {'long_name': 'United States', 'short_name': 'US',
     'types': ['country', 'political']},
    {'long_name': '19123', 'short_name': '19123', 'types': ['postal_code']}
    ],
    'formatted_address': '990 Spring Garden St, Philadelphia, PA 19123, USA',
    'geometry': {'bounds': {
        'northeast': {'lat': 39.9614743, 'lng': -75.15379639999999},
        'southwest': {'lat': 39.9611391, 'lng': -75.1545269}},
        'location': {'lat': 39.961265, 'lng': -75.15412760000001},
    'location_type': 'ROOFTOP',
    'viewport': {
        'northeast': {'lat': 39.9626556802915, 'lng': -75.1528126697085},
        'southwest': {'lat': 39.9599577197085, 'lng': -75.1555106302915}}},
    'place_id': 'ChIJ8cV_ZH_IxokRA_ETpdB5R3Y',
    'types': ['premise']}],
    'status': 'OK'}

geocoding_data_no_country = {
   "results": [
      {
         "address_components": [
            {
               "long_name": "4WHM+QCX",
               "short_name": "4WHM+QCX",
               "types": ["plus_code"]
            },
            {
               "long_name": "Famagusta",
               "short_name": "Famagusta",
               "types": ["locality", "political"]
            },
            {
               "long_name": "99450",
               "short_name": "99450",
               "types": ["postal_code"]
            }
         ],
         "formatted_address": "4WHM+QCX, Famagusta 99450",
         "geometry": {
            "location": {"lat": 35.1294866, "lng": 33.9336133},
            "location_type": "GEOMETRIC_CENTER",
            "viewport": {
               "northeast": {
                  "lat": 35.1308355802915,
                  "lng": 33.9349622802915
               },
               "southwest": {
                  "lat": 35.1281376197085,
                  "lng": 33.9322643197085
               }
            }
         },
         "partial_match": True,
         "place_id": "ChIJDzTqezLI3xQRW5kjGEGLRzA",
         "types": ["establishment", "point_of_interest"]
      }
   ],
   "status": "OK"
}

geocoding_data_second_country = {
   "results": [
      {
         "address_components": [
            {
               "long_name": "Noor Bagh",
               "short_name": "Noor Bagh",
               "types": ["political", "sublocality", "sublocality_level_1"]
            },
            {
               "long_name": "Srinagar",
               "short_name": "Srinagar",
               "types": ["locality", "political"]
            },
            {
               "long_name": "190009",
               "short_name": "190009",
               "types": ["postal_code"]
            }
         ],
         "formatted_address": "Noor Bagh, Srinagar 190009",
         "geometry": {
            "bounds": {
               "northeast": {"lat": 34.07490560000001, "lng": 74.8043599},
               "southwest": {"lat": 34.0734423, "lng": 74.8023449}
            },
            "location": {"lat": 34.0742387, "lng": 74.8035549},
            "location_type": "APPROXIMATE",
            "viewport": {
               "northeast": {
                  "lat": 34.07552293029151,
                  "lng": 74.8047013802915
               },
               "southwest": {
                  "lat": 34.07282496970851,
                  "lng": 74.80200341970848
               }
            }
         },
         "partial_match": True,
         "place_id": "ChIJmQlkN-2P4TgR6u3glWeOMTE",
         "types": ["political", "sublocality", "sublocality_level_1"]
      },
      {
         "address_components": [
            {
               "long_name": "Kaliakair",
               "short_name": "Kaliakair",
               "types": ["locality", "political"]
            },
            {
               "long_name": "Gazipur District",
               "short_name": "Gazipur District",
               "types": ["administrative_area_level_2", "political"]
            },
            {
               "long_name": "Dhaka Division",
               "short_name": "Dhaka Division",
               "types": ["administrative_area_level_1", "political"]
            },
            {
               "long_name": "Bangladesh",
               "short_name": "BD",
               "types": ["country", "political"]
            }
         ],
         "formatted_address": "Kaliakair, Bangladesh",
         "geometry": {
            "bounds": {
               "northeast": {"lat": 24.0840819, "lng": 90.2365494},
               "southwest": {"lat": 24.0535966, "lng": 90.2030754}
            },
            "location": {"lat": 24.0694528, "lng": 90.2221213},
            "location_type": "APPROXIMATE",
            "viewport": {
               "northeast": {
                  "lat": 24.0840819,
                  "lng": 90.2365494
               },
               "southwest": {
                  "lat": 24.0535966,
                  "lng": 90.2030754
               }
            }
         },
         "partial_match":  True,
         "place_id": "ChIJTfs3ierjVTcRysKYnbOKmZM",
         "types": ["locality", "political"]
      }
   ],
   "status": "OK"
}


class GeocodingTest(TestCase):
    @patch('api.geocoding.requests.get')
    def test_geocode_response_contains_expected_keys(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        geocoded_data = geocode_address('990 Spring Garden St, Philly', 'US')
        self.assertIn('full_response', geocoded_data)
        self.assertIn('geocoded_address', geocoded_data)
        self.assertIn('geocoded_point', geocoded_data)
        self.assertIn('lat', geocoded_data['geocoded_point'])
        self.assertIn('lng', geocoded_data['geocoded_point'])

    @patch('api.geocoding.requests.get')
    def test_ungeocodable_address_returns_zero_resusts(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = {'results': [],
                                                   'status': 'ZERO_RESULTS'}
        results = geocode_address('@#$^@#$^', 'XX')
        self.assertEqual(0, results['result_count'])

    @patch('api.geocoding.requests.get')
    def test_incorrect_country_code_raises_error(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data

        with self.assertRaises(ValueError) as cm:
            geocode_address('Datong Bridge, Qiucun town, Fenghua District, ' +
                            'Tirupur, Tamilnadu, 641604', 'IN')

        self.assertEqual(
            cm.exception.args,
            ("Geocoding results of US did not match " +
             "provided country code of IN.",)
        )

    @patch('api.geocoding.requests.get')
    def test_accepts_inexact_address(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data_no_country

        expected_result = geocoding_data_no_country['results'][0]
        expected_point = expected_result["geometry"]["location"]
        expected_address = expected_result["formatted_address"]

        results = geocode_address('PortİSBİ Serbest Bölge Office:4, ' +
                                  'Gazimağusa, North Cyprus, 99450',
                                  'TR')
        self.assertEqual(expected_point, results["geocoded_point"])
        self.assertEqual(expected_address, results['geocoded_address'])

    @patch('api.geocoding.requests.get')
    def test_accepts_alternate_address(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data_second_country

        expected_result = geocoding_data_second_country['results'][1]
        expected_point = expected_result["geometry"]["location"]
        expected_address = expected_result["formatted_address"]

        results = geocode_address('Noorbagh, Kaliakoir Gazipur Dhaka 1704',
                                  'BD')
        self.assertEqual(expected_point, results["geocoded_point"])
        self.assertEqual(expected_address, results['geocoded_address'])

    @patch('api.geocoding.requests.get')
    def test_geocode_non_200_response(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=400)
        with self.assertRaisesRegexp(ValueError, '400'):
            geocode_address('Noorbagh, Kaliakoir Gazipur Dhaka 1704', 'BD')


class FacilityAndProcessingTypeTest(TestCase):
    def test_exact_processing_type_match(self):
        processing_type_input = 'assembly'
        expected_output = (
            PROCESSING_TYPE, EXACT_MATCH,
            PROCESSING_TYPES_TO_FACILITY_TYPES[
                ALL_PROCESSING_TYPES[ASSEMBLY]
            ],
            ALL_PROCESSING_TYPES[ASSEMBLY]
        )
        output = get_facility_and_processing_type(processing_type_input)
        self.assertEqual(output, expected_output)

    def test_exact_facility_type_match(self):
        facility_type_input = 'raw material processing or production'
        facility_type_value = ALL_FACILITY_TYPES[facility_type_input]
        expected_output = (
            FACILITY_TYPE, EXACT_MATCH,
            facility_type_value,
            facility_type_value,
        )
        output = get_facility_and_processing_type(facility_type_input)
        self.assertEqual(output, expected_output)

    def test_alias_processing_type_match(self):
        processing_type_input = 'wet printing'
        expected_output = (
            PROCESSING_TYPE, ALIAS_MATCH,
            PROCESSING_TYPES_TO_FACILITY_TYPES[
                ALL_PROCESSING_TYPES[WET_ROLLER_PRINTING]
            ],
            ALL_PROCESSING_TYPES[WET_ROLLER_PRINTING]
        )
        output = get_facility_and_processing_type(processing_type_input)
        self.assertEqual(output, expected_output)

    def test_fuzzy_processing_type_match(self):
        processing_type_input = 'asembley'
        expected_output = (
            PROCESSING_TYPE, FUZZY_MATCH,
            PROCESSING_TYPES_TO_FACILITY_TYPES[
                ALL_PROCESSING_TYPES[ASSEMBLY]
            ],
            ALL_PROCESSING_TYPES[ASSEMBLY]
        )
        output = get_facility_and_processing_type(processing_type_input)
        self.assertEqual(output, expected_output)

    def test_fuzzy_facility_type_match(self):
        facility_type_input = 'raw mater process'
        facility_type_value = ALL_FACILITY_TYPES[
            RAW_MATERIAL_PROCESSING_OR_PRODUCTION
        ]
        expected_output = (
            FACILITY_TYPE, FUZZY_MATCH,
            facility_type_value,
            facility_type_value,
        )
        output = get_facility_and_processing_type(facility_type_input)
        self.assertEqual(output, expected_output)


listitem_geocode_data = {
   "results": [
      {
         "address_components": [
            {
               "long_name": "Linjiacunzhen",
               "short_name": "Linjiacunzhen",
               "types": ["political", "sublocality", "sublocality_level_2"]
            },
            {
               "long_name": "Zhucheng",
               "short_name": "Zhucheng",
               "types": ["political", "sublocality", "sublocality_level_1"]
            },
            {
               "long_name": "Weifang",
               "short_name": "Weifang",
               "types": ["locality", "political"]
            },
            {
               "long_name": "Shandong",
               "short_name": "Shandong",
               "types": ["administrative_area_level_1", "political"]
            },
            {
               "long_name": "China",
               "short_name": "CN",
               "types": ["country", "political"]
            },
            {
               "long_name": "262232",
               "short_name": "262232",
               "types": ["postal_code"]
            }
         ],
         "formatted_address": "Linjiacunzhen, Zhucheng, " +
                              "Weifang, Shandong, China, 262232",
         "geometry": {
            "location": {"lat": 35.994813, "lng": 119.65418},
            "location_type": "APPROXIMATE",
            "viewport": {
               "northeast": {"lat": 36.0038401, "lng": 119.6701874},
               "southwest": {"lat": 35.9857849, "lng": 119.6381726}
            }
         },
         "partial_match": True,
         "place_id": "ChIJV5SPiTEkvjUR7ErhmnSRTR8",
         "types": ["political", "sublocality", "sublocality_level_2"]
      },
      {
         "address_components": [
            {
               "long_name": "396210",
               "short_name": "396210",
               "types": ["postal_code"]
            },
            {
               "long_name": "Daman",
               "short_name": "Daman",
               "types": ["administrative_area_level_2", "political"]
            },
            {
               "long_name": "Dadra and Nagar Haveli and Daman and Diu",
               "short_name": "DH",
               "types": ["administrative_area_level_1", "political"]
            },
            {
               "long_name": "India",
               "short_name": "IN",
               "types": ["country", "political"]
            }
         ],
         "formatted_address":  "Dadra and Nagar Haveli and Daman and " +
                               "Diu 396210, India",
         "geometry": {
            "bounds": {
               "northeast": {"lat": 20.4696847, "lng": 72.8751228},
               "southwest": {"lat": 20.4051972, "lng": 72.82805549999999}
            },
            "location": {"lat": 20.4346424, "lng": 72.8456399},
            "location_type": "APPROXIMATE",
            "viewport": {
               "northeast": {"lat": 20.4696847, "lng": 72.8751228},
               "southwest": {"lat": 20.4051972, "lng": 72.82805549999999}
            }
         },
         "partial_match": True,
         "place_id": "ChIJ8d4EeIra4DsR3xkUSZh_-ng",
         "types": ["postal_code"]
      }
   ],
   "status": "OK"
}


class FacilityListItemGeocodingTest(ProcessingTestCase):
    def test_invalid_argument_raises_error(self):
        with self.assertRaises(ValueError) as cm:
            geocode_facility_list_item("hello")

        self.assertEqual(
            cm.exception.args,
            ('Argument must be a FacilityListItem',)
        )

    def test_unparsed_item_raises_error(self):
        facility_list = FacilityList(header='address,country,name')
        source = Source(source_type=Source.LIST,
                        facility_list=facility_list)
        item = FacilityListItem(
            raw_data='1400 JFK Blvd, Philly,us,Shirts!',
            source=source
        )

        with self.assertRaises(ValueError) as cm:
            geocode_facility_list_item(item)

        self.assertEqual(
            cm.exception.args,
            ('Items to be geocoded must be in the PARSED status',),
        )

    @patch('api.geocoding.requests.get')
    def test_nested_correct_country_code_succeeds(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = listitem_geocode_data
        facility_list = FacilityList.objects.create(
            header='address,country,name')
        source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=facility_list)
        item = FacilityListItem(
            raw_data='"Linjiacun Town, Zhucheng City Weifang, Daman, ' +
                     'Daman, 396210",IN,Shirts!',
            source=source
        )
        parse_facility_list_item(item)
        geocode_facility_list_item(item)

        expected_result = listitem_geocode_data['results'][1]
        expected_address = expected_result['formatted_address']

        self.assertEqual(item.status, FacilityListItem.GEOCODED)
        self.assertEqual(item.geocoded_address, expected_address)
        self.assertIsInstance(item.geocoded_point, Point)

    @patch('api.geocoding.requests.get')
    def test_incorrect_country_code_has_error_status(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = listitem_geocode_data
        facility_list = FacilityList.objects.create(
            header='address,country,name')
        source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=facility_list)
        item = FacilityListItem(
            raw_data='"Linjiacun Town, Zhucheng City Weifang, Daman, ' +
                     'Daman, 396210",BD,Shirts!',
            source=source
        )
        parse_facility_list_item(item)
        geocode_facility_list_item(item)

        self.assertEqual(item.status, FacilityListItem.ERROR_GEOCODING)
        self.assertIsNone(item.geocoded_address)
        self.assertIsNone(item.geocoded_point)

    def test_successfully_geocoded_item_has_correct_results(self):
        facility_list = FacilityList.objects.create(
            header='address,country,name')
        source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=facility_list)
        item = FacilityListItem(
            raw_data='"City Hall, Philly, PA",us,Shirts!',
            source=source
        )

        parse_facility_list_item(item)
        geocode_facility_list_item(item)

        self.assertIsNotNone(item.geocoded_address)
        self.assertIsInstance(item.geocoded_point, Point)
        self.assertEqual(item.status, FacilityListItem.GEOCODED)
        self.assertIn(
            'results',
            self.get_first_status(item, ProcessingAction.GEOCODE)['data']
        )

    def test_failed_geocoded_item_has_no_resuts_status(self):
        facility_list = FacilityList.objects.create(
            header='address,country,name')
        source = Source.objects.create(source_type=Source.LIST,
                                       facility_list=facility_list)
        item = FacilityListItem(
            raw_data='"hello, world, foo, bar, baz",us,Shirts!',
            source=source
        )
        parse_facility_list_item(item)
        item.country_code = "$%"
        geocode_facility_list_item(item)

        self.assertEqual(item.status, FacilityListItem.GEOCODED_NO_RESULTS)
        self.assertIsNone(item.geocoded_address)
        self.assertIsNone(item.geocoded_point)


class FacilityNamesAddressesAndContributorsTest(TestCase):
    def setUp(self):
        self.name_one = 'name_one'
        self.name_two = 'name_two'
        self.address_one = 'address_one'
        self.address_two = 'address_two'
        self.email_one = 'one@example.com'
        self.email_two = 'two@example.com'
        self.contrib_one_name = 'contributor one'
        self.contrib_two_name = 'contributor two'
        self.country_code = 'US'
        self.list_one_name = 'one'
        self.list_two_name = 'two'
        self.user_one = User.objects.create(email=self.email_one)
        self.user_two = User.objects.create(email=self.email_two)

        self.contrib_one = Contributor \
            .objects \
            .create(admin=self.user_one,
                    name=self.contrib_one_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_two = Contributor \
            .objects \
            .create(admin=self.user_two,
                    name=self.contrib_two_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_one = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one",
                    name=self.list_one_name)

        self.source_one = Source \
            .objects \
            .create(facility_list=self.list_one,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contrib_one)

        self.list_item_one = FacilityListItem \
            .objects \
            .create(name=self.name_one,
                    address=self.address_one,
                    country_code=self.country_code,
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_one)

        self.list_two = FacilityList \
            .objects \
            .create(header="header",
                    file_name="two",
                    name=self.list_two_name)

        self.source_two = Source \
            .objects \
            .create(facility_list=self.list_two,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contrib_two)

        self.list_item_two = FacilityListItem \
            .objects \
            .create(name=self.name_two,
                    address=self.address_two,
                    country_code=self.country_code,
                    row_index="2",
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_two)

        self.facility = Facility \
            .objects \
            .create(name=self.name_one,
                    address=self.address_one,
                    country_code=self.country_code,
                    location=Point(0, 0),
                    created_from=self.list_item_one)

        self.extended_field_one = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value=self.name_one,
                contributor=self.contrib_one,
                facility=self.facility,
                facility_list_item=self.list_item_one
            )

        self.extended_field_two = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value=self.name_two,
                contributor=self.contrib_two,
                facility=self.facility,
                facility_list_item=self.list_item_two
            )

        self.facility_match_one = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.CONFIRMED,
                    facility=self.facility,
                    results="",
                    facility_list_item=self.list_item_one)

        self.facility_match_two = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.CONFIRMED,
                    facility=self.facility,
                    results="",
                    facility_list_item=self.list_item_two)

    def test_returns_contributors(self):
        sources = self.facility.sources()
        self.assertIn(self.source_one, sources)
        self.assertIn(self.source_two, sources)
        self.assertEqual(len(sources), 2)

    def test_returns_extended_fields(self):
        fields = self.facility.extended_fields()
        self.assertIn(self.extended_field_one, fields)
        self.assertIn(self.extended_field_two, fields)
        self.assertEqual(len(fields), 2)

    def test_excludes_canonical_name_from_other_names(self):
        other_names = self.facility.other_names()
        self.assertIn(self.name_two, other_names)
        self.assertNotIn(self.name_one, other_names)

    def test_excludes_canonical_address_from_other_addresses(self):
        other_addresses = self.facility.other_addresses()
        self.assertIn(self.address_two, other_addresses)
        self.assertNotIn(self.address_one, other_addresses)

    def test_excludes_other_names_from_inactive_lists(self):
        self.source_two.is_active = False
        self.source_two.save()
        other_names = self.facility.other_names()
        self.assertNotIn(self.name_two, other_names)
        self.assertEqual(len(other_names), 0)

    def test_excludes_other_addresses_from_inactive_lists(self):
        self.source_two.is_active = False
        self.source_two.save()
        other_addresses = self.facility.other_addresses()
        self.assertNotIn(self.address_two, other_addresses)
        self.assertEqual(len(other_addresses), 0)

    def test_excludes_contributors_from_inactive_lists(self):
        self.source_two.is_active = False
        self.source_two.save()
        sources = self.facility.sources()
        self.assertIn(self.source_one, sources)
        self.assertNotIn(self.source_two, sources)

    def test_excludes_other_names_from_non_public_lists(self):
        self.source_two.is_active = False
        self.source_two.save()
        other_names = self.facility.other_names()
        self.assertNotIn(self.name_two, other_names)
        self.assertEqual(len(other_names), 0)

    def test_excludes_other_addresses_from_non_public_lists(self):
        self.source_two.is_active = False
        self.source_two.save()
        other_addresses = self.facility.other_addresses()
        self.assertNotIn(self.address_two, other_addresses)
        self.assertEqual(len(other_addresses), 0)

    def test_excludes_contributors_from_non_public_lists(self):
        self.source_two.is_active = False
        self.source_two.save()
        sources = self.facility.sources()
        self.assertIn(self.source_one, sources)
        self.assertNotIn(self.source_two, sources)

    def test_excludes_unmatched_facilities_from_other_names(self):
        self.facility_match_two.status = FacilityMatch.REJECTED
        self.facility_match_two.save()
        other_names = self.facility.other_names()
        self.assertNotIn(self.name_two, other_names)
        self.assertEqual(len(other_names), 0)

    def test_excludes_unmatched_facilities_from_other_addresses(self):
        self.facility_match_two.status = FacilityMatch.REJECTED
        self.facility_match_two.save()
        other_addresses = self.facility.other_addresses()
        self.assertNotIn(self.name_one, other_addresses)
        self.assertEqual(len(other_addresses), 0)

    def test_excludes_unmatched_facilities_from_contributors(self):
        self.facility_match_two.status = FacilityMatch.REJECTED
        self.facility_match_two.save()
        sources = self.facility.sources()
        self.assertIn(self.source_one, sources)
        self.assertNotIn(self.source_two, sources)
        self.assertEqual(len(sources), 1)

    def test_excludes_inactive_facility_matches_from_details(self):
        self.facility_match_two.is_active = False
        self.facility_match_two.save()

        sources = self.facility.sources()
        self.assertIn(self.source_one, sources)
        self.assertNotIn(self.source_two, sources)
        self.assertIn("One Other", sources)
        self.assertEqual(len(sources), 2)

        other_names = self.facility.other_names()
        self.assertNotIn(self.name_two, other_names)
        self.assertEqual(len(other_names), 0)

        other_addresses = self.facility.other_addresses()
        self.assertNotIn(self.address_two, other_addresses)
        self.assertEqual(len(other_addresses), 0)

        fields = self.facility.extended_fields()
        self.assertIn(self.extended_field_one, fields)
        self.assertNotIn(self.extended_field_two, fields)
        self.assertEqual(len(fields), 1)

    def test_excludes_private_matches_from_details(self):
        self.source_two.is_public = False
        self.source_two.save()

        sources = self.facility.sources()
        self.assertIn(self.source_one, sources)
        self.assertNotIn(self.source_two, sources)
        self.assertIn("One Other", sources)
        self.assertEqual(len(sources), 2)

        other_names = self.facility.other_names()
        self.assertNotIn(self.name_two, other_names)
        self.assertEqual(len(other_names), 0)

        other_addresses = self.facility.other_addresses()
        self.assertNotIn(self.address_two, other_addresses)
        self.assertEqual(len(other_addresses), 0)


class ConfirmRejectAndRemoveAndDissociateFacilityMatchTest(TestCase):
    def setUp(self):
        self.country_code = 'US'

        self.prior_user_name = 'prior_user_name'
        self.prior_user_email = 'prioruser@example.com'
        self.prior_contrib_name = 'prior_contrib_name'
        self.prior_list_name = 'prior_list_name'
        self.prior_user = User.objects.create(email=self.prior_user_email)
        self.prior_address_one = 'prior_address_one'
        self.prior_address_two = 'prior_address_two'
        self.prior_name_one = 'prior_name_one'
        self.prior_name_two = 'prior_name_two'
        self.prior_user_password = 'example-password'
        self.prior_user.set_password(self.prior_user_password)
        self.prior_user.save()

        self.prior_contrib = Contributor \
            .objects \
            .create(admin=self.prior_user,
                    name=self.prior_contrib_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.prior_list = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one",
                    name=self.prior_list_name)

        self.prior_source = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.prior_list,
                    contributor=self.prior_contrib)

        self.prior_list_item_one = FacilityListItem \
            .objects \
            .create(name=self.prior_name_one,
                    address=self.prior_address_one,
                    country_code=self.country_code,
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.prior_source)

        self.prior_facility_one = Facility \
            .objects \
            .create(name=self.prior_list_item_one.name,
                    address=self.prior_list_item_one.address,
                    country_code=self.country_code,
                    location=Point(0, 0),
                    created_from=self.prior_list_item_one)

        self.prior_facility_match_one = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.prior_facility_one,
                    results="",
                    facility_list_item=self.prior_list_item_one)

        self.prior_list_item_two = FacilityListItem \
            .objects \
            .create(name=self.prior_name_two,
                    address=self.prior_address_two,
                    country_code=self.country_code,
                    row_index=2,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.prior_source)

        self.prior_facility_two = Facility \
            .objects \
            .create(name=self.prior_list_item_two.name,
                    address=self.prior_list_item_two.address,
                    country_code=self.country_code,
                    location=Point(0, 0),
                    created_from=self.prior_list_item_two)

        self.prior_facility_match_one = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.prior_facility_two,
                    results="",
                    facility_list_item=self.prior_list_item_one)

        self.current_user_name = 'current_user_name'
        self.current_user_email = 'currentuser@example.com'
        self.current_contrib_name = 'current_contrib_name'
        self.current_list_name = 'current_list_name'
        self.current_user = User.objects.create(email=self.current_user_email)
        self.current_user_password = "password123"
        self.current_user.set_password(self.current_user_password)
        self.current_user.save()

        self.current_address = 'current_address'
        self.current_name = 'current_name'

        self.current_contrib = Contributor \
            .objects \
            .create(admin=self.current_user,
                    name=self.current_contrib_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.current_list = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one",
                    name=self.current_list_name)

        self.current_source = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.current_list,
                    contributor=self.current_contrib)

        self.current_list_item = FacilityListItem \
            .objects \
            .create(name=self.current_name,
                    address=self.current_address,
                    country_code=self.country_code,
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.POTENTIAL_MATCH,
                    source=self.current_source)

        self.potential_facility_match_one = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.PENDING,
                    facility=self.prior_facility_one,
                    results="",
                    facility_list_item=self.current_list_item)

        self.potential_facility_match_two = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.PENDING,
                    facility=self.prior_facility_two,
                    results="",
                    facility_list_item=self.current_list_item)

        self.client.login(email=self.current_user_email,
                          password=self.current_user_password)

        self.confirm_url = '/api/facility-lists/{}/confirm/' \
            .format(self.current_list.id)

        self.reject_url = '/api/facility-lists/{}/reject/' \
            .format(self.current_list.id)

        self.remove_url = '/api/facility-lists/{}/remove/' \
            .format(self.current_list.id)

    def match_url(self, match, action='detail'):
        return reverse('facility-match-{}'.format(action),
                       kwargs={'pk': match.pk})

    def test_confirming_match_rejects_other_potential_matches(self):
        confirm_response = self.client.post(
            self.match_url(self.potential_facility_match_one, action='confirm')
        )

        confirmed_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_one.id)

        rejected_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_two.id)

        self.assertEqual(confirm_response.status_code, 200)
        self.assertEqual(confirmed_match.status, FacilityMatch.CONFIRMED)
        self.assertEqual(rejected_match.status, FacilityMatch.REJECTED)

    def test_confirming_match_changes_list_item_status(self):
        confirm_response = self.client.post(
            self.match_url(self.potential_facility_match_one, action='confirm')
        )

        updated_list_item = FacilityListItem \
            .objects \
            .get(pk=self.current_list_item.id)

        self.assertEqual(confirm_response.status_code, 200)
        self.assertEqual(updated_list_item.status,
                         FacilityListItem.CONFIRMED_MATCH)

    def test_confirming_match_doesnt_create_new_facility(self):
        confirm_response = self.client.post(
            self.match_url(self.potential_facility_match_one, action='confirm')
        )

        facilities = Facility.objects.all()

        self.assertEqual(confirm_response.status_code, 200)
        self.assertEqual(facilities.count(), 2)

    def test_rejecting_last_potential_match_changes_list_item_status(self):
        reject_response_one = self.client.post(
            self.match_url(self.potential_facility_match_one, action='reject')
        )

        reject_response_two = self.client.post(
            self.match_url(self.potential_facility_match_two, action='reject')
        )

        self.assertEqual(reject_response_one.status_code, 200)
        self.assertEqual(reject_response_two.status_code, 200)

        updated_list_item = FacilityListItem \
            .objects \
            .get(pk=self.current_list_item.id)

        self.assertEqual(updated_list_item.status,
                         FacilityListItem.CONFIRMED_MATCH)

    def test_rejecting_one_of_several_matches_changes_match_to_rejected(self):
        reject_response_one = self.client.post(
            self.match_url(self.potential_facility_match_one, action='reject')
        )

        self.assertEqual(reject_response_one.status_code, 200)

        updated_potential_match_one = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_one.id)

        self.assertEqual(updated_potential_match_one.status,
                         FacilityMatch.REJECTED)

        updated_potential_match_two = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_two.id)

        self.assertEqual(updated_potential_match_two.status,
                         FacilityMatch.PENDING)

    def test_rejecting_one_of_several_matches_doesnt_change_item_status(self):
        reject_response_one = self.client.post(
            self.match_url(self.potential_facility_match_one, action='reject')
        )

        self.assertEqual(reject_response_one.status_code, 200)

        updated_list_item = FacilityListItem \
            .objects \
            .get(pk=self.current_list_item.id)

        self.assertEqual(updated_list_item.status,
                         FacilityListItem.POTENTIAL_MATCH)

    def test_rejecting_last_potential_match_creates_new_facility(self):
        reject_response_one = self.client.post(
            self.match_url(self.potential_facility_match_one, action='reject')
        )

        reject_response_two = self.client.post(
            self.match_url(self.potential_facility_match_two, action='reject')
        )

        self.assertEqual(reject_response_one.status_code, 200)
        self.assertEqual(reject_response_two.status_code, 200)

        facilities = Facility.objects.all()
        self.assertEqual(facilities.count(), 3)

    def test_rejecting_last_potential_match_creates_a_new_facility_match(self):
        initial_facility_matches_count = FacilityMatch.objects.all().count()

        reject_response_one = self.client.post(
            self.match_url(self.potential_facility_match_one, action='reject')
        )

        reject_response_two = self.client.post(
            self.match_url(self.potential_facility_match_two, action='reject')
        )

        self.assertEqual(reject_response_one.status_code, 200)
        self.assertEqual(reject_response_two.status_code, 200)

        facilities = Facility.objects.all()
        self.assertEqual(facilities.count(), 3)

        new_facility_matches_count = FacilityMatch.objects.all().count()
        self.assertEqual(
            initial_facility_matches_count + 1,
            new_facility_matches_count,
        )

    def test_removing_a_list_item_sets_its_matches_to_inactive(self):
        confirm_response = self.client.post(
            self.match_url(self.potential_facility_match_one, action='confirm')
        )

        confirmed_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_one.id)

        rejected_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_two.id)

        self.assertEqual(confirm_response.status_code, 200)
        self.assertEqual(confirmed_match.status, FacilityMatch.CONFIRMED)
        self.assertEqual(rejected_match.status, FacilityMatch.REJECTED)
        self.assertEqual(confirmed_match.is_active, True)
        self.assertEqual(rejected_match.is_active, True)

        remove_response = self.client.post(
            self.remove_url,
            {"list_item_id": self.current_list_item.id},
        )

        self.assertEqual(remove_response.status_code, 200)

        updated_confirmed_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_one.id)

        updated_rejected_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_two.id)

        self.assertEqual(updated_confirmed_match.is_active, False)
        self.assertEqual(updated_rejected_match.is_active, False)

    def test_only_list_contribtutor_can_remove_a_list_item(self):
        confirm_response = self.client.post(
            self.match_url(self.potential_facility_match_one, action='confirm')
        )

        confirmed_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_one.id)

        rejected_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_two.id)

        self.assertEqual(confirm_response.status_code, 200)
        self.assertEqual(confirmed_match.status, FacilityMatch.CONFIRMED)
        self.assertEqual(rejected_match.status, FacilityMatch.REJECTED)
        self.assertEqual(confirmed_match.is_active, True)
        self.assertEqual(rejected_match.is_active, True)

        self.client.logout()

        self.client.login(email=self.prior_user_email,
                          password=self.prior_user_password)

        remove_response = self.client.post(
            self.remove_url,
            {"list_item_id": self.current_list_item.id},
        )

        self.assertEqual(remove_response.status_code, 404)

    def test_dissociate_sets_matches_to_inactive(self):
        confirm_response = self.client.post(
            self.match_url(self.potential_facility_match_one, action='confirm')
        )

        confirmed_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_one.id)

        self.assertEqual(confirm_response.status_code, 200)
        self.assertEqual(confirmed_match.is_active, True)

        dissociate_url = reverse('facility-dissociate',
                                 kwargs={'pk': confirmed_match.facility.pk})
        dissociate_response = self.client.post(dissociate_url)

        self.assertEqual(dissociate_response.status_code, 200)

        updated_confirmed_match = FacilityMatch \
            .objects \
            .get(pk=self.potential_facility_match_one.id)

        self.assertEqual(updated_confirmed_match.is_active, False)


def junk_chars(string):
    return 'AA' + string + 'YY'


class DedupeMatchingTests(TestCase):
    fixtures = ['users', 'contributors', 'facility_lists', 'sources',
                'facility_list_items', 'facilities', 'facility_matches']

    def setUp(self):
        self.contributor = Contributor.objects.first()

    def tearDown(self):
        GazetteerCache._gazetter = None
        GazetteerCache._facility_version = None
        GazetteerCache._match_version = None

    def create_list(self, items, status=FacilityListItem.GEOCODED):
        facility_list = FacilityList(
            name='test', description='',
            file_name='test.csv', header='country,name,address')
        facility_list.save()
        source = Source(source_type=Source.LIST, facility_list=facility_list,
                        contributor=self.contributor)
        source.save()
        for index, item in enumerate(items):
            country_code, name, address = item
            list_item = FacilityListItem(
                source=source, row_index=index, raw_data='',
                status=status, name=name, address=address,
                country_code=country_code, geocoded_address='')
            list_item.save()
        return facility_list

    def test_matches(self):
        facility = Facility.objects.first()
        facility_list = self.create_list([
            (facility.country_code, junk_chars(facility.name.upper()),
             junk_chars(facility.address.upper()))])
        facility_list = self.create_list([
            (facility.country_code, facility.name.upper(),
             junk_chars(facility.address.upper()))])
        result = match_facility_list_items(facility_list)
        matches = result['item_matches']
        item_id = str(facility_list.source.facilitylistitem_set.all()[0].id)
        self.assertIn(item_id, matches)
        self.assertEqual(1, len(matches[item_id]))
        self.assertEqual(str(facility.id), matches[item_id][0][0])
        self.assertEqual(0.5, result['results']['gazetteer_threshold'])
        self.assertFalse(result['results']['no_gazetteer_matches'])
        self.assertFalse(result['results']['no_geocoded_items'])

    def assert_match_count_after_delete(
            self, delete_facility=True, match_count=0):
        # First create a list and match it. We use a multiple item list so that
        # there is valid training data available after we delete records later
        # in the test case.
        facility_list = self.create_list([
            (f.country_code, junk_chars(f.name.upper()),
             junk_chars(f.address.upper()))
            for f in Facility.objects.all()[:3]])
        result = match_facility_list_items(facility_list)
        matches = result['item_matches']
        item_ids = [
            str(i.id)
            for i in facility_list.source.facilitylistitem_set.all()]
        matched_ids = [id for id in item_ids if id in matches.keys()]
        self.assertTrue(len(matched_ids) > 0)

        # Now we delete the previously matched facility and try to match it
        # again
        matched_facility_id = matches[matched_ids[0]][0][0]

        facility = Facility.objects.get(pk=matched_facility_id)
        for item in FacilityListItem.objects.filter(facility=facility):
            item.facility = None
            item.save()

        for match in FacilityMatch.objects.filter(facility=facility):
            match.delete()

        facility_list = self.create_list([
            (facility.country_code, facility.name, facility.address)])
        item_id = str(facility_list.source.facilitylistitem_set.all()[0].id)

        if delete_facility:
            facility.delete()

        result = match_facility_list_items(facility_list)
        matches = result['item_matches']
        self.assertEqual(match_count, len(matches[item_id]))

    def test_does_not_match_after_delete_facility_and_match(self):
        self.assert_match_count_after_delete(
            delete_facility=True, match_count=0)

    def test_matches_after_delete_match(self):
        self.assert_match_count_after_delete(
            delete_facility=False, match_count=1)

    def test_does_not_match(self):
        facility_list = self.create_list([
            ('US', 'Azavea', '990 Spring Garden St.')])
        result = match_facility_list_items(facility_list)
        matches = result['item_matches']
        self.assertEqual(0, len(matches))
        self.assertTrue(result['results']['no_gazetteer_matches'])

    def test_no_geocoded_items(self):
        facility = Facility.objects.first()
        facility_list = self.create_list([
            (facility.country_code, junk_chars(facility.name.upper()),
             junk_chars(facility.address.upper()))],
                                         status=FacilityListItem.PARSED)
        result = match_facility_list_items(facility_list)
        self.assertTrue(result['results']['no_geocoded_items'])

    def test_reduce_matches(self):
        matches = [
            ('US2020052GKF19F', 75),
            ('US2020052GKF19F_MATCH-23', 88),
            ('US2020052YDVKBQ', 45)
        ]
        expected = [
            ('US2020052GKF19F', 88),
            ('US2020052YDVKBQ', 45)
        ]
        self.assertEqual(expected, reduce_matches(matches))

    def test_is_string_match(self):
        # The clean function will remove stray characters
        item = FacilityListItem(
            country_code='US',
            name='Pants Ahoy',
            address='123 Main St')
        facility = Facility(
            country_code='US',
            name='"PANTS AHOY"',
            address='123     MAIN     ST')
        self.assertTrue(is_string_match(item, facility))

        # Needs to be an exact character match
        item = FacilityListItem(
            country_code='US',
            name='Pants Ahoy',
            address='123 Main St')
        facility = Facility(
            country_code='US',
            name='Pants Ahoy',
            address='123 Main Street')
        self.assertFalse(is_string_match(item, facility))


class OarIdTests(TestCase):

    def test_make_and_validate_oar_id(self):
        id = make_oar_id('US')
        validate_oar_id(id)
        self.assertEqual(id[:2], 'US')

    def test_id_too_long(self):
        self.assertRaises(ValueError, validate_oar_id, 'US2019070KTWK4x')

    def test_invalid_checksum(self):
        self.assertRaises(ValueError, validate_oar_id, 'USX019070KTWK4')

    def test_invalid_country(self):
        self.assertRaises(ValueError, make_oar_id, '99')


class ContributorsListAPIEndpointTests(TestCase):
    def setUp(self):
        self.name_one = 'name_one'
        self.name_two = 'name_two'
        self.name_three = 'name_three'
        self.name_four = 'name_four'

        self.address_one = 'address_one'
        self.address_two = 'address_two'
        self.address_three = 'address_three'
        self.address_four = 'address_four'

        self.email_one = 'one@example.com'
        self.email_two = 'two@example.com'
        self.email_three = 'three@example.com'
        self.email_four = 'four@example.com'
        self.email_five = 'five@example.com'
        self.email_six = 'six@example.com'
        self.email_seven = 'seven@example.com'

        self.contrib_one_name = 'contributor that should be included'
        self.contrib_two_name = 'contributor with no lists'
        self.contrib_three_name = 'contributor with an inactive list'
        self.contrib_four_name = 'contributor with a non public list'
        self.contrib_five_name = 'contributor with only error items'
        self.contrib_six_name = 'contributor with one good and one error item'
        self.contrib_seven_name = 'contributor with create=False API source'

        self.country_code = 'US'
        self.list_one_name = 'one'
        self.list_one_b_name = 'one-b'
        self.list_three_name = 'three'
        self.list_four_name = 'four'
        self.list_five_name = 'five'
        self.list_six_name = 'six'
        self.list_seven_name = 'seven'

        self.user_one = User.objects.create(email=self.email_one)
        self.user_two = User.objects.create(email=self.email_two)
        self.user_three = User.objects.create(email=self.email_three)
        self.user_four = User.objects.create(email=self.email_four)
        self.user_five = User.objects.create(email=self.email_five)
        self.user_six = User.objects.create(email=self.email_six)
        self.user_seven = User.objects.create(email=self.email_seven)

        self.contrib_one = Contributor \
            .objects \
            .create(admin=self.user_one,
                    name=self.contrib_one_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_two = Contributor \
            .objects \
            .create(admin=self.user_two,
                    name=self.contrib_two_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_three = Contributor \
            .objects \
            .create(admin=self.user_three,
                    name=self.contrib_three_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_four = Contributor \
            .objects \
            .create(admin=self.user_four,
                    name=self.contrib_four_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_five = Contributor \
            .objects \
            .create(admin=self.user_five,
                    name=self.contrib_five_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_six = Contributor \
            .objects \
            .create(admin=self.user_six,
                    name=self.contrib_six_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_seven = Contributor \
            .objects \
            .create(admin=self.user_seven,
                    name=self.contrib_seven_name,
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_one = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one",
                    name=self.list_one_name)

        self.source_one = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_one,
                    contributor=self.contrib_one)

        self.list_item_one = FacilityListItem \
            .objects \
            .create(row_index=0,
                    source=self.source_one,
                    status=FacilityListItem.MATCHED)

        self.list_one_b = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one-b",
                    name=self.list_one_b_name)

        self.source_one_b = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_one_b,
                    contributor=self.contrib_one)

        self.list_item_one_b = FacilityListItem \
            .objects \
            .create(row_index=0,
                    source=self.source_one_b,
                    status=FacilityListItem.MATCHED)

        # Contributor two has no lists

        self.list_three = FacilityList \
            .objects \
            .create(header="header",
                    file_name="three",
                    name=self.list_three_name)

        self.source_three = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_three,
                    is_public=True,
                    is_active=False,
                    contributor=self.contrib_three)

        self.list_four = FacilityList \
            .objects \
            .create(header="header",
                    file_name="four",
                    name=self.list_four_name)

        self.source_four = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_four,
                    is_public=False,
                    is_active=True,
                    contributor=self.contrib_four)

        self.list_five = FacilityList \
            .objects \
            .create(header="header",
                    file_name="five",
                    name=self.list_five_name)

        self.source_five = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_five,
                    contributor=self.contrib_five)

        self.list_item_five = FacilityListItem \
            .objects \
            .create(row_index=0,
                    source=self.source_five,
                    status=FacilityListItem.ERROR_PARSING)

        self.list_six = FacilityList \
            .objects \
            .create(header="header",
                    file_name="six",
                    name=self.list_six_name)

        self.source_six = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_six,
                    contributor=self.contrib_six)

        self.list_item_six_a = FacilityListItem \
            .objects \
            .create(row_index=0,
                    source=self.source_six,
                    status=FacilityListItem.ERROR_PARSING)

        self.list_item_six_b = FacilityListItem \
            .objects \
            .create(row_index=1,
                    source=self.source_six,
                    status=FacilityListItem.MATCHED)

        # Test to ensure contributors don't appear in the list if all of their
        # sources have create=False
        self.source_seven = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    create=False,
                    contributor=self.contrib_seven)

        self.list_item_seven = FacilityListItem \
            .objects \
            .create(row_index=0,
                    source=self.source_seven,
                    status=FacilityListItem.MATCHED)

    def test_contributors_list_has_only_contributors_with_active_lists(self):
        response = self.client.get('/api/contributors/')
        response_data = response.json()
        contributor_names = list(zip(*response_data))[1]

        self.assertIn(
            self.contrib_one_name,
            contributor_names,
        )

        self.assertNotIn(
            self.contrib_two_name,
            contributor_names,
        )

        self.assertNotIn(
            self.contrib_three_name,
            contributor_names,
        )

        self.assertNotIn(
            self.contrib_four_name,
            contributor_names,
        )

        self.assertNotIn(
            self.contrib_five_name,
            contributor_names,
        )

        self.assertIn(
            self.contrib_six_name,
            contributor_names,
        )

        self.assertEqual(
            2,
            len(contributor_names),
        )


class FacilityListItemTests(APITestCase):
    def setUp(self):
        self.email = 'test@example.com'
        self.password = 'example123'
        self.user = User.objects.create(email=self.email)
        self.user.set_password(self.password)
        self.user.save()

        self.contributor = Contributor \
            .objects \
            .create(admin=self.user,
                    name='test contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.facility_list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='test list')

        self.source = Source.objects.create(
            source_type=Source.LIST,
            facility_list=self.facility_list,
            contributor=self.contributor)

        statuses = [c[0] for c in FacilityListItem.STATUS_CHOICES]
        for i, possible_status in enumerate(statuses):
            FacilityListItem \
                .objects \
                .create(name='test name',
                        address='test address',
                        country_code='US',
                        source=self.source,
                        row_index=i,
                        status=possible_status)

        self.client.login(email=self.email,
                          password=self.password)

    def test_get_all_items(self):
        response = self.client.get(
            reverse('facility-list-items',
                    kwargs={'pk': self.facility_list.pk}))
        self.assertEqual(200, response.status_code)
        content = json.loads(response.content)
        self.assertEqual(len(FacilityListItem.STATUS_CHOICES),
                         len(content['results']))

    def test_get_items_filtered_by_status(self):
        url = reverse('facility-list-items',
                      kwargs={'pk': self.facility_list.pk})

        response = self.client.get('{}?status=GEOCODED'.format(url))
        self.assertEqual(200, response.status_code)
        content = json.loads(response.content)
        self.assertEqual(1, len(content['results']))
        self.assertEqual('GEOCODED', content['results'][0]['status'])

    def test_get_multiple_statuses(self):
        url = reverse('facility-list-items',
                      kwargs={'pk': self.facility_list.pk})

        response = self.client.get(
            '{}?status=GEOCODED&status=MATCHED'.format(url))
        self.assertEqual(200, response.status_code)
        content = json.loads(response.content)
        self.assertEqual(2, len(content['results']))

    def test_invalid_status_returns_400(self):
        url = reverse('facility-list-items',
                      kwargs={'pk': self.facility_list.pk})

        response = self.client.get('{}?status=FOO'.format(url))
        self.assertEqual(400, response.status_code)
        content = json.loads(response.content)
        self.assertTrue('status' in content)

    def test_empty_status_filter_returns_400(self):
        url = reverse('facility-list-items',
                      kwargs={'pk': self.facility_list.pk})
        response = self.client.get('{}?status='.format(url))
        self.assertEqual(400, response.status_code)
        content = json.loads(response.content)
        self.assertTrue('status' in content)

    def test_new_facility(self):
        url = reverse('facility-list-items',
                      kwargs={'pk': self.facility_list.pk})

        # First assert that there are no NEW_FACILITYs in our test data.
        response = self.client.get(
            '{}?status=NEW_FACILITY'.format(url))
        self.assertEqual(200, response.status_code)
        content = json.loads(response.content)
        self.assertEqual(0, len(content['results']))

        # Create a facility from the test list item
        list_item = FacilityListItem.objects.filter(
            source=self.source,
            status=FacilityListItem.MATCHED,
        ).first()
        facility = Facility.objects.create(
            country_code=list_item.country_code,
            created_from=list_item,
            location=Point(0, 0),
        )
        list_item.facility = facility
        list_item.save()

        # Assert that we now have a NEW_FACILITY
        response = self.client.get(
            '{}?status=NEW_FACILITY'.format(url))
        self.assertEqual(200, response.status_code)
        content = json.loads(response.content)
        self.assertEqual(1, len(content['results']))


class FacilityClaimAdminDashboardTests(APITestCase):
    def setUp(self):
        self.email = 'test@example.com'
        self.password = 'example123'
        self.user = User.objects.create(email=self.email)
        self.user.set_password(self.password)
        self.user.save()

        self.contributor = Contributor \
            .objects \
            .create(admin=self.user,
                    name='test contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one",
                    name='List')

        self.source = Source \
            .objects \
            .create(facility_list=self.list,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        self.list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source)

        self.facility = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item)

        self.facility_match = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.CONFIRMED,
                    facility=self.facility,
                    results="",
                    facility_list_item=self.list_item)

        self.facility_claim = FacilityClaim \
            .objects \
            .create(
                contributor=self.contributor,
                facility=self.facility,
                contact_person='Name',
                email=self.email,
                phone_number=12345,
                company_name='Test',
                website='http://example.com',
                facility_description='description',
                preferred_contact_method=FacilityClaim.EMAIL)

        self.superuser = User \
            .objects \
            .create_superuser('superuser@example.com',
                              'superuser')

        self.client.login(email='superuser@example.com',
                          password='superuser')

    @override_switch('claim_a_facility', active=True)
    def test_user_cannot_submit_second_facility_claim_with_one_pending(self):
        self.client.logout()
        self.client.login(email=self.email,
                          password=self.password)

        error_response = self.client.post(
            '/api/facilities/{}/claim/'.format(self.facility.id),
            {
                'contact_person': 'contact_person',
                'email': 'example@example.com',
                'phone_number': '12345',
                'company_name': 'company_name',
                'website': 'http://example.com',
                'facility_description': 'facility_description',
                'verification_method': 'verification_method',
                'preferred_contact_method': FacilityClaim.EMAIL,
            }
        )

        self.assertEqual(400, error_response.status_code)

        self.assertEqual(
            error_response.json()['detail'],
            'User already has a pending claim on this facility'
        )

    @override_switch('claim_a_facility', active=True)
    def test_approve_claim_and_email_claimant_and_contributors(self):
        self.assertEqual(len(mail.outbox), 0)

        response = self.client.post(
            '/api/facility-claims/{}/approve/'.format(self.facility_claim.id)
        )

        self.assertEqual(200, response.status_code)

        # Expect two emails to have been sent:
        #   - one to the user who submitted the facility claim
        #   - one to a contributor who has this facility on a list
        self.assertEqual(len(mail.outbox), 2)

        updated_facility_claim = FacilityClaim \
            .objects \
            .get(pk=self.facility_claim.id)

        self.assertEqual(
            FacilityClaim.APPROVED,
            updated_facility_claim.status,
        )

        notes_count = FacilityClaimReviewNote \
            .objects \
            .filter(claim=updated_facility_claim) \
            .count()

        self.assertEqual(notes_count, 1)

        error_response = self.client.post(
            '/api/facility-claims/{}/approve/'.format(self.facility_claim.id)
        )

        self.assertEqual(400, error_response.status_code)

    @override_switch('claim_a_facility', active=True)
    def test_can_approve_at_most_one_facility_claim(self):
        response = self.client.post(
            '/api/facility-claims/{}/approve/'.format(self.facility_claim.id)
        )

        self.assertEqual(200, response.status_code)

        updated_facility_claim = FacilityClaim \
            .objects \
            .get(pk=self.facility_claim.id)

        self.assertEqual(
            FacilityClaim.APPROVED,
            updated_facility_claim.status,
        )

        new_user = User.objects.create(email='new_user@example.com')
        new_contributor = Contributor \
            .objects \
            .create(admin=new_user,
                    name='new_contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        new_facility_claim = FacilityClaim \
            .objects \
            .create(
                contributor=new_contributor,
                facility=self.facility,
                contact_person='Name',
                email='new_user@example.com',
                phone_number=12345,
                company_name='Test',
                website='http://example.com',
                facility_description='description',
                preferred_contact_method=FacilityClaim.EMAIL)

        error_response = self.client.post(
            '/api/facility-claims/{}/approve/'.format(new_facility_claim.id)
        )

        self.assertEqual(400, error_response.status_code)

    @override_switch('claim_a_facility', active=True)
    def test_deny_facility_claim(self):
        self.assertEqual(len(mail.outbox), 0)

        response = self.client.post(
            '/api/facility-claims/{}/deny/'.format(self.facility_claim.id)
        )

        self.assertEqual(200, response.status_code)
        self.assertEqual(len(mail.outbox), 1)

        updated_facility_claim = FacilityClaim \
            .objects \
            .get(pk=self.facility_claim.id)

        self.assertEqual(
            FacilityClaim.DENIED,
            updated_facility_claim.status,
        )

        notes_count = FacilityClaimReviewNote \
            .objects \
            .filter(claim=updated_facility_claim) \
            .count()

        self.assertEqual(notes_count, 1)

        error_response = self.client.post(
            '/api/facility-claims/{}/deny/'.format(self.facility_claim.id)
        )

        self.assertEqual(400, error_response.status_code)

    @override_switch('claim_a_facility', active=True)
    def test_revoke_facility_claim(self):
        self.assertEqual(len(mail.outbox), 0)

        error_response = self.client.post(
            '/api/facility-claims/{}/revoke/'.format(self.facility_claim.id)
        )

        self.assertEqual(400, error_response.status_code)

        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.save()

        response = self.client.post(
            '/api/facility-claims/{}/revoke/'.format(self.facility_claim.id)
        )

        self.assertEqual(200, response.status_code)
        self.assertEqual(len(mail.outbox), 1)

        updated_facility_claim = FacilityClaim \
            .objects \
            .get(pk=self.facility_claim.id)

        self.assertEqual(
            FacilityClaim.REVOKED,
            updated_facility_claim.status,
        )

        notes_count = FacilityClaimReviewNote \
            .objects \
            .filter(claim=updated_facility_claim) \
            .count()

        self.assertEqual(notes_count, 1)

        another_error_response = self.client.post(
            '/api/facility-claims/{}/revoke/'.format(self.facility_claim.id)
        )

        self.assertEqual(400, another_error_response.status_code)

    @override_switch('claim_a_facility', active=True)
    def test_add_claim_review_note(self):
        api_url = '/api/facility-claims/{}/note/'.format(
            self.facility_claim.id
        )
        response = self.client.post(api_url, {'note': 'note'})

        self.assertEqual(200, response.status_code)

        notes_count = FacilityClaimReviewNote \
            .objects \
            .filter(claim=self.facility_claim) \
            .count()

        self.assertEqual(notes_count, 1)

    @override_switch('claim_a_facility', active=True)
    def test_claims_list_API_accessible_only_to_superusers(self):
        response = self.client.get('/api/facility-claims/')
        self.assertEqual(200, response.status_code)

        self.client.logout()
        self.client.login(email=self.email,
                          password=self.password)

        error_response = self.client.get('/api/facility-claims/')
        self.assertEqual(403, error_response.status_code)


class ApprovedFacilityClaimTests(APITestCase):
    def setUp(self):
        self.email = 'test@example.com'
        self.password = 'example123'
        self.user = User.objects.create(email=self.email)
        self.user.set_password(self.password)
        self.user.save()

        self.user_contributor = Contributor \
            .objects \
            .create(admin=self.user,
                    name='text contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contributor_email = 'contributor@example.com'
        self.contributor_user = User.objects \
                                    .create(email=self.contributor_email)

        self.list_contributor = Contributor \
            .objects \
            .create(admin=self.contributor_user,
                    name='test contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.facility_list = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one",
                    name='List')

        self.source = Source \
            .objects \
            .create(facility_list=self.facility_list,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.list_contributor)

        self.list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source)

        self.facility = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item)

        self.facility_match = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.CONFIRMED,
                    facility=self.facility,
                    results="",
                    facility_list_item=self.list_item)

        self.facility_claim = FacilityClaim \
            .objects \
            .create(
                contributor=self.user_contributor,
                facility=self.facility,
                contact_person='Name',
                email=self.email,
                phone_number=12345,
                company_name='Test',
                website='http://example.com',
                facility_description='description',
                preferred_contact_method=FacilityClaim.EMAIL)

        self.superuser_email = 'superuser@example.com'
        self.superuser_password = 'superuser'

        self.superuser = User \
            .objects \
            .create_superuser(self.superuser_email,
                              self.superuser_password)

        self.client.login(email=self.email,
                          password=self.password)

    @override_switch('claim_a_facility', active=True)
    def test_non_approved_facility_claim_is_not_visible(self):
        api_url = '/api/facility-claims/{}/claimed/'.format(
            self.facility_claim.id
        )
        pending_response = self.client.get(api_url)
        self.assertEqual(404, pending_response.status_code)

        self.facility_claim.status = FacilityClaim.DENIED
        self.facility_claim.save()
        denied_response = self.client.get(api_url)
        self.assertEqual(404, denied_response.status_code)

        self.facility_claim.status = FacilityClaim.REVOKED
        self.facility_claim.save()
        revoked_response = self.client.get(api_url)
        self.assertEqual(404, revoked_response.status_code)

    @override_switch('claim_a_facility', active=True)
    def test_approved_facility_claim_is_visible_to_claimant(self):
        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.save()

        response = self.client.get(
            '/api/facility-claims/{}/claimed/'.format(self.facility_claim.id)
        )
        self.assertEqual(200, response.status_code)

    @override_switch('claim_a_facility', active=True)
    def test_other_user_approved_facility_claims_are_not_visible(self):
        self.client.logout()

        self.client.login(email=self.superuser_email,
                          pasword=self.superuser_password)

        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.save()

        response = self.client.get(
            '/api/facility-claims/{}/claimed/'.format(self.facility_claim.id)
        )
        self.assertEqual(401, response.status_code)

    @override_switch('claim_a_facility', active=True)
    def test_approved_facility_claim_info_is_in_details_response(self):
        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.facility_description = 'new_description'
        self.facility_claim.save()

        response_data = self.client.get(
            '/api/facilities/{}/'.format(self.facility_claim.facility.id)
        ).json()['properties']['claim_info']['facility']

        self.assertEqual(response_data['description'], 'new_description')

    @override_switch('claim_a_facility', active=True)
    def test_updating_claim_profile_sends_email_to_contributor(self):
        self.assertEqual(len(mail.outbox), 0)
        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.save()

        response = self.client.put(
            '/api/facility-claims/{}/claimed/'.format(self.facility_claim.id),
            {
                'facility_description': 'test_facility_description',
                'facility_phone_number_publicly_visible': False,
                'point_of_contact_publicly_visible': False,
                'office_info_publicly_visible': False,
                'facility_website_publicly_visible': False,
            }
        )

        self.assertEqual(200, response.status_code)

        updated_description = FacilityClaim \
            .objects \
            .get(pk=self.facility_claim.id) \
            .facility_description

        self.assertEqual(updated_description, 'test_facility_description')

        self.assertEqual(len(mail.outbox), 1)

    @override_switch('claim_a_facility', active=True)
    def test_non_visible_facility_phone_is_not_in_details_response(self):
        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.save()

        response_data = self.client.get(
            '/api/facilities/{}/'.format(self.facility_claim.facility.id)
        ).json()['properties']['claim_info']['facility']

        self.assertIsNone(response_data['phone_number'])

    @override_switch('claim_a_facility', active=True)
    def test_non_visible_office_info_is_not_in_details_response(self):
        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.save()

        response_data = self.client.get(
            '/api/facilities/{}/'.format(self.facility_claim.facility.id)
        ).json()['properties']['claim_info']

        self.assertIsNone(response_data['contact'])

    @override_switch('claim_a_facility', active=True)
    def test_non_visible_contact_info_is_not_in_details_response(self):
        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.save()

        response_data = self.client.get(
            '/api/facilities/{}/'.format(self.facility_claim.facility.id)
        ).json()['properties']['claim_info']

        self.assertIsNone(response_data['office'])

    @override_switch('claim_a_facility', active=True)
    def test_non_visible_website_is_not_in_details_response(self):
        self.facility_claim.status = FacilityClaim.APPROVED
        self.facility_claim.save()

        response_data = self.client.get(
            '/api/facilities/{}/'.format(self.facility_claim.facility.id)
        ).json()['properties']['claim_info']['facility']

        self.assertIsNone(response_data['website'])


class FacilityClaimTest(APITestCase):
    def setUp(self):
        self.email = 'test@example.com'
        self.password = 'example123'
        self.user = User.objects.create(email=self.email)
        self.user.set_password(self.password)
        self.user.save()

        self.contributor = Contributor(name='Contributor', admin=self.user)
        self.contributor.save()

        self.list_one = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='list')

        self.source_one = Source \
            .objects \
            .create(facility_list=self.list_one,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        self.list_item_one = FacilityListItem \
            .objects \
            .create(name='name',
                    address='address',
                    country_code='US',
                    source=self.source_one,
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH)

        self.facility = Facility \
            .objects \
            .create(name='name',
                    address='address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item_one)

        self.facility_match_one = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.CONFIRMED,
                    facility=self.facility,
                    results="",
                    facility_list_item=self.list_item_one)

    @override_switch('claim_a_facility', active=True)
    def test_requires_login(self):
        url = reverse('facility-claimed')
        response = self.client.get(url)
        self.assertEqual(401, response.status_code)

    @override_switch('claim_a_facility', active=True)
    def test_claimed_facilities_list(self):
        self.client.post('/user-login/',
                         {"email": self.email,
                          "password": self.password},
                         format="json")
        url = reverse('facility-claimed')
        response = self.client.get(url)
        self.assertEqual(200, response.status_code)
        data = json.loads(response.content)
        self.assertEqual([], data)

        claim = FacilityClaim(
            facility=self.facility, contributor=self.contributor)
        claim.save()

        # A new claim should NOT appear in the claimed list
        response = self.client.get(url)
        self.assertEqual(200, response.status_code)
        data = json.loads(response.content)
        self.assertEqual([], data)

        claim.status = FacilityClaim.APPROVED
        claim.save()
        response = self.client.get(url)
        self.assertEqual(200, response.status_code)
        data = json.loads(response.content)
        self.assertEqual(1, len(data))


class DashboardListTests(APITestCase):
    def setUp(self):
        self.user_email = 'test@example.com'
        self.user_password = 'example123'
        self.user = User.objects.create(email=self.user_email)
        self.user.set_password(self.user_password)
        self.user.save()

        self.contributor = Contributor \
            .objects \
            .create(admin=self.user,
                    name='test contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='First List')

        self.source = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list,
                    contributor=self.contributor)

        self.list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    source=self.source,
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH)

        self.facility = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item)

        self.inactive_list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='Second List')

        Source.objects.create(
            source_type=Source.LIST,
            facility_list=self.inactive_list,
            is_active=False,
            contributor=self.contributor)

        self.private_list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='Third List')

        Source.objects.create(
            source_type=Source.LIST,
            facility_list=self.private_list,
            is_public=False,
            contributor=self.contributor)

        self.superuser_email = 'superuser@example.com'
        self.superuser_password = 'superuser'

        self.superuser = User \
            .objects \
            .create_superuser(self.superuser_email,
                              self.superuser_password)

        self.supercontributor = Contributor \
            .objects \
            .create(admin=self.superuser,
                    name='test super contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.superlist = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='Super List')

        Source.objects.create(
            source_type=Source.LIST,
            facility_list=self.superlist,
            contributor=self.supercontributor)

    def test_user_can_list_own_lists(self):
        self.client.login(email=self.user_email,
                          password=self.user_password)
        response = self.client.get('/api/facility-lists/')

        self.assertEqual(200, response.status_code)

        lists = response.json()

        # Ensure we get all three lists
        self.assertEqual(3, len(lists))
        # Ensure they are ordered newest first
        self.assertEqual(
            ['Third List', 'Second List', 'First List'],
            [l['name'] for l in lists])

    def test_superuser_can_list_own_lists(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.get('/api/facility-lists/')

        self.assertEqual(200, response.status_code)

        lists = response.json()

        # Ensure we get the one list
        self.assertEqual(1, len(lists))
        # Ensure it is the right one
        self.assertEqual('Super List', lists[0]['name'])

    def test_superuser_can_list_other_contributors_lists(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.get(
            '/api/facility-lists/?contributor={}'.format(
                self.contributor.id))

        self.assertEqual(200, response.status_code)

        lists = response.json()

        # Ensure we get all three lists, private and public,
        # active and inactive
        self.assertEqual(3, len(lists))
        self.assertEqual(
            ['Third List', 'Second List', 'First List'],
            [l['name'] for l in lists])

    def test_user_cannot_list_other_contributors_lists(self):
        # Regular users, even if they ask for lists by other
        # contributors, will just be given their own lists

        self.client.login(email=self.user_email,
                          password=self.user_password)
        response = self.client.get(
            '/api/facility-lists/?contributor={}'.format(
                self.supercontributor.id))

        self.assertEqual(200, response.status_code)

        lists = response.json()

        # Ensure we get the user's, not the superuser's, lists
        self.assertEqual(3, len(lists))
        self.assertEqual(
            ['Third List', 'Second List', 'First List'],
            [l['name'] for l in lists])

    def test_user_can_view_own_lists(self):
        self.client.login(email=self.user_email,
                          password=self.user_password)

        for l in [self.list, self.inactive_list, self.private_list]:
            response = self.client.get('/api/facility-lists/{}/'.format(l.id))
            self.assertEqual(200, response.status_code)
            self.assertEqual(l.name, response.json()['name'])

    def test_user_cannot_view_others_lists(self):
        self.client.login(email=self.user_email,
                          password=self.user_password)

        response = self.client.get(
            '/api/facility-lists/{}/'.format(
                self.superlist.id))

        self.assertEqual(404, response.status_code)

    def test_superuser_can_view_others_lists(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        superuser_lists = [self.superlist]
        user_lists = [self.list, self.inactive_list, self.private_list]

        for l in superuser_lists + user_lists:
            response = self.client.get('/api/facility-lists/{}/'.format(l.id))
            self.assertEqual(200, response.status_code)
            self.assertEqual(l.name, response.json()['name'])


class FacilityDeleteTest(APITestCase):
    def setUp(self):
        self.user_email = 'test@example.com'
        self.user_password = 'example123'
        self.user = User.objects.create(email=self.user_email)
        self.user.set_password(self.user_password)
        self.user.save()

        self.other_user_email = 'other@example.com'
        self.other_user_password = 'other123'
        self.other_user = User.objects.create(email=self.other_user_email)
        self.other_user.set_password(self.other_user_password)
        self.other_user.save()

        self.superuser_email = 'super@example.com'
        self.superuser_password = 'example123'
        self.superuser = User.objects.create_superuser(
            email=self.superuser_email,
            password=self.superuser_password)

        self.contributor = Contributor \
            .objects \
            .create(admin=self.user,
                    name='test contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.other_contributor = Contributor \
            .objects \
            .create(admin=self.other_user,
                    name='other contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='First List')

        self.source = Source \
            .objects \
            .create(facility_list=self.list,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        self.list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    source=self.source,
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH)

        self.facility = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item)
        self.facility_url = '/api/facilities/{}/'.format(self.facility.id)

        self.list_item.facility = self.facility
        self.list_item.save()

        self.facility_match = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.CONFIRMED,
                    facility=self.facility,
                    results="",
                    facility_list_item=self.list_item)

    def test_requires_auth(self):
        response = self.client.delete(self.facility_url)
        self.assertEqual(401, response.status_code)

    def test_requires_superuser(self):
        self.client.login(email=self.user_email,
                          password=self.user_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(403, response.status_code)

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

    def test_delete(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        self.assertEqual(
            0, Facility.objects.filter(id=self.facility.id).count())
        self.assertEqual(
            0, FacilityMatch.objects.filter(facility=self.facility).count())

        self.list_item.refresh_from_db()
        self.assertEqual(
            FacilityListItem.DELETED, self.list_item.status)
        self.assertEqual(
            ProcessingAction.DELETE_FACILITY,
            self.list_item.processing_results[-1]['action'])
        self.assertEqual(
            self.facility.id,
            self.list_item.processing_results[-1]['deleted_oar_id'])

    def test_cant_delete_if_there_is_an_appoved_claim(self):
        FacilityClaim.objects.create(
            contributor=self.contributor,
            facility=self.facility,
            contact_person='test',
            email='test@test.com',
            phone_number='1234567890',
            status=FacilityClaim.APPROVED)
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(400, response.status_code)

    def test_unapproved_claims_are_deleted(self):
        FacilityClaim.objects.create(
            contributor=self.contributor,
            facility=self.facility,
            contact_person='test',
            email='test@test.com',
            phone_number='1234567890',
            status=FacilityClaim.PENDING)
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)
        self.assertEqual(
            0, FacilityClaim.objects.filter(facility=self.facility).count())

    def test_other_match_is_promoted(self):
        initial_facility_count = Facility.objects.all().count()
        list_2 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        source_2 = Source \
            .objects \
            .create(facility_list=list_2,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.other_contributor)

        list_item_2 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    geocoded_point=Point(1, 1),
                    row_index=1,
                    status=FacilityListItem.MATCHED,
                    facility=self.facility,
                    source=source_2)

        match_2 = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_2,
                    confidence=0.65,
                    results='')

        list_3 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='three',
                    name='Third List')

        source_3 = Source \
            .objects \
            .create(facility_list=list_3,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.other_contributor)

        list_item_3 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    geocoded_point=Point(2, 2),
                    source=source_3,
                    row_index=1,
                    status=FacilityListItem.MATCHED,
                    facility=self.facility)

        match_3 = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_3,
                    confidence=0.85,
                    results='')

        alias = FacilityAlias.objects.create(
            facility=self.facility,
            oar_id='US1234567ABCDEF')

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        # We should have "promoted" the matched facility to replace the deleted
        # facility
        facility_count = Facility.objects.all().count()
        self.assertEqual(facility_count, initial_facility_count)
        self.assertEqual(2, FacilityAlias.objects.all().count())

        # We should have created a new alias
        new_alias = FacilityAlias.objects.exclude(
            oar_id='US1234567ABCDEF').first()
        self.assertEqual(FacilityAlias.DELETE, new_alias.reason)
        self.assertEqual(self.facility.id, new_alias.oar_id)

        # The line item previously matched to the deleted facility should now
        # be matched to a new facility
        match_3.refresh_from_db()
        list_item_2.refresh_from_db()
        self.assertEqual(match_3.facility, list_item_2.facility)
        match_2.refresh_from_db()
        self.assertEqual(match_3.facility, match_2.facility)

        # We should have replaced the alias with one pointing to the new
        # facility
        alias.refresh_from_db()
        self.assertEqual(match_3.facility, alias.facility)

    def test_match_from_other_contributor_is_promoted(self):
        initial_facility_count = Facility.objects.all().count()
        list_2 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        source_2 = Source \
            .objects \
            .create(facility_list=list_2,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        list_item_2 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    geocoded_point=Point(1, 1),
                    row_index=1,
                    status=FacilityListItem.MATCHED,
                    facility=self.facility,
                    source=source_2)

        match_2 = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_2,
                    confidence=0.65,
                    results='')

        list_3 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='three',
                    name='Third List')

        source_3 = Source \
            .objects \
            .create(facility_list=list_3,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.other_contributor)

        list_item_3 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    geocoded_point=Point(2, 2),
                    source=source_3,
                    row_index=1,
                    status=FacilityListItem.MATCHED,
                    facility=self.facility)

        match_3 = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_3,
                    confidence=0.85,
                    results='')

        alias = FacilityAlias.objects.create(
            facility=self.facility,
            oar_id='US1234567ABCDEF')

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        # The original facility should be deleted and have no matches
        self.assertEqual(
            0, FacilityMatch.objects.filter(facility=self.facility).count())
        self.assertEqual(
            0, Facility.objects.filter(id=self.facility.id).count())

        # We should have "promoted" the other contributor's
        # matched facility to replace the deleted facility
        facility_count = Facility.objects.all().count()
        self.assertEqual(facility_count, initial_facility_count)
        self.assertEqual(2, FacilityAlias.objects.all().count())

        # match 2 should be deleted because it's from the
        # deleted facility's contributor
        list_item_2.refresh_from_db()
        self.assertEqual(
            FacilityListItem.DELETED, list_item_2.status)
        self.assertEqual(
            0, FacilityMatch.objects.filter(id=match_2.id).count())

        # We should have created a new alias
        new_alias = FacilityAlias.objects.exclude(
            oar_id='US1234567ABCDEF').first()
        self.assertEqual(FacilityAlias.DELETE, new_alias.reason)
        self.assertEqual(self.facility.id, new_alias.oar_id)

        # We should have replaced the alias with one pointing to the new
        # facility
        match_3.refresh_from_db()
        alias.refresh_from_db()
        self.assertEqual(match_3.facility, alias.facility)

    def test_matches_without_locations_are_ignored(self):
        list_2 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        source_2 = Source \
            .objects \
            .create(facility_list=list_2,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        list_item_2 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    geocoded_point=None,
                    row_index=1,
                    status=FacilityListItem.MATCHED,
                    facility=self.facility,
                    source=source_2)

        FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_2,
                    confidence=0.65,
                    results='')

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        self.assertEqual(
            0, Facility.objects.filter(id=self.facility.id).count())
        self.assertEqual(
            0, FacilityMatch.objects.filter(facility=self.facility).count())

        self.list_item.refresh_from_db()
        self.assertEqual(
            FacilityListItem.DELETED, self.list_item.status)
        self.assertEqual(
            ProcessingAction.DELETE_FACILITY,
            self.list_item.processing_results[-1]['action'])
        self.assertEqual(
            self.facility.id,
            self.list_item.processing_results[-1]['deleted_oar_id'])

    def test_other_inactive_match_is_promoted(self):
        initial_facility_count = Facility.objects.all().count()
        list_2 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        source_2 = Source \
            .objects \
            .create(facility_list=list_2,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.other_contributor)

        list_item_2 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    geocoded_point=Point(1, 1),
                    row_index=1,
                    status=FacilityListItem.MATCHED,
                    facility=self.facility,
                    source=source_2)

        FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_2,
                    confidence=0.65,
                    results='',
                    is_active=False)

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        # We should have "promoted" the matched facility to replace the deleted
        # facility
        facility_count = Facility.objects.all().count()
        self.assertEqual(facility_count, initial_facility_count)
        self.assertEqual(1, FacilityAlias.objects.all().count())
        alias = FacilityAlias.objects.first()
        self.assertEqual(list_item_2, alias.facility.created_from)

    def test_other_matches_from_same_contributor_are_deleted(self):
        initial_facility_count = Facility.objects.all().count()
        list_2 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        source_2 = Source \
            .objects \
            .create(facility_list=list_2,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        list_item_2 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    geocoded_point=Point(1, 1),
                    row_index=1,
                    status=FacilityListItem.MATCHED,
                    facility=self.facility,
                    source=source_2)

        FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_2,
                    confidence=0.65,
                    results='',
                    is_active=False)

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        # The facility should be deleted and not be replaced
        # since both items/matches came from the same contributor.
        facility_count = Facility.objects.all().count()
        self.assertEqual(facility_count, initial_facility_count - 1)
        self.assertEqual(0, FacilityAlias.objects.all().count())

    def test_rejected_match_is_deleted_not_promoted(self):
        list_2 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='Second List')

        source_2 = Source \
            .objects \
            .create(facility_list=list_2,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        list_item_2 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    geocoded_point=Point(1, 1),
                    row_index=1,
                    status=FacilityListItem.MATCHED,
                    source=source_2)

        FacilityMatch \
            .objects \
            .create(status=FacilityMatch.REJECTED,
                    facility_list_item=list_item_2,
                    confidence=0,
                    facility=self.facility,
                    results='')

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        self.assertEqual(0, Facility.objects.all().count())
        self.assertEqual(0, FacilityMatch.objects.all().count())
        self.assertEqual(0, FacilityAlias.objects.all().count())

        list_item_2.refresh_from_db()
        self.assertEqual(FacilityListItem.DELETED, list_item_2.status)
        self.assertIsNone(list_item_2.facility)
        self.assertEqual(
            ProcessingAction.DELETE_FACILITY,
            list_item_2.processing_results[-1]['action'])

    def test_delete_with_alias(self):
        FacilityAlias.objects.create(facility=self.facility,
                                     oar_id='US1234567ABCDEF')
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

    def test_delete_removed_item(self):
        self.facility_match.is_active = False
        self.facility_match.save()

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        self.assertEqual(
            0, Facility.objects.filter(id=self.facility.id).count())
        self.assertEqual(
            0, FacilityMatch.objects.filter(facility=self.facility).count())

    def test_can_delete_multiple_created_froms(self):
        FacilityMatch \
            .objects \
            .create(status=FacilityMatch.PENDING,
                    facility=self.facility,
                    facility_list_item=self.facility.created_from,
                    confidence=0.85,
                    results='')

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.delete(self.facility_url)
        self.assertEqual(204, response.status_code)

        self.assertEqual(
            0, Facility.objects.filter(id=self.facility.id).count())
        self.assertEqual(
            0, FacilityMatch.objects.filter(facility=self.facility).count())


class FacilityMergeTest(APITestCase):
    def setUp(self):
        self.user_email = 'test@example.com'
        self.user_password = 'example123'
        self.user = User.objects.create(email=self.user_email)
        self.user.set_password(self.user_password)
        self.user.save()

        self.superuser_email = 'super@example.com'
        self.superuser_password = 'example123'
        self.superuser = User.objects.create_superuser(
            email=self.superuser_email,
            password=self.superuser_password)

        self.api_user_email = 'api@example.com'
        self.api_user_password = 'example123'
        self.api_user = User.objects.create(email=self.api_user_email)
        self.api_user.set_password(self.api_user_password)
        self.api_user.save()

        self.contributor_1 = Contributor \
            .objects \
            .create(admin=self.user,
                    name='test contributor 1',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_1 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='First List')

        self.source_1 = Source \
            .objects \
            .create(facility_list=self.list_1,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor_1)

        self.list_item_1 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_1)

        self.facility_1 = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item_1)

        self.match_1 = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_1,
                    facility_list_item=self.list_item_1,
                    confidence=0.85,
                    results='')

        self.list_item_1.facility = self.facility_1
        self.list_item_1.save()

        self.facility_1_claim = FacilityClaim \
            .objects \
            .create(
                contributor=self.contributor_1,
                facility=self.facility_1,
                contact_person='test 1',
                job_title='test 1',
                email='test1@test.com',
                phone_number='1234567890',
                company_name='test 1',
                facility_description='test 1',
                preferred_contact_method=FacilityClaim.EMAIL,
                status=FacilityClaim.APPROVED)

        self.contributor_2 = Contributor \
            .objects \
            .create(admin=self.superuser,
                    name='test contributor 2',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_2 = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        self.source_2 = Source \
            .objects \
            .create(facility_list=self.list_2,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor_2)

        self.list_item_2 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    ppe_product_types=['two_a', 'two_b'],
                    ppe_contact_phone='222-222-2222',
                    ppe_contact_email='ppe_two@example.com',
                    ppe_website='https://example.com/ppe_two',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_2)

        self.facility_2 = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    ppe_product_types=self.list_item_2.ppe_product_types,
                    ppe_contact_phone=self.list_item_2.ppe_contact_phone,
                    ppe_contact_email=self.list_item_2.ppe_contact_email,
                    ppe_website=self.list_item_2.ppe_website,
                    location=Point(0, 0),
                    created_from=self.list_item_2)

        self.match_2 = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_2,
                    facility_list_item=self.list_item_2,
                    confidence=0.85,
                    results='')

        self.list_item_2.facility = self.facility_2
        self.list_item_2.save()

        self.facility_2_claim = FacilityClaim \
            .objects \
            .create(
                contributor=self.contributor_2,
                facility=self.facility_2,
                contact_person='test 2',
                job_title='test 2',
                email='test2@test.com',
                phone_number='1234567890',
                company_name='test 2',
                facility_description='test 2',
                preferred_contact_method=FacilityClaim.EMAIL)

        self.contributor_3 = Contributor \
            .objects \
            .create(admin=self.api_user,
                    name='test contributor 3',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.source_3 = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True,
                    create=False,
                    contributor=self.contributor_3)

        self.list_item_3 = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    facility=self.facility_2,
                    source=self.source_3)

        self.extended_field_1 = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value='name one',
                contributor=self.contributor_1,
                facility=self.facility_1,
                facility_list_item=self.list_item_1,
                is_verified=True
            )

        self.extended_field_2 = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value='name two',
                contributor=self.contributor_2,
                facility=self.facility_2,
                facility_list_item=self.list_item_2
            )

        self.existing_alias = FacilityAlias.objects.create(
            facility=self.facility_2,
            oar_id='US1234567ABCDEF')

        self.merge_endpoint = '/api/facilities/merge/'
        self.merge_url = '{}?target={}&merge={}'.format(
            self.merge_endpoint,
            self.facility_1.id,
            self.facility_2.id)

    def test_requires_auth(self):
        response = self.client.post(self.merge_url)
        self.assertEqual(401, response.status_code)

    def test_requires_superuser(self):
        self.client.login(email=self.user_email,
                          password=self.user_password)
        response = self.client.post(self.merge_url)
        self.assertEqual(403, response.status_code)

    def test_merge(self):
        original_facility_count = Facility.objects.all().count()
        original_alias_count = FacilityAlias.objects.all().count()
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.post(self.merge_url)
        self.assertEqual(200, response.status_code)
        data = json.loads(response.content)
        self.assertEqual(self.facility_1.id, data['id'])
        self.assertEqual(original_facility_count - 1,
                         Facility.objects.all().count())

        # The target models should not have changed
        self.list_item_1.refresh_from_db()
        self.assertEqual(self.list_item_1.facility, self.facility_1)
        self.match_1.refresh_from_db()
        self.assertEqual(self.match_1.facility, self.facility_1)
        self.assertEqual(self.match_1.facility_list_item, self.list_item_1)

        self.list_item_2.refresh_from_db()
        self.assertEqual(self.list_item_2.facility, self.facility_1)
        self.match_2.refresh_from_db()
        self.assertEqual(self.match_2.facility, self.facility_1)
        self.assertEqual(self.match_2.facility_list_item, self.list_item_2)
        self.assertEqual(ProcessingAction.MERGE_FACILITY,
                         self.list_item_2.processing_results[-1]['action'])

        self.facility_1.refresh_from_db()
        self.assertIn(self.source_1, self.facility_1.sources())
        self.assertIn(self.source_2, self.facility_1.sources())

        self.assertEqual(original_alias_count + 1,
                         FacilityAlias.objects.all().count())
        alias = FacilityAlias.objects.first()
        for alias in FacilityAlias.objects.all():
            self.assertEqual(self.facility_1, alias.facility)
            self.assertIn(alias.oar_id,
                          (self.facility_2.id, self.existing_alias.oar_id))
            self.assertEqual(FacilityAlias.MERGE, alias.reason)

        # The PPE fields should have been copied
        self.assertEqual(self.list_item_2.ppe_product_types,
                         self.facility_1.ppe_product_types)
        self.assertEqual(self.list_item_2.ppe_contact_phone,
                         self.facility_1.ppe_contact_phone)
        self.assertEqual(self.list_item_2.ppe_contact_email,
                         self.facility_1.ppe_contact_email)
        self.assertEqual(self.list_item_2.ppe_website,
                         self.facility_1.ppe_website)

        self.facility_2_claim.refresh_from_db()
        # The pending claim on the merge facility should have been updated
        self.assertEqual(self.facility_1, self.facility_2_claim.facility)
        self.assertEqual(FacilityClaim.DENIED, self.facility_2_claim.status)

    def test_merge_with_extended_fields(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.post(self.merge_url)
        self.assertEqual(200, response.status_code)

        # Fields pointing to target facility should be unchanged
        self.extended_field_1.refresh_from_db()
        self.assertEqual(self.extended_field_1.facility, self.facility_1)
        self.assertEqual(self.extended_field_1.facility_list_item,
                         self.list_item_1)

        # Fields pointing to merge facility should be updated
        self.extended_field_2.refresh_from_db()
        self.assertEqual(self.extended_field_2.facility, self.facility_1)
        self.assertEqual(self.extended_field_2.facility_list_item,
                         self.list_item_2)

    def test_updates_facility_index(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.post(self.merge_url)
        self.assertEqual(200, response.status_code)

        facility_index = FacilityIndex.objects.get(id=self.facility_1.id)
        self.assertIn(self.extended_field_1.value,
                      facility_index.native_language_name)
        self.assertIn(self.extended_field_2.value,
                      facility_index.native_language_name)

    def test_merge_with_two_approved_claims(self):
        self.facility_1_claim.status = FacilityClaim.APPROVED
        self.facility_1_claim.save()
        self.facility_2_claim.status = FacilityClaim.APPROVED
        self.facility_2_claim.save()

        just_before_change = datetime.utcnow()
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.post(self.merge_url)
        self.assertEqual(200, response.status_code)

        self.facility_1_claim.refresh_from_db()
        self.facility_2_claim.refresh_from_db()

        self.assertEqual(self.facility_1, self.facility_2_claim.facility)
        self.assertEqual(FacilityClaim.APPROVED, self.facility_1_claim.status)
        self.assertEqual(FacilityClaim.REVOKED, self.facility_2_claim.status)
        self.assertEqual(self.superuser,
                         self.facility_2_claim.status_change_by)
        # Modifying tzinfo avoids
        # TypeError: can't compare offset-naive and offset-aware datetimes
        self.assertGreater(
            self.facility_2_claim.status_change_date.replace(tzinfo=None),
            just_before_change.replace(tzinfo=None))

    def test_required_params(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.post(self.merge_endpoint)
        self.assertEqual(400, response.status_code)
        data = json.loads(response.content)
        self.assertIn('target', data)
        self.assertIn('merge', data)

    def test_params_reference_existing_objects(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        url = '{}?target={}&merge={}'.format(self.merge_endpoint, 'foo', 'bar')
        response = self.client.post(url)
        self.assertEqual(400, response.status_code)
        data = json.loads(response.content)
        self.assertIn('target', data)
        self.assertIn('merge', data)

    def test_requires_distinct_params(self):
        self.client.login(email=self.user_email,
                          password=self.user_password)
        response = self.client.post(self.merge_url)
        self.assertEqual(403, response.status_code)

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        url = '{}?target={}&merge={}'.format(self.merge_endpoint,
                                             self.facility_1.id,
                                             self.facility_1.id)
        response = self.client.post(url)
        self.assertEqual(400, response.status_code)
        data = json.loads(response.content)
        self.assertIn('target', data)
        self.assertIn('merge', data)


class FacilitySplitTest(APITestCase):
    def setUp(self):
        self.user_one_email = 'one@example.com'
        self.user_one_password = 'example123'
        self.user_one = User.objects.create(email=self.user_one_email)
        self.user_one.set_password(self.user_one_password)
        self.user_one.save()

        self.user_two_email = 'two@example.com'
        self.user_two_password = 'example123'
        self.user_two = User.objects.create(email=self.user_two_email)
        self.user_two.set_password(self.user_two_password)
        self.user_two.save()

        self.superuser_email = 'super@example.com'
        self.superuser_password = 'example123'
        self.superuser = User.objects.create_superuser(
            email=self.superuser_email,
            password=self.superuser_password)

        self.contributor_one = Contributor \
            .objects \
            .create(admin=self.user_one,
                    name='test contributor 1',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_one = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='First List')

        self.source_one = Source \
            .objects \
            .create(facility_list=self.list_one,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor_one)

        self.list_item_one = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_one)

        self.facility_one = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    ppe_product_types=['two_a', 'two_b'],
                    ppe_contact_phone='222-222-2222',
                    ppe_contact_email='ppe_two@example.com',
                    ppe_website='https://example.com/ppe_two',
                    created_from=self.list_item_one)

        self.match_one = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_one,
                    facility_list_item=self.list_item_one,
                    confidence=0.85,
                    results='')

        self.list_item_one.facility = self.facility_one
        self.list_item_one.save()

        self.contributor_two = Contributor \
            .objects \
            .create(admin=self.user_two,
                    name='test contributor 2',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_two = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        self.source_two = Source \
            .objects \
            .create(facility_list=self.list_two,
                    source_type=Source.LIST,
                    contributor=self.contributor_two)

        self.list_item_two = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    ppe_product_types=['two_a', 'two_b'],
                    ppe_contact_phone='222-222-2222',
                    ppe_contact_email='ppe_two@example.com',
                    ppe_website='https://example.com/ppe_two',
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_two)

        self.match_two = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_one,
                    facility_list_item=self.list_item_two,
                    confidence=0.85,
                    results='')

        self.list_item_two.facility = self.facility_one
        self.list_item_two.save()

        self.extended_field_one = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value='name one',
                contributor=self.contributor_one,
                facility=self.facility_one,
                facility_list_item=self.list_item_one,
                is_verified=True
            )

        self.extended_field_two = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value='name two',
                contributor=self.contributor_two,
                facility=self.facility_one,
                facility_list_item=self.list_item_two
            )

        self.split_url = '/api/facilities/{}/split/'.format(
            self.facility_one.id,
        )

    def test_split_is_unauthorized_for_anonymous_users(self):
        get_response = self.client.get(self.split_url)
        self.assertEqual(get_response.status_code, 401)

        post_response = self.client.post(self.split_url)
        self.assertEqual(post_response.status_code, 401)

    def test_split_is_unauthorized_for_non_administrators(self):
        self.client.login(email=self.user_one_email,
                          password=self.user_one_password)
        get_response = self.client.get(self.split_url)
        self.assertEqual(get_response.status_code, 403)

        post_response = self.client.post(self.split_url)
        self.assertEqual(post_response.status_code, 403)

    def test_get_returns_facility_details_with_match_data(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        get_response = self.client.get(self.split_url)
        self.assertEqual(get_response.status_code, 200)

        data = json.loads(get_response.content)
        self.assertEqual(
            len(data['properties']['matches']),
            1,
        )

        self.assertEqual(
            data['properties']['matches'][0]['match_id'],
            self.match_two.id,
        )

    def test_post_returns_bad_request_without_match_id(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.split_url, {})
        self.assertEqual(post_response.status_code, 400)

    def test_post_reverts_to_created_facility(self):
        self.facility_two = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    ppe_product_types=['two_a', 'two_b'],
                    ppe_contact_phone='222-222-2222',
                    ppe_contact_email='ppe_two@example.com',
                    ppe_website='https://example.com/ppe_two',
                    created_from=self.list_item_two)

        initial_facility_count = Facility \
            .objects \
            .all() \
            .count()
        self.assertEqual(initial_facility_count, 2)
        self.assertEqual(self.match_two.facility, self.facility_one)

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.split_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 200)

        updated_facility_count = Facility \
            .objects \
            .all() \
            .count()

        self.assertEqual(updated_facility_count, initial_facility_count)

        self.match_two.refresh_from_db()
        self.assertEqual(self.match_two.facility, self.facility_two)

    def test_post_creates_a_new_facility(self):
        initial_facility_count = Facility \
            .objects \
            .all() \
            .count()
        self.assertEqual(initial_facility_count, 1)

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.split_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 200)

        updated_facility_count = Facility \
            .objects \
            .all() \
            .count()

        self.assertEqual(updated_facility_count, 2)

    def test_post_decrements_prior_facility_matches(self):
        self.assertEqual(
            self.facility_one.facilitymatch_set.count(),
            2,
        )

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.split_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 200)

        self.facility_one.refresh_from_db()

        self.assertEqual(
            self.facility_one.facilitymatch_set.count(),
            1,
        )

    def test_post_returns_match_id_and_new_oar_id(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.split_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 200)

        data = json.loads(post_response.content)

        self.assertEqual(
            self.match_two.id,
            data['match_id'],
        )

        self.match_two.refresh_from_db()

        self.assertEqual(
            self.match_two.facility.id,
            data['new_oar_id'],
        )

    def test_post_reverts_ppe_data(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.split_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 200)

        self.facility_one.refresh_from_db()
        self.assertEqual(self.facility_one.created_from.ppe_product_types,
                         self.facility_one.ppe_product_types)
        self.assertEqual(self.facility_one.created_from.ppe_contact_phone,
                         self.facility_one.ppe_contact_phone)
        self.assertEqual(self.facility_one.created_from.ppe_contact_email,
                         self.facility_one.ppe_contact_email)
        self.assertEqual(self.facility_one.created_from.ppe_website,
                         self.facility_one.ppe_website)

    def test_post_updates_extended_fields(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.split_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 200)
        data = json.loads(post_response.content)
        new_facility_id = data['new_oar_id']

        # Fields not associated with the split-off match should be unchanged
        self.extended_field_one.refresh_from_db()
        self.assertEqual(self.extended_field_one.facility, self.facility_one)
        self.assertEqual(self.extended_field_one.facility_list_item,
                         self.list_item_one)

        # Field associated with the split-off match should be updated
        self.extended_field_two.refresh_from_db()
        self.assertEqual(self.extended_field_two.facility.id, new_facility_id)
        self.assertEqual(self.extended_field_two.facility_list_item,
                         self.list_item_two)

    def test_post_updates_facility_index(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.split_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 200)
        data = json.loads(post_response.content)
        new_facility_id = data['new_oar_id']

        facility_index_one = FacilityIndex.objects.get(id=self.facility_one.id)
        self.assertIn(self.extended_field_one.value,
                      facility_index_one.native_language_name)
        self.assertNotIn(self.extended_field_two.value,
                         facility_index_one.native_language_name)
        facility_index_two = FacilityIndex.objects.get(id=new_facility_id)
        self.assertNotIn(self.extended_field_one.value,
                         facility_index_two.native_language_name)
        self.assertIn(self.extended_field_two.value,
                      facility_index_two.native_language_name)


class FacilityMatchPromoteTest(APITestCase):
    def setUp(self):
        self.user_one_email = 'one@example.com'
        self.user_one_password = 'example123'
        self.user_one = User.objects.create(email=self.user_one_email)
        self.user_one.set_password(self.user_one_password)
        self.user_one.save()

        self.user_two_email = 'two@example.com'
        self.user_two_password = 'example123'
        self.user_two = User.objects.create(email=self.user_two_email)
        self.user_two.set_password(self.user_two_password)
        self.user_two.save()

        self.superuser_email = 'super@example.com'
        self.superuser_password = 'example123'
        self.superuser = User.objects.create_superuser(
            email=self.superuser_email,
            password=self.superuser_password)

        self.contributor_one = Contributor \
            .objects \
            .create(admin=self.user_one,
                    name='test contributor 1',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_one = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='First List')

        self.source_one = Source \
            .objects \
            .create(facility_list=self.list_one,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor_one)

        self.address_one = 'Address One'
        self.address_two = 'Address Two'
        self.name_one = 'Name One'
        self.name_two = 'Name Two'
        self.country_code_one = 'US'
        self.country_code_two = 'CA'
        self.location_one = Point(1, 1)
        self.location_two = Point(2, 2)

        self.ppe_product_types_two = ['Masks', 'Gloves']
        self.ppe_contact_phone_two = '123-456-7890'
        self.ppe_contact_email_two = 'ppe@example.com'
        self.ppe_website_two = 'https://example.com/ppe'

        self.list_item_one = FacilityListItem \
            .objects \
            .create(name=self.name_one,
                    address=self.name_one,
                    country_code=self.country_code_one,
                    row_index=1,
                    geocoded_point=self.location_one,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_one)

        self.facility_one = Facility \
            .objects \
            .create(name=self.name_one,
                    address=self.address_one,
                    country_code=self.country_code_one,
                    location=self.location_one,
                    created_from=self.list_item_one)

        self.match_one = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_one,
                    facility_list_item=self.list_item_one,
                    confidence=0.85,
                    results='')

        self.list_item_one.facility = self.facility_one
        self.list_item_one.save()

        self.contributor_two = Contributor \
            .objects \
            .create(admin=self.user_two,
                    name='test contributor 2',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_two = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        self.source_two = Source \
            .objects \
            .create(facility_list=self.list_two,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor_two)

        self.list_item_two = FacilityListItem \
            .objects \
            .create(name=self.name_two,
                    address=self.address_two,
                    country_code=self.country_code_two,
                    row_index=1,
                    geocoded_point=self.location_two,
                    ppe_product_types=self.ppe_product_types_two,
                    ppe_contact_phone=self.ppe_contact_phone_two,
                    ppe_contact_email=self.ppe_contact_email_two,
                    ppe_website=self.ppe_website_two,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_two)

        self.match_two = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_one,
                    facility_list_item=self.list_item_two,
                    confidence=0.85,
                    results='')

        self.list_item_three = FacilityListItem \
            .objects \
            .create(name='third facility',
                    address='third address',
                    country_code='US',
                    source=self.source_one,
                    row_index=2,
                    geocoded_point=Point(3, 3),
                    status=FacilityListItem.MATCHED)

        self.other_facility = Facility \
            .objects \
            .create(name='third facility',
                    address='third address',
                    country_code='US',
                    location=Point(3, 3),
                    created_from=self.list_item_three)

        self.list_item_three.facility = self.other_facility
        self.list_item_three.save()

        self.list_item_two.facility = self.facility_one
        self.list_item_two.save()

        self.promote_url = '/api/facilities/{}/promote/'.format(
            self.facility_one.id,
        )

    def test_promote_is_unauthorized_for_anon_users(self):
        post_response = self.client.post(self.promote_url)
        self.assertEqual(post_response.status_code, 401)

    def test_promote_is_unauth_for_non_admins(self):
        self.client.login(email=self.user_one_email,
                          password=self.user_one_password)
        post_response = self.client.post(self.promote_url)
        self.assertEqual(post_response.status_code, 403)

    def test_returns_error_if_list_item_not_in_matched_status(self):
        self.list_item_two.status = FacilityListItem.PARSED
        self.list_item_two.save()

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.promote_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 400)

    def test_returns_error_if_match_is_not_to_facility(self):
        self.match_two.facility = self.other_facility
        self.match_two.facility_list_item = self.list_item_three
        self.match_two.save()
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        post_response = self.client.post(self.promote_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 400)

    def test_returns_error_if_facility_created_from_match_list_item(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        post_response = self.client.post(self.promote_url,
                                         {'match_id': self.match_one.id})
        self.assertEqual(post_response.status_code, 400)

    def test_updates_facility_if_list_item_is_in_matched_status(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.promote_url,
                                         {'match_id': self.match_two.id})
        self.assertEqual(post_response.status_code, 200)

        self.facility_one.refresh_from_db()

        self.assertEqual(
            self.facility_one.name,
            self.list_item_two.name,
        )

        self.assertEqual(
            self.facility_one.address,
            self.list_item_two.address,
        )

        self.assertEqual(
            self.facility_one.country_code,
            self.list_item_two.country_code,
        )

        self.assertEqual(
            self.facility_one.location,
            self.list_item_two.geocoded_point,
        )

        reason = 'Promoted item {} in list {} over item {} in list {}'.format(
            self.list_item_two.id,
            self.list_two.id,
            self.list_item_one.id,
            self.list_one.id,
        )

        self.assertEqual(
            Facility.history.first().history_change_reason,
            reason,
        )

        self.assertEqual(
            self.facility_one.ppe_product_types,
            self.list_item_two.ppe_product_types,
        )

        self.assertEqual(
            self.facility_one.ppe_contact_phone,
            self.list_item_two.ppe_contact_phone,
        )

        self.assertEqual(
            self.facility_one.ppe_contact_email,
            self.list_item_two.ppe_contact_email,
        )

        self.assertEqual(
            self.facility_one.ppe_website,
            self.list_item_two.ppe_website,
        )

    def test_can_promote_single_item_over_list_item(self):
        single_source = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor_two)

        single_item = FacilityListItem \
            .objects \
            .create(name='single',
                    address='single',
                    country_code='US',
                    row_index=0,
                    geocoded_point=self.location_one,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=single_source)

        single_match = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_one,
                    facility_list_item=single_item,
                    confidence=0.85,
                    results='')

        single_item.facility = self.facility_one
        single_item.save()

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        post_response = self.client.post(self.promote_url,
                                         {'match_id': single_match.id})
        self.assertEqual(post_response.status_code, 200)


class PermissionsTests(TestCase):
    class MockRequest(object):
        def __init__(self, referer):
            self.META = {}
            if referer is not None:
                self.META['HTTP_REFERER'] = referer

    @override_settings(ALLOWED_HOSTS=['.allowed.org'])
    def test_is_referer_allowed(self):
        def check_host(url):
            return referring_host_is_allowed(
                referring_host(
                    PermissionsTests.MockRequest(url)))

        self.assertTrue(check_host('http://allowed.org'))
        self.assertTrue(check_host('http://subdomain.allowed.org'))
        self.assertTrue(check_host('http://allowed.org:6543'))
        self.assertTrue(check_host('http://allowed.org:6543/api/countries'))

        self.assertFalse(check_host('http://notallowed.org'))
        self.assertFalse(check_host('http://allowed.com'))
        self.assertFalse(check_host(''))
        self.assertFalse(check_host(None))
        self.assertFalse(check_host('foo'))


class FacilityClaimChangesTest(TestCase):
    def setUp(self):
        self.email = 'test@example.com'
        self.password = 'password'
        self.name = 'Test User'
        self.user = User(email=self.email)
        self.user.set_password(self.password)
        self.user.save()

        self.contributor = Contributor.objects.create(
            name=self.name, admin=self.user)

        self.facility_list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='one')

        self.source = Source \
            .objects \
            .create(facility_list=self.facility_list,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        self.list_item = FacilityListItem \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source)

        self.facility = Facility \
            .objects \
            .create(name=self.list_item.name,
                    address=self.list_item.address,
                    country_code=self.list_item.country_code,
                    location=Point(0, 0),
                    created_from=self.list_item)

        self.claim = FacilityClaim.objects.create(
            contributor=self.contributor,
            facility=self.facility,
            facility_name_english='Facility Name',
            facility_name_native_language='Објекат Име'
        )

    def test_no_changes(self):
        self.assertIsNone(self.claim.get_changes())

    def test_changes(self):
        prev_name = self.claim.facility_name_english
        new_name = 'Changed'
        self.claim.facility_name_english = new_name
        self.claim.save()
        changes = self.claim.get_changes()
        self.assertIsNotNone(changes)
        self.assertEqual(1, len(changes))
        self.assertEqual('facility_name_english', changes[0]['name'])
        self.assertEqual(prev_name, changes[0]['previous'])
        self.assertEqual(new_name, changes[0]['current'])
        self.assertEqual(
            'facility name in English', changes[0]['verbose_name'])

    def test_do_not_include_field_change(self):
        self.claim.facility_name = 'Changed'
        self.claim.save()
        changes = self.claim.get_changes(include='foo')
        self.assertIsNone(changes)

    def test_non_public_changes(self):
        self.claim.facility_phone_number = 'Changed'
        self.claim.point_of_contact_person_name = 'Changed'
        self.claim.office_official_name = 'Changed'
        self.claim.office_address = 'Changed'
        self.claim.office_country_code = 'CN'
        self.claim.office_phone_number = 'Changed'
        self.claim.save()
        self.assertIsNone(self.claim.get_changes())

        self.claim.office_info_publicly_visible = True
        self.claim.office_official_name = 'Changed again'
        self.claim.office_address = 'Changed again'
        self.claim.office_country_code = 'GB'
        self.claim.office_phone_number = 'Changed again'
        self.claim.save()
        changes = self.claim.get_changes()
        self.assertIsNotNone(changes)
        field_names = [c['name'] for c in changes]
        self.assertIn('office_official_name', field_names)
        self.assertIn('office_address', field_names)
        self.assertIn('office_country_code', field_names)
        self.assertIn('office_phone_number', field_names)
        self.assertNotIn('facility_phone_number', field_names)
        self.assertNotIn('point_of_contact_person_name', field_names)

    def test_change_serializers(self):
        self.claim.parent_company = self.contributor
        self.claim.save()
        changes = self.claim.get_changes()
        self.assertIsNotNone(changes)
        self.assertEqual(1, len(changes))
        change = changes[0]
        self.assertEqual(change['name'], 'parent_company')
        self.assertEqual(change['current'], self.name)


class FacilityClaimSerializerTests(TestCase):
    def setUp(self):
        self.email = 'test@example.com'
        self.password = 'password'
        self.name = 'Test User'
        self.user = User(email=self.email)
        self.user.set_password(self.password)
        self.user.save()

        self.contributor = Contributor.objects.create(
            name=self.name, admin=self.user)

        self.facility_list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='one')

        self.source = Source \
            .objects \
            .create(facility_list=self.facility_list,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        self.list_item = FacilityListItem \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source)

        self.facility = Facility \
            .objects \
            .create(name=self.list_item.name,
                    address=self.list_item.address,
                    country_code=self.list_item.country_code,
                    location=Point(0, 0),
                    created_from=self.list_item)

        self.claim = FacilityClaim.objects.create(
            contributor=self.contributor,
            facility=self.facility,
            status=FacilityClaim.APPROVED,
        )

    def test_product_and_production_options_are_serialized(self):
        data = ApprovedFacilityClaimSerializer(self.claim).data

        self.assertIn('production_type_choices', data)
        self.assertIsNotNone(data['production_type_choices'])
        self.assertNotEqual([], data['production_type_choices'])


class TilePermissionsTest(APITestCase):
    def setUp(self):
        self.tile_path = reverse('tile', kwargs={
            'layer': 'facilitygrid',
            'cachekey': '1567700347-1-95f951f7',
            'z': 6, 'x': 15, 'y': 29,
            'ext': 'pbf',
        })

    @override_settings(ALLOWED_HOSTS=['testserver', '.allowed.org'])
    @override_switch('vector_tile', active=True)
    def test_allowed_hosts_can_fetch_tiles(self):
        response = self.client.get(self.tile_path, {},
                                   HTTP_REFERER='http://allowed.org/')
        self.assertEqual(200, response.status_code)

    def test_disallowed_hosts_cannot_fetch_tiles(self):
        response = self.client.get(self.tile_path)
        self.assertEqual(401, response.status_code)


class FacilityAPITestCaseBase(APITestCase):
    def setUp(self):
        self.user_email = 'test@example.com'
        self.user_password = 'example123'
        self.user = User.objects.create(email=self.user_email)
        self.user.set_password(self.user_password)
        self.user.save()

        self.superuser_email = 'super@example.com'
        self.superuser_password = 'example123'
        self.superuser = User.objects.create_superuser(
            email=self.superuser_email,
            password=self.superuser_password)

        self.contributor = Contributor \
            .objects \
            .create(admin=self.user,
                    name='test contributor 1',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='First List')

        self.source = Source \
            .objects \
            .create(facility_list=self.list,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        self.list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source)

        self.facility = Facility \
            .objects \
            .create(name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item)

        self.match = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=self.list_item,
                    confidence=0.85,
                    results='')

        self.list_item.facility = self.facility
        self.list_item.save()

    def join_group_and_login(self):
        self.client.logout()
        group = auth.models.Group.objects.get(
            name=FeatureGroups.CAN_SUBMIT_FACILITY,
        )
        self.user.groups.set([group.id])
        self.user.save()
        self.client.login(email=self.user_email,
                          password=self.user_password)


class SearchByList(APITestCase):
    def setUp(self):
        self.user_email = 'test@example.com'
        self.user_password = 'example123'
        self.user = User.objects.create(email=self.user_email)
        self.user.set_password(self.user_password)
        self.user.save()

        self.superuser_email = 'super@example.com'
        self.superuser_password = 'example123'
        self.superuser = User.objects.create_superuser(
            email=self.superuser_email,
            password=self.superuser_password)

        self.contributor = Contributor \
            .objects \
            .create(id=111,
                    admin=self.user,
                    name='test contributor 1',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list = FacilityList \
            .objects \
            .create(id=123,
                    header='header',
                    file_name='one',
                    name='First List')

        self.source = Source \
            .objects \
            .create(id=456,
                    facility_list=self.list,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        self.list_item = FacilityListItem \
            .objects \
            .create(id=789,
                    name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source)

        self.facility = Facility \
            .objects \
            .create(id='US2021067MMRTSD',
                    name='Name',
                    address='Address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item)

        self.match = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=self.list_item,
                    confidence=0.85,
                    results='')

        self.list_item.facility = self.facility
        self.list_item.save()

    def test_fetched_by_contributor(self):
        url = '/api/facilities/?contributors={}'.format(self.contributor.id)
        response = self.client.get(url)
        self.assertEqual(200, response.status_code)
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], self.facility.id)

    def test_fetched_by_list(self):
        url = '/api/facilities/?lists={}'.format(self.list.id)
        response = self.client.get(url)
        self.assertEqual(200, response.status_code)
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], self.facility.id)

    def test_fetched_by_list_and_contributor(self):
        url = '/api/facilities/?contributors={}&lists={}'.format(
            self.contributor.id, self.list.id)
        response = self.client.get(url)
        self.assertEqual(200, response.status_code)
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], self.facility.id)


class UpdateLocationTest(FacilityAPITestCaseBase):
    def setUp(self):
        super(UpdateLocationTest, self).setUp()
        self.url = reverse('facility-update-location',
                           kwargs={'pk': self.facility.id})

    def test_requires_auth(self):
        response = self.client.post(self.url)
        self.assertEqual(401, response.status_code)

    def test_requires_superuser(self):
        self.client.login(email=self.user_email,
                          password=self.user_password)
        response = self.client.post(self.url)
        self.assertEqual(403, response.status_code)

    def test_facility_exisits(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        self.url = reverse('facility-update-location',
                           kwargs={'pk': 'DOES_NOT_EXIST'})
        response = self.client.post(self.url)
        self.assertEqual(404, response.status_code)

    def test_required_arguments(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.post(self.url)
        self.assertEqual(400, response.status_code)
        data = json.loads(response.content)
        self.assertIn('lat', data)
        self.assertIn('lng', data)

    def test_valid_arguments(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        response = self.client.post(self.url, {'lat': 1000, 'lng': 0})
        self.assertEqual(400, response.status_code)
        data = json.loads(response.content)
        self.assertIn('lat', data)
        self.assertNotIn('lng', data)

        response = self.client.post(self.url, {'lat': 0, 'lng': 1000})
        self.assertEqual(400, response.status_code)
        data = json.loads(response.content)
        self.assertIn('lng', data)
        self.assertNotIn('lat', data)

    def test_updates_location(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        response = self.client.post(
            self.url, {
                UpdateLocationParams.LAT: 41,
                UpdateLocationParams.LNG: 43,
                UpdateLocationParams.NOTES: 'A note',
                UpdateLocationParams.CONTRIBUTOR_ID: self.contributor.id,
            })
        self.assertEqual(200, response.status_code)
        data = json.loads(response.content)
        self.assertEqual(data['id'], self.facility.id)
        self.assertEqual(data['geometry']['coordinates'], [43.0, 41.0])

        facility_locations = FacilityLocation.objects.filter(
            facility=self.facility)
        self.assertTrue(facility_locations.count() == 1)
        facility_location = facility_locations.first()
        self.assertEqual(facility_location.facility, self.facility)
        self.assertEqual(facility_location.created_by, self.superuser)
        self.assertEqual(facility_location.location.x, 43.0)
        self.assertEqual(facility_location.location.y, 41.0)
        self.assertEqual(facility_location.contributor, self.contributor)
        self.assertEqual(facility_location.notes, 'A note')


class SerializeOtherLocationsTest(FacilityAPITestCaseBase):
    def setUp(self):
        super(SerializeOtherLocationsTest, self).setUp()
        self.url = '/api/facilities/{}'.format(self.facility.id)

        self.other_user_email = 'hello@example.com'
        self.other_user_password = 'example123'
        self.other_user = User.objects.create(email=self.other_user_email)
        self.other_user.set_password(self.other_user_password)
        self.other_user.save()

        self.other_contributor = Contributor \
            .objects \
            .create(admin=self.other_user,
                    name='test contributor 2',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.assertFalse(self.other_contributor.id == self.other_user.id,
                         'We want to verify that we serialize the proper ID '
                         'and we can only do that if we have distinct '
                         'Contributor and User ID values.')

        self.other_list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        self.other_source = Source \
            .objects \
            .create(facility_list=self.other_list,
                    is_active=True,
                    is_public=True,
                    contributor=self.other_contributor)

        self.other_list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(5, 5),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.other_source)

        self.other_match = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.CONFIRMED,
                    facility=self.facility,
                    facility_list_item=self.other_list_item,
                    confidence=0.85,
                    results='')

        self.extended_field_one = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value='name one',
                contributor=self.contributor,
                facility=self.facility,
                facility_list_item=self.list_item,
                is_verified=True
            )

        self.extended_field_two = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value='name two',
                contributor=self.other_contributor,
                facility=self.facility,
                facility_list_item=self.other_list_item
            )

        self.extended_field_three = ExtendedField \
            .objects \
            .create(
                field_name='native_language_name',
                value='name two',
                contributor=self.contributor,
                facility=self.facility,
                facility_list_item=self.list_item
            )

    def test_excludes_match_if_geocoded_point_is_none(self):
        self.other_list_item.geocoded_point = None
        self.other_list_item.save()
        response = self.client.get(
            '/api/facilities/{}/'.format(self.facility.id)
        )
        data = json.loads(response.content)
        self.assertEqual(
            len(data['properties']['other_locations']),
            0,
        )

    def test_serializes_other_match_location_in_facility_details(self):
        response = self.client.get(
            '/api/facilities/{}/'.format(self.facility.id)
        )

        data = json.loads(response.content)
        self.assertEqual(
            len(data['properties']['other_locations']),
            1,
        )

        self.assertIsNone(
            data['properties']['other_locations'][0]['notes'],
        )

        self.assertEqual(
            data['properties']['other_locations'][0]['lat'],
            5,
        )

        # The UI needs to build profile page links that use the ID of the User
        # who is the "admin" of the Contributor, not the ID of the Contributor
        # itself.
        self.assertEqual(
            data['properties']['other_locations'][0]['contributor_id'],
            self.other_user.id
        )

    def test_does_not_serialize_inactive_list_item_matches(self):
        self.other_source.is_active = False
        self.other_source.save()
        response = self.client.get(
            '/api/facilities/{}/'.format(self.facility.id)
        )

        data = json.loads(response.content)
        self.assertEqual(
            len(data['properties']['other_locations']),
            0,
        )

    def test_does_not_serialize_non_public_list_item_matches(self):
        self.other_source.is_public = False
        self.other_source.save()
        response = self.client.get(
            '/api/facilities/{}/'.format(self.facility.id)
        )

        data = json.loads(response.content)
        self.assertEqual(
            len(data['properties']['other_locations']),
            0,
        )

    def test_serializes_other_locations_in_facility_details(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        self.client.post(
            '/api/facilities/{}/update-location/'.format(self.facility.id), {
                UpdateLocationParams.LAT: 41,
                UpdateLocationParams.LNG: 43,
                UpdateLocationParams.NOTES: 'A note',
                UpdateLocationParams.CONTRIBUTOR_ID: self.other_contributor.id,
            })

        self.client.logout()

        response = self.client.get(
            '/api/facilities/{}/'.format(self.facility.id)
        )

        data = json.loads(response.content)
        self.assertEqual(
            len(data['properties']['other_locations']),
            3,
        )

        self.assertEqual(
            data['properties']['other_locations'][0]['notes'],
            'A note',
        )

        # The UI needs to build profile page links that use the ID of the User
        # who is the "admin" of the Contributor, not the ID of the Contributor
        # itself.
        self.assertEqual(
            data['properties']['other_locations'][0]['contributor_id'],
            self.other_user.id
        )

    def test_serializes_other_location_without_note_or_contributor(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        self.client.post(
            '/api/facilities/{}/update-location/'.format(self.facility.id), {
                UpdateLocationParams.LAT: 41,
                UpdateLocationParams.LNG: 43,
                UpdateLocationParams.NOTES: '',
            })

        self.client.logout()

        response = self.client.get(
            '/api/facilities/{}/'.format(self.facility.id)
        )

        data = json.loads(response.content)

        self.assertEqual(
            len(data['properties']['other_locations']),
            3,
        )

        self.assertEqual(
            data['properties']['other_locations'][0]['notes'],
            '',
        )

        self.assertEqual(
            data['properties']['other_locations'][0]['contributor_name'],
            None,
        )

    def test_serializes_extended_fields_in_facility_details(self):
        response = self.client.get(
            '/api/facilities/{}/'.format(self.facility.id)
        )
        data = json.loads(response.content)
        fields = data['properties']['extended_fields']['native_language_name']

        self.assertEqual(len(fields), 3)
        self.assertEqual(fields[0]['value'], 'name one')

        self.extended_field_one.is_verified = False
        self.extended_field_one.save()

        response = self.client.get(
            '/api/facilities/{}/'.format(self.facility.id)
        )
        data = json.loads(response.content)
        fields = data['properties']['extended_fields']['native_language_name']

        self.assertEqual(len(fields), 3)
        self.assertEqual(fields[0]['value'], 'name two')


class FacilityHistoryEndpointTest(FacilityAPITestCaseBase):
    def setUp(self):
        super(FacilityHistoryEndpointTest, self).setUp()
        self.history_url = '/api/facilities/{}/history/'.format(
            self.facility.id
        )

        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        self.list_two = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        self.source_two = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_two,
                    contributor=self.contributor)

        self.list_item_two = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_two)

        self.facility_two = Facility \
            .objects \
            .create(name='Name Two',
                    address='Address Two',
                    country_code='US',
                    location=Point(5, 5),
                    created_from=self.list_item_two)

        self.match_two = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_two,
                    facility_list_item=self.list_item_two,
                    confidence=0.85,
                    results='')

        self.list_item_two.facility = self.facility_two
        self.list_item_two.save()

        self.list_for_confirm_or_remove = FacilityList \
            .objects \
            .create(header='List for confirm or reject',
                    file_name='list for confirm or reject',
                    name='List for confirm or reject')

        self.source_for_confirm_or_remove = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_for_confirm_or_remove,
                    contributor=self.contributor)

        self.list_item_for_confirm_or_remove = FacilityListItem \
            .objects \
            .create(name='List item for confirmed match',
                    address='Address for confirmed match',
                    country_code='US',
                    row_index=2,
                    geocoded_point=Point(12, 34),
                    status=FacilityListItem.POTENTIAL_MATCH,
                    source=self.source_for_confirm_or_remove)

        self.match_for_confirm_or_remove = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.PENDING,
                    facility_list_item=self.list_item_for_confirm_or_remove,
                    facility=self.facility_two,
                    confidence=0.65,
                    results='')

        self.facility_two_history_url = '/api/facilities/{}/history/'.format(
            self.facility_two.id
        )

    @override_flag(FeatureGroups.CAN_GET_FACILITY_HISTORY, active=True)
    def test_serializes_deleted_facility_history(self):
        delete_facility_url = '/api/facilities/{}/'.format(self.facility.id)
        delete_response = self.client.delete(delete_facility_url)

        self.assertEqual(
            delete_response.status_code,
            204,
        )

        response = self.client.get(self.history_url)
        data = json.loads(response.content)

        self.assertEqual(
            data[0]['action'],
            'DELETE',
        )

        self.assertIn(
            'Deleted',
            data[0]['detail'],
        )

        self.assertEqual(
            len(data),
            4,
        )

    @override_flag('can_get_facility_history', active=True)
    def test_serializes_facility_location_change(self):
        location_change_url = '/api/facilities/{}/update-location/' \
            .format(self.facility.id)
        location_change_response = self.client.post(
            location_change_url, {
                UpdateLocationParams.LAT: 41,
                UpdateLocationParams.LNG: 43,
            })

        self.assertEqual(location_change_response.status_code, 200)

        response = self.client.get(self.history_url)
        data = json.loads(response.content)

        self.assertEqual(
            data[0]['action'],
            'UPDATE',
        )

        self.assertIn(
            'FacilityLocation',
            data[0]['detail'],
        )

        self.assertEqual(
            data[0]['changes']['location']['new']['coordinates'][0],
            43,
        )

        self.assertEqual(
            data[0]['changes']['location']['old']['coordinates'][0],
            0,
        )

        self.assertEqual(
            len(data),
            3,
        )

    @override_flag('can_get_facility_history', active=True)
    def test_serializes_facility_merge(self):
        merge_facilities_url = '/api/facilities/merge/?target={}&merge={}' \
            .format(self.facility_two.id, self.facility.id)

        merge_facilities_response = self.client.post(
            merge_facilities_url,
        )

        self.assertEqual(merge_facilities_response.status_code, 200)

        response = self.client.get(self.history_url)
        data = json.loads(response.content)

        self.assertEqual(
            data[0]['action'],
            'MERGE',
        )

        self.assertEqual(
            data[0]['detail'],
            'Merged with {}'.format(self.facility_two.id)
        )

        self.assertEqual(
            len(data),
            3,
        )

    @override_flag('can_get_facility_history', active=True)
    def test_serializes_facility_match_promotion(self):
        merge_facilities_url = '/api/facilities/merge/?target={}&merge={}' \
            .format(self.facility_two.id, self.facility.id)

        merge_facilities_response = self.client.post(
            merge_facilities_url,
        )

        self.assertEqual(merge_facilities_response.status_code, 200)

        self.facility_two.refresh_from_db()

        promote_facility_url = '/api/facilities/{}/promote/'.format(
            self.facility_two.id)

        promote_facility_data = {
            'match_id': self.match.id,
        }

        promote_facility_response = self.client.post(
            promote_facility_url,
            promote_facility_data,
        )

        self.assertEqual(promote_facility_response.status_code, 200)

        response = self.client.get(self.facility_two_history_url)
        data = json.loads(response.content)

        self.assertEqual(
            data[0]['action'],
            'UPDATE',
        )

        self.assertIn(
            'Promoted',
            data[0]['detail'],
        )

        self.assertEqual(
            data[0]['changes']['location']['new']['coordinates'][0],
            0,
        )

        self.assertEqual(
            data[0]['changes']['location']['old']['coordinates'][0],
            5,
        )

        self.assertEqual(
            data[0]['changes']['name']['new'],
            self.list_item.name,
        )

        self.assertEqual(
            data[0]['changes']['name']['old'],
            'Name Two',
        )

        self.assertEqual(
            data[0]['changes']['address']['new'],
            self.list_item.address,
        )

        self.assertEqual(
            data[0]['changes']['address']['old'],
            'Address Two',
        )

        self.assertEqual(
            len(data),
            4,
        )

    @override_flag('can_get_facility_history', active=True)
    def test_serializes_facility_match_split(self):
        merge_facilities_url = '/api/facilities/merge/?target={}&merge={}' \
            .format(self.facility_two.id, self.facility.id)

        merge_facilities_response = self.client.post(
            merge_facilities_url,
        )

        self.assertEqual(merge_facilities_response.status_code, 200)

        self.facility_two.refresh_from_db()

        split_facility_url = '/api/facilities/{}/split/'.format(
            self.facility_two.id)

        split_facility_data = {
            'match_id': self.match.id,
        }

        split_facility_response = self.client.post(
            split_facility_url,
            split_facility_data,
        )

        self.assertEqual(split_facility_response.status_code, 200)

        self.match.refresh_from_db()

        response = self.client.get(self.facility_two_history_url)
        data = json.loads(response.content)

        self.assertEqual(
            data[0]['action'],
            'SPLIT',
        )

        self.assertEqual(
            data[0]['detail'],
            '{} was split from {}'.format(
                self.match.facility.id,
                self.facility_two.id,
            )
        )

        self.assertEqual(
            len(data),
            4,
        )

    @override_flag('can_get_facility_history', active=True)
    def test_serializes_facility_match_move(self):
        move_facility_url = '/api/facilities/{}/move/'.format(
            self.facility.id)

        move_facility_data = {
            'match_id': self.match_two.id,
        }

        move_facility_response = self.client.post(
            move_facility_url,
            move_facility_data,
        )

        self.assertEqual(move_facility_response.status_code, 200)

        self.match_two.refresh_from_db()

        response = self.client.get(self.facility_two_history_url)
        data = json.loads(response.content)

        self.assertEqual(
            data[0]['action'],
            'MOVE',
        )

        self.assertEqual(
            data[0]['detail'],
            'Match {} was moved from {}'.format(
                self.match_two.id,
                self.facility_two.id,
            )
        )

        self.assertEqual(
            len(data),
            3,
        )

    @override_flag('can_get_facility_history', active=True)
    def test_handles_request_for_invalid_facility_id(self):
        invalid_history_url = '/api/facilities/hello/history/'
        invalid_history_response = self.client.get(invalid_history_url)
        self.assertEqual(invalid_history_response.status_code, 404)

    @override_flag('can_get_facility_history', active=True)
    def test_includes_association_for_automatic_match(self):
        automatic_match_response = self.client.get(
            self.facility_two_history_url,
        )

        data = json.loads(automatic_match_response.content)

        self.assertEqual(
            data[0]['action'],
            'ASSOCIATE',
        )

        self.assertEqual(
            data[0]['detail'],
            'Associate facility {} with {} via list {}'.format(
                self.facility_two.id,
                self.contributor.name,
                self.list_two.name,
            ),
        )

        self.assertEqual(
            len(data),
            2,
        )

        self.superuser.groups.add(
            Group.objects.get(name=FeatureGroups.CAN_SUBMIT_PRIVATE_FACILITY))
        self.superuser.save()

        automatic_match_response = self.client.get(
            self.facility_two_history_url,
        )

        data = json.loads(automatic_match_response.content)

        self.assertEqual(
            data[0]['detail'],
            'Associate facility {} with an Other'.format(
                self.facility_two.id,
            ),
        )

    @override_flag('can_get_facility_history', active=True)
    def test_associate_appears_after_create_in_history_data(self):
        history_response = self.client.get(
            self.history_url,
        )

        self.assertEqual(
            history_response.status_code,
            200,
        )

        data = json.loads(history_response.content)

        self.assertEqual(
            data[0]['action'],
            'ASSOCIATE',
        )

        self.assertEqual(
            data[1]['action'],
            'CREATE',
        )

        self.assertEqual(
            len(data),
            2,
        )

    @override_flag('can_get_facility_history', active=True)
    def test_includes_association_for_confirmed_match(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        confirm_url = '/api/facility-matches/{}/confirm/'.format(
            self.match_for_confirm_or_remove.id,
        )

        confirm_response = self.client.post(confirm_url)

        self.assertEqual(
            confirm_response.status_code,
            200,
        )

        confirmed_match_response = self.client.get(
            self.facility_two_history_url,
        )

        data = json.loads(confirmed_match_response.content)

        self.assertEqual(
            data[0]['action'],
            'ASSOCIATE',
        )

        self.assertEqual(
            data[0]['detail'],
            'Associate facility {} with {} via list {}'.format(
                self.facility_two.id,
                self.contributor.name,
                self.list_for_confirm_or_remove.name,
            ),
        )

        self.assertEqual(
            len(data),
            3,
        )

        self.user.groups.add(
            Group.objects.get(name=FeatureGroups.CAN_SUBMIT_PRIVATE_FACILITY))
        self.user.save()

        confirmed_match_response = self.client.get(
            self.facility_two_history_url,
        )

        data = json.loads(confirmed_match_response.content)

        self.assertEqual(
            data[0]['detail'],
            'Associate facility {} with an Other'.format(
                self.facility_two.id,
            ),
        )

    @override_flag('can_get_facility_history', active=True)
    def test_includes_dissociation_record_when_match_is_severed(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        confirm_url = '/api/facility-matches/{}/confirm/'.format(
            self.match_for_confirm_or_remove.id,
        )

        confirm_response = self.client.post(confirm_url)

        self.assertEqual(
            confirm_response.status_code,
            200,
        )

        remove_item_url = '/api/facility-lists/{}/remove/'.format(
            self.list_for_confirm_or_remove.id,
        )

        remove_item_data = {
            'list_item_id': self.list_item_for_confirm_or_remove.id,
        }

        remove_item_response = self.client.post(
            remove_item_url,
            remove_item_data,
        )

        self.assertEqual(
            remove_item_response.status_code,
            200,
        )

        removed_match_response = self.client.get(
            self.facility_two_history_url,
        )

        data = json.loads(removed_match_response.content)

        self.assertEqual(
            data[0]['action'],
            'DISSOCIATE',
        )

        self.assertEqual(
            data[0]['detail'],
            'Dissociate facility {} from {} via list {}'.format(
                self.facility_two.id,
                self.contributor.name,
                self.list_for_confirm_or_remove.name,
            ),
        )

        self.assertEqual(
            len(data),
            4,
        )

        self.user.groups.add(
            Group.objects.get(name=FeatureGroups.CAN_SUBMIT_PRIVATE_FACILITY))
        self.user.save()

        confirmed_match_response = self.client.get(
            self.facility_two_history_url,
        )

        data = json.loads(confirmed_match_response.content)

        self.assertEqual(
            data[0]['detail'],
            'Dissociate facility {} from an Other'.format(
                self.facility_two.id,
            ),
        )

    @override_flag('can_get_facility_history', active=True)
    def test_includes_dissociation_record_when_item_removed(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        confirm_url = '/api/facility-matches/{}/confirm/'.format(
            self.match_for_confirm_or_remove.id,
        )

        confirm_response = self.client.post(confirm_url)

        self.assertEqual(
            confirm_response.status_code,
            200,
        )

        # Upload replacement
        csv_file = SimpleUploadedFile('facilities.csv',
                                      b'country,name,address\n',
                                      content_type='text/csv')
        replace_response = self.client.post(
            reverse('facility-list-list'),
            {'file': csv_file,
             'replaces': self.list_for_confirm_or_remove.id},
            format='multipart')
        self.assertEqual(replace_response.status_code, status.HTTP_200_OK)

        removed_match_response = self.client.get(
            self.facility_two_history_url,
        )

        data = json.loads(removed_match_response.content)

        self.assertEqual(
            data[0]['action'],
            'DISSOCIATE',
        )

        self.assertEqual(
            data[0]['detail'],
            'Dissociate facility {} from {} via list {}'.format(
                self.facility_two.id,
                self.contributor.name,
                self.list_for_confirm_or_remove.name,
            ),
        )

    @override_flag('can_get_facility_history', active=True)
    def test_includes_dissociation_record_when_dissociate_api_is_called(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        confirm_url = '/api/facility-matches/{}/confirm/'.format(
            self.match_for_confirm_or_remove.id,
        )

        confirm_response = self.client.post(confirm_url)
        self.assertEqual(
            confirm_response.status_code,
            200,
        )

        dissociate_url = reverse('facility-dissociate',
                                 kwargs={'pk': self.facility_two.pk})
        dissociate_response = self.client.post(dissociate_url)
        self.assertEqual(
            dissociate_response.status_code,
            200,
        )

        history_url = reverse('facility-get-facility-history',
                              kwargs={'pk': self.facility_two.pk})
        history_response = self.client.get(history_url)
        self.assertEqual(
            history_response.status_code,
            200,
        )
        data = json.loads(history_response.content)
        self.assertEqual(
            data[0]['action'],
            'DISSOCIATE',
        )

        self.assertEqual(
            data[0]['detail'],
            'Dissociate facility {} from {} via list {}'.format(
                self.facility_two.id,
                self.contributor.name,
                self.list_for_confirm_or_remove.name,
            ),
        )

    @override_flag('can_get_facility_history', active=True)
    @override_switch('claim_a_facility', active=True)
    def test_includes_entry_for_claim_approval(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        claim_facility_url = '/api/facilities/{}/claim/'.format(
            self.facility_two.id,
        )

        claim_facility_data = {
            'contact_person': 'contact_person',
            'job_title': 'job_title',
            'company_name': 'company_name',
            'email': 'email@example.com',
            'phone_number': 1234567,
            'website': 'https://example.com',
            'facility_description': 'facility_description',
            'verification_method': 'verification_method',
            'preferred_contact_method': 'email',
        }

        claim_response = self.client.post(
            claim_facility_url,
            claim_facility_data,
        )

        self.assertEqual(
            claim_response.status_code,
            200,
        )

        self.client.logout()
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        claim = FacilityClaim.objects.first()

        approve_claim_url = '/api/facility-claims/{}/approve/'.format(
            claim.id,
        )

        approve_claim_response = self.client.post(
            approve_claim_url,
            {'reason': 'reason'},
        )

        self.assertEqual(
            approve_claim_response.status_code,
            200,
        )

        history_response = self.client.get(self.facility_two_history_url)

        self.assertEqual(
            history_response.status_code,
            200
        )

        data = json.loads(history_response.content)

        self.assertEqual(
            data[0]['action'],
            'CLAIM',
        )

        self.assertEqual(
            len(data),
            3,
        )

    @override_flag('can_get_facility_history', active=True)
    @override_switch('claim_a_facility', active=True)
    def test_handles_deleted_facility(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        claim_facility_url = '/api/facilities/{}/claim/'.format(
            self.facility_two.id,
        )

        claim_facility_data = {
            'contact_person': 'contact_person',
            'job_title': 'job_title',
            'company_name': 'company_name',
            'email': 'email@example.com',
            'phone_number': 1234567,
            'website': 'https://example.com',
            'facility_description': 'facility_description',
            'verification_method': 'verification_method',
            'preferred_contact_method': 'email',
        }

        claim_response = self.client.post(
            claim_facility_url,
            claim_facility_data,
        )

        self.assertEqual(
            claim_response.status_code,
            200,
        )

        self.client.logout()
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        claim = FacilityClaim.objects.first()

        approve_claim_url = '/api/facility-claims/{}/approve/'.format(
            claim.id,
        )

        approve_claim_response = self.client.post(
            approve_claim_url,
            {'reason': 'reason'},
        )

        self.assertEqual(
            approve_claim_response.status_code,
            200,
        )

        list = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='List')

        source = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=list,
                    contributor=self.contributor)

        list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=source)

        facility = Facility \
            .objects \
            .create(name='Name Two',
                    address='Address Two',
                    country_code='US',
                    location=Point(5, 5),
                    created_from=list_item)

        FacilityClaim.history.update(facility=facility)
        facility.delete()

        history_response = self.client.get(self.facility_two_history_url)

        self.assertEqual(
            history_response.status_code,
            200
        )

        data = json.loads(history_response.content)

        self.assertEqual(
            data[0]['action'],
            'ASSOCIATE',
        )

        self.assertEqual(
            len(data),
            2,
        )

    @override_flag('can_get_facility_history', active=True)
    @override_switch('claim_a_facility', active=True)
    def test_includes_entry_for_claim_revocation(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        claim_facility_url = '/api/facilities/{}/claim/'.format(
            self.facility_two.id,
        )

        claim_facility_data = {
            'contact_person': 'contact_person',
            'job_title': 'job_title',
            'company_name': 'company_name',
            'email': 'email@example.com',
            'phone_number': 1234567,
            'website': 'https://example.com',
            'facility_description': 'facility_description',
            'verification_method': 'verification_method',
            'preferred_contact_method': 'email',
        }

        claim_response = self.client.post(
            claim_facility_url,
            claim_facility_data,
        )

        self.assertEqual(
            claim_response.status_code,
            200,
        )

        self.client.logout()
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        claim = FacilityClaim.objects.first()

        approve_claim_url = '/api/facility-claims/{}/approve/'.format(
            claim.id,
        )

        approve_claim_response = self.client.post(
            approve_claim_url,
            {'reason': 'reason'},
        )

        self.assertEqual(
            approve_claim_response.status_code,
            200,
        )

        revoke_claim_url = '/api/facility-claims/{}/revoke/'.format(
            claim.id,
        )

        revoke_claim_response = self.client.post(
            revoke_claim_url,
            {'reason': 'reason'},
        )

        self.assertEqual(
            revoke_claim_response.status_code,
            200,
        )

        history_response = self.client.get(self.facility_two_history_url)

        self.assertEqual(
            history_response.status_code,
            200
        )

        data = json.loads(history_response.content)

        self.assertEqual(
            data[0]['action'],
            'CLAIM_REVOKE',
        )

        self.assertEqual(
            len(data),
            4,
        )

    @override_flag('can_get_facility_history', active=True)
    @override_switch('claim_a_facility', active=True)
    def test_includes_entry_for_public_claimed_facility_data_changes(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        claim_facility_url = '/api/facilities/{}/claim/'.format(
            self.facility_two.id,
        )

        claim_facility_data = {
            'contact_person': 'contact_person',
            'job_title': 'job_title',
            'company_name': 'company_name',
            'email': 'email@example.com',
            'phone_number': 1234567,
            'website': 'https://example.com',
            'facility_description': 'facility_description',
            'verification_method': 'verification_method',
            'preferred_contact_method': 'email',
        }

        claim_response = self.client.post(
            claim_facility_url,
            claim_facility_data,
        )

        self.assertEqual(
            claim_response.status_code,
            200,
        )

        self.client.logout()
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        claim = FacilityClaim.objects.first()

        approve_claim_url = '/api/facility-claims/{}/approve/'.format(
            claim.id,
        )

        approve_claim_response = self.client.post(
            approve_claim_url,
            {'reason': 'reason'},
        )

        self.assertEqual(
            approve_claim_response.status_code,
            200,
        )

        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        update_claim_url = '/api/facility-claims/{}/claimed/'.format(
            claim.id,
        )

        update_claim_data = {
            'id': claim.id,
            'facility_name_english': 'facility_name_english',
            'facility_name_native_language': 'facility_name_native_language',
            'facility_address': 'facility_address',
            'facility_description': 'facility_description',
            'facility_phone_number': 1234567,
            'facility_phone_number_publicly_visible': True,
            'facility_website': 'https://openapparel.org',
            'facility_website_publicly_visible': True,
            'facility_minimum_order_quantity': 10,
            'facility_average_lead_time': '2 months',
            'point_of_contact_person_name': 'point_of_contact_person_name',
            'point_of_contact_email': 'point_of_contact_email',
            'facility_workers_count': 20,
            'facility_female_workers_percentage': 50,
            'point_of_contact_publicly_visible': True,
            'office_official_name': 'office_official_name',
            'office_address': 'office_address',
            'office_country_code': 'US',
            'office_phone_number': 2345678,
            'office_info_publicly_visible': True,
            'facility_type': 'Cut and Sew / RMG',
        }

        update_claim_response = self.client.put(
            update_claim_url,
            update_claim_data,
        )

        self.assertEqual(
            update_claim_response.status_code,
            200,
        )

        history_response = self.client.get(self.facility_two_history_url)

        self.assertEqual(
            history_response.status_code,
            200
        )

        data = json.loads(history_response.content)

        self.assertEqual(
            data[0]['action'],
            'CLAIM_UPDATE',
        )

        self.assertEqual(
            len(data),
            4,
        )

        non_public_keys = [
            'facility_website',
            'facility_website_publicly_visible',
            'point_of_contact_publicly_visible',
            'point_of_contact_person_name',
            'point_of_contact_email',
            'office_info_publicly_visible',
            'office_official_name',
            'office_country_code',
            'office_address',
            'office_phone_number',
        ]

        for non_public_key in non_public_keys:
            self.assertNotIn(
                non_public_key,
                data[0]['changes'],
            )

    def test_unauthenticated_receives_401(self):
        self.client.logout()
        history_response = self.client.get(self.history_url)

        self.assertEqual(
            history_response.status_code,
            401,
        )

    def test_superuser_can_access_endpoint(self):
        # superuser is already signed in via `setUp`
        history_response = self.client.get(self.history_url)

        self.assertEqual(
            history_response.status_code,
            200,
        )

    def test_not_in_group_receives_403(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        history_response = self.client.get(self.history_url)

        self.assertEqual(
            history_response.status_code,
            403,
        )

    def test_in_group_receives_200(self):
        self.client.logout()

        history_group = auth.models.Group.objects.get(
            name='can_get_facility_history',
        )

        self.user.groups.set([history_group.id])
        self.user.save()

        self.client.login(email=self.user_email,
                          password=self.user_password)

        history_response = self.client.get(self.history_url)

        self.assertEqual(
            history_response.status_code,
            200,
        )


def is_json(myjson):
    try:
        json.loads(myjson)
    except ValueError:
        return False
    return True


class FacilitySubmitTest(FacilityAPITestCaseBase):
    def setUp(self):
        super(FacilitySubmitTest, self).setUp()
        self.url = reverse('facility-list')
        self.valid_facility = {
            'country': 'United States',
            'name': 'Pants Hut',
            'address': '123 Main St, Anywhereville, PA',
            'extra_1': 'Extra data'
        }

    def test_unauthenticated_receives_401(self):
        self.client.logout()
        response = self.client.post(self.url)
        self.assertEqual(response.status_code, 401)

    def test_not_in_group_receives_403(self):
        self.client.logout()
        self.client.login(email=self.user_email,
                          password=self.user_password)

        response = self.client.post(self.url)

        self.assertEqual(response.status_code, 403)

    @skip('DB is read-only')
    def test_empty_body_is_invalid(self):
        self.join_group_and_login()
        response = self.client.post(self.url)
        self.assertEqual(response.status_code, 400)

    @skip('DB is read-only')
    def test_missing_fields_are_invalid(self):
        self.join_group_and_login()

        response = self.client.post(self.url, {
            'country': 'US',
            'name': 'Something',
        })
        self.assertEqual(response.status_code, 400)

        response = self.client.post(self.url, {
            'country': 'US',
            'address': 'Some street',
        })
        self.assertEqual(response.status_code, 400)

        response = self.client.post(self.url, {
            'name': 'Something',
            'address': 'Some street',
        })
        self.assertEqual(response.status_code, 400)

    @skip('DB is read-only')
    def test_valid_request(self):
        self.join_group_and_login()
        response = self.client.post(self.url, self.valid_facility)
        self.assertEqual(response.status_code, 201)

    @skip('DB is read-only')
    def test_raw_data_json_formatted_with_singlequote(self):
        self.join_group_and_login()
        response = self.client.post(self.url, self.valid_facility)
        data = json.loads(response.content)
        list_item = FacilityListItem.objects.get(id=data['item_id'])
        self.assertTrue(is_json(list_item.raw_data))

    @skip('DB is read-only')
    def test_raw_data_json_formatted_with_doublequote(self):
        self.join_group_and_login()
        response = self.client.post(self.url, {
            "country": "United States",
            "name": "Pants Hut",
            "address": "123 Main St, Anywhereville, PA",
            "extra_1": "Extra data"
        })
        data = json.loads(response.content)
        list_item = FacilityListItem.objects.get(id=data['item_id'])
        self.assertTrue(is_json(list_item.raw_data))

    @skip('DB is read-only')
    def test_raw_data_json_formatted_with_querydict(self):
        self.join_group_and_login()
        query_dict = QueryDict('', mutable=True)
        query_dict.update(self.valid_facility)
        response = self.client.post(self.url, query_dict)
        data = json.loads(response.content)
        list_item = FacilityListItem.objects.get(id=data['item_id'])
        self.assertTrue(is_json(list_item.raw_data))

    def text_raw_data_with_internal_quotes(self):
        self.join_group_and_login()
        response = self.client.post(self.url, {
            'country': "US",
            'name': "Item",
            'address': "Address",
            'extra_2': "d'ataé"
        })
        data = json.loads(response.content)
        list_item = FacilityListItem.objects.get(id=data['item_id'])
        self.assertTrue(is_json(list_item.raw_data))

    @skip('DB is read-only')
    def test_valid_request_with_params(self):
        self.join_group_and_login()
        url_with_query = '{}?create=false&public=true'.format(self.url)
        response = self.client.post(url_with_query, self.valid_facility)
        self.assertEqual(response.status_code, 200)

    @skip('DB is read-only')
    def test_private_permission(self):
        self.join_group_and_login()
        url_with_query = '{}?public=false'.format(self.url)
        response = self.client.post(url_with_query, self.valid_facility)
        self.assertEqual(response.status_code, 403)

        group = auth.models.Group.objects.get(
            name=FeatureGroups.CAN_SUBMIT_PRIVATE_FACILITY,
        )
        self.user.groups.add(group.id)
        self.user.save()

        response = self.client.post(url_with_query, self.valid_facility)
        self.assertEqual(response.status_code, 201)

    @skip('DB is read-only')
    def test_creates_nonstandard_fields(self):
        self.join_group_and_login()
        self.client.post(self.url, self.valid_facility)
        fields = NonstandardField.objects.filter(
            contributor=self.user.contributor).values_list(
            'column_name', flat=True)
        self.assertEquals(1, len(fields))
        self.assertIn('extra_1', fields)

    @skip('DB is read-only')
    def test_exact_matches_with_create_false(self):
        self.join_group_and_login()
        url_with_query = '{}?create=false&public=true'.format(self.url)
        response = self.client.post(url_with_query, self.valid_facility)
        self.assertEqual(response.status_code, 200)
        response_two = self.client.post(url_with_query, self.valid_facility)
        self.assertEqual(response_two.status_code, 200)

    @skip('DB is read-only')
    def test_handles_exact_matches_with_empty_strings(self):
        self.join_group_and_login()
        url_with_query = '{}?public=true'.format(self.url)
        response = self.client.post(url_with_query, {
            'country': 'United States',
            'name': ',',
            'address': '123 Main St, Anywhereville, PA',
            'extra_1': 'Extra data'
        })
        self.assertEqual(response.status_code, 400)
        response_two = self.client.post(url_with_query, {
            'country': 'United States',
            'name': 'Pants Hut',
            'address': ',,',
            'extra_1': 'Extra data'
        })
        self.assertEqual(response_two.status_code, 400)


class FacilityCreateBodySerializerTest(TestCase):
    def test_valid_data(self):
        serializer = FacilityCreateBodySerializer(data={
            'country': 'United States',
            'name': 'Pants Hut',
            'address': '123 Main St, Anywhereville, PA'
        })
        self.assertTrue(serializer.is_valid())

    def test_missing_fields(self):
        serializer = FacilityCreateBodySerializer(data={
            'name': 'Pants Hut',
            'address': '123 Main St, Anywhereville, PA'
        })
        self.assertFalse(serializer.is_valid())
        self.assertIn('country', serializer.errors)
        self.assertNotIn('name', serializer.errors)
        self.assertNotIn('address', serializer.errors)

        serializer = FacilityCreateBodySerializer(data={
            'country': 'United States',
            'address': '123 Main St, Anywhereville, PA'
        })
        self.assertFalse(serializer.is_valid())
        self.assertIn('name', serializer.errors)
        self.assertNotIn('country', serializer.errors)
        self.assertNotIn('address', serializer.errors)

        serializer = FacilityCreateBodySerializer(data={
            'country': 'United States',
            'name': 'Pants Hut',
        })
        self.assertFalse(serializer.is_valid())
        self.assertIn('address', serializer.errors)
        self.assertNotIn('country', serializer.errors)
        self.assertNotIn('name', serializer.errors)

    def test_invalid_country(self):
        serializer = FacilityCreateBodySerializer(data={
            'country': 'Notrealia',
            'name': 'Pants Hut',
            'address': '123 Main St, Anywhereville, PA'
        })
        self.assertFalse(serializer.is_valid())
        self.assertIn('country', serializer.errors)


class FacilitySearchContributorTest(FacilityAPITestCaseBase):
    def setUp(self):
        super(FacilitySearchContributorTest, self).setUp()
        self.url = reverse('facility-list')
        self.private_user = User.objects.create(email='shh@hush.com')
        self.private_user_password = 'shhh'
        self.private_user.set_password(self.private_user_password)
        self.private_user.groups.set(
            auth.models.Group.objects.filter(
                name__in=[
                    FeatureGroups.CAN_SUBMIT_FACILITY,
                    FeatureGroups.CAN_SUBMIT_PRIVATE_FACILITY
                ]).values_list('id', flat=True))
        self.private_user.save()
        self.client.logout()

    def fetch_facility_contributors(self, facility):
        facility_url = '{}{}/'.format(self.url, facility.id)
        response = self.client.get(facility_url)
        data = json.loads(response.content)
        return data.get('properties', {}).get('contributors', [])

    def test_names(self):
        self.source.is_active = False
        self.source.save()

        self.contributor.contrib_type = (
            'Auditor / Certification Scheme / Service Provider'
        )
        self.contributor.save()
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual(
                'An Auditor / Certification Scheme / Service Provider',
                contributors[0].get('name'))

        self.contributor.contrib_type = 'Brand / Retailer'
        self.contributor.save()
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual('A Brand / Retailer',
                         contributors[0].get('name'))

    def test_inactive_contributor(self):
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual('test contributor 1 (First List)',
                         contributors[0].get('name'))

        self.source.is_active = False
        self.source.save()
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual('One Other', contributors[0].get('name'))

    def test_private_contributor(self):
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual('test contributor 1 (First List)',
                         contributors[0].get('name'))

        self.source.is_public = False
        self.source.save()
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual('One Other', contributors[0].get('name'))

    def test_multiple(self):
        user_two = User.objects.create(email='2@two.com')
        user_two.set_password('shhh')
        user_two.save()

        contributor_two = Contributor \
            .objects \
            .create(admin=user_two,
                    name='test contributor 2',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        source_two = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True,
                    contributor=contributor_two)

        list_item_two = FacilityListItem \
            .objects \
            .create(name='Item 2',
                    address='Address',
                    country_code='US',
                    row_index=0,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=source_two,
                    facility=self.facility)

        FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_two,
                    confidence=0.85,
                    results='')

        source_three = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True)

        list_item_three = FacilityListItem \
            .objects \
            .create(name='Item 3',
                    address='Address',
                    country_code='US',
                    row_index=0,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=source_three,
                    facility=self.facility)

        FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility,
                    facility_list_item=list_item_three,
                    confidence=0.85,
                    results='')

        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(2, len(contributors))

        source_two.is_active = False
        source_two.save()
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(2, len(contributors))
        self.assertEqual('test contributor 1 (First List)',
                         contributors[0].get('name'))
        self.assertEqual('One Other',
                         contributors[1].get('name'))

        self.match.is_active = False
        self.match.save()
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual('2 Others', contributors[0].get('name'))

    def test_private_user(self):
        self.client.login(email=self.private_user.email,
                          password=self.private_user_password)
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual('One Other', contributors[0].get('name'))

        self.private_user.groups.set(
            auth.models.Group.objects.filter(
                name__in=[
                    FeatureGroups.CAN_SUBMIT_FACILITY,
                    FeatureGroups.CAN_SUBMIT_PRIVATE_FACILITY,
                    FeatureGroups.CAN_VIEW_FULL_CONTRIB_DETAIL
                ]).values_list('id', flat=True))
        self.private_user.save()
        contributors = self.fetch_facility_contributors(self.facility)
        self.assertEqual(1, len(contributors))
        self.assertEqual('test contributor 1 (First List)',
                         contributors[0].get('name'))

    def test_inactive_or_private_contributor_omitted(self):
        def get_facility_count():
            url = '{}?contributors={}'.format(
                self.url, self.contributor.id)
            response = self.client.get(url)
            data = json.loads(response.content)
            return int(data.get('count'))

        self.assertEqual(1, get_facility_count())

        self.source.is_public = False
        self.source.is_active = True
        self.source.save()
        self.assertEqual(0, get_facility_count())

        self.source.is_public = True
        self.source.is_active = False
        self.source.save()
        self.assertEqual(0, get_facility_count())


class SingleItemFacilityMatchTest(FacilityAPITestCaseBase):
    def setUp(self):
        super(SingleItemFacilityMatchTest, self).setUp()
        self.contributor_two = Contributor \
            .objects \
            .create(admin=self.superuser,
                    name='test contributor 2',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.source_two = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor_two)

        self.list_item_two = FacilityListItem \
            .objects \
            .create(name='Item 2',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.POTENTIAL_MATCH,
                    source=self.source_two)

        self.match_two = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.PENDING,
                    facility=self.facility,
                    facility_list_item=self.list_item_two,
                    confidence=0.75,
                    results='')

        self.list_item_two.facility = self.facility
        self.list_item_two.save()

    def match_url(self, match, action='detail'):
        return reverse('facility-match-{}'.format(action),
                       kwargs={'pk': match.pk})

    def test_get_match_detail(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        response = self.client.get(self.match_url(self.match_two))
        self.assertEqual(200, response.status_code)

    def test_only_contributor_can_get_match_detail(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        response = self.client.get(self.match_url(self.match))
        self.assertEqual(404, response.status_code)

    def test_confirm(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        response = self.client.post(
            self.match_url(self.match_two, action='confirm'))
        self.assertEqual(200, response.status_code)

    def test_only_contributor_can_confirm(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)

        response = self.client.post(
            self.match_url(self.match, action='confirm'))
        self.assertEqual(404, response.status_code)

    def test_reject(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.post(
            self.match_url(self.match_two, action='reject'))
        self.assertEqual(200, response.status_code)

    def test_only_contributor_can_reject(self):
        self.client.login(email=self.superuser_email,
                          password=self.superuser_password)
        response = self.client.post(
            self.match_url(self.match, action='reject'))
        self.assertEqual(404, response.status_code)


class FacilitySearchTest(FacilityAPITestCaseBase):
    def setUp(self):
        super(FacilitySearchTest, self).setUp()

        self.user_two_email = 'two@example.com'
        self.user_two_password = 'example123'
        self.user_two = User.objects.create(email=self.user_two_email)
        self.user_two.set_password(self.user_two_password)
        self.user_two.save()

        self.contributor_two = Contributor \
            .objects \
            .create(admin=self.user_two,
                    name='test contributor 2',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_two = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='Second List')

        self.source_two = Source \
            .objects \
            .create(facility_list=self.list_two,
                    source_type=Source.LIST,
                    contributor=self.contributor_two)

        self.list_item_two = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_two)

        self.facility_two = Facility \
            .objects \
            .create(name='Name Two',
                    address='Address Two',
                    country_code='US',
                    location=Point(5, 5),
                    created_from=self.list_item_two)

        self.match_two = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_two,
                    facility_list_item=self.list_item_two,
                    confidence=0.85,
                    results='')

        self.list_item_two.facility = self.facility_two
        self.list_item_two.save()

        self.source_two_b = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    contributor=self.contributor)

        self.list_item_two_b = FacilityListItem \
            .objects \
            .create(name='Item 2b',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_two_b)

        self.match_two_b = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.AUTOMATIC,
                    facility=self.facility_two,
                    facility_list_item=self.list_item_two_b,
                    confidence=0.85,
                    results='')

        self.list_item_two_b.facility = self.facility_two
        self.list_item_two_b.save()

        self.base_url = reverse('facility-list')
        self.contributor_or_url = self.base_url + \
            '?contributors={}&contributors={}'
        self.contributor_and_url = self.base_url + \
            '?contributors={}&contributors={}&combine_contributors=AND'

        self.client.login(email=self.user_email,
                          password=self.user_password)

    def assert_response_count(self, response, count):
        data = json.loads(response.content)
        self.assertEqual(count, int(data['count']))

    def test_contributor_or_search(self):
        response = self.client.get(
            self.contributor_or_url.format(
                self.contributor.id,
                self.contributor_two.id))
        self.assert_response_count(response, 2)

    def test_contributor_and_search(self):
        response = self.client.get(
            self.contributor_and_url.format(
                self.contributor.id,
                self.contributor_two.id))
        self.assert_response_count(response, 1)

    def test_contributor_and_inactive_match(self):
        self.match_two_b.is_active = False
        self.match_two_b.save()

        response = self.client.get(
            self.contributor_and_url.format(
                self.contributor.id,
                self.contributor_two.id))
        self.assert_response_count(response, 0)

    def test_contributor_and_inactive_source(self):
        self.source_two_b.is_active = False
        self.source_two_b.save()

        response = self.client.get(
            self.contributor_and_url.format(
                self.contributor.id,
                self.contributor_two.id))
        self.assert_response_count(response, 0)

    def test_contributor_and_private_source(self):
        self.source_two_b.is_public = False
        self.source_two_b.save()

        response = self.client.get(
            self.contributor_and_url.format(
                self.contributor.id,
                self.contributor_two.id))
        self.assert_response_count(response, 0)

    def test_contributor_and_pending_match(self):
        self.match_two_b.status = FacilityMatch.PENDING
        self.match_two_b.save()

        response = self.client.get(
            self.contributor_and_url.format(
                self.contributor.id,
                self.contributor_two.id))
        self.assert_response_count(response, 0)


class ListWithoutSourceTest(TestCase):
    def test_str(self):
        facility_list = FacilityList()
        # We are checking that the __str__ method does not raise an exception.
        str(facility_list)

    def test_serializer(self):
        facility_list = FacilityList()
        # Checking the `data` property triggers the serialization. We are
        # checking that it does not raise an exception.
        FacilityListSerializer(facility_list).data


class ContributorTypesTest(FacilityAPITestCaseBase):
    def get_contributor_types(self):
        return self.client.get(reverse('all_contributor_types'))

    def fetch_and_assert_all_counts_are_zero(self):
        response = self.get_contributor_types()
        data = json.loads(response.content)
        for (id, label) in data:
            self.assertEqual(id, label)

    def test_all_types_are_returned(self):
        response = self.get_contributor_types()
        data = json.loads(response.content)
        self.assertEqual(len(Contributor.CONTRIB_TYPE_CHOICES), len(data))

    def test_only_public_sources_are_counted(self):
        self.source.is_public = False
        self.source.save()
        self.fetch_and_assert_all_counts_are_zero()

    def test_only_active_sources_are_counted(self):
        self.source.is_active = False
        self.source.save()
        self.fetch_and_assert_all_counts_are_zero()

    def test_only_confirmed_items_are_counted(self):
        self.list_item.status = FacilityListItem.GEOCODED
        self.list_item.save()
        self.fetch_and_assert_all_counts_are_zero()


class PPEFieldTest(TestCase):
    def setUp(self):
        self.email_one = 'one@example.com'
        self.email_two = 'two@example.com'
        self.user_one = User.objects.create(email=self.email_one)
        self.user_two = User.objects.create(email=self.email_two)
        self.user_two.set_password('password')
        self.user_two.save()

        self.contrib_one = Contributor \
            .objects \
            .create(admin=self.user_one,
                    name='contributor one',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_two = Contributor \
            .objects \
            .create(admin=self.user_two,
                    name='contributor two',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_one = FacilityList \
            .objects \
            .create(header='header',
                    file_name='one',
                    name='list_one')

        self.list_two = FacilityList \
            .objects \
            .create(header='header',
                    file_name='two',
                    name='list_two')

        self.source_one = Source \
            .objects \
            .create(facility_list=self.list_one,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contrib_one)

        self.source_two = Source \
            .objects \
            .create(facility_list=self.list_two,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    create=True,
                    contributor=self.contrib_two)

        self.list_item_one = FacilityListItem \
            .objects \
            .create(name='name',
                    address='address',
                    country_code='US',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_one)

        self.list_item_two = FacilityListItem \
            .objects \
            .create(name='name',
                    address='address',
                    country_code='US',
                    ppe_product_types=['Masks_Two', 'Gloves_Two'],
                    ppe_contact_phone='222-222-2222',
                    ppe_contact_email='two@example.com',
                    ppe_website='http://example.com/two',
                    row_index=1,
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.source_two,
                    geocoded_point=Point(0, 0))

        self.facility = Facility \
            .objects \
            .create(name='name',
                    address='address',
                    country_code='US',
                    location=Point(0, 0),
                    created_from=self.list_item_one)

        self.facility_match_one = FacilityMatch \
            .objects \
            .create(status=FacilityMatch.CONFIRMED,
                    facility=self.facility,
                    results="",
                    facility_list_item=self.list_item_one)

    def make_match_results(self, list_item_id, facility_id, score):
        return {
            'processed_list_item_ids': [list_item_id],
            'item_matches': {
                list_item_id: [(facility_id, np.float32(score))]
            },
            'results': {
                'gazetteer_threshold': 50,
                'automatic_threshold': 80,
                'recall_weight': 0.5,
                'code_version': 'abcd1234',
            },
            'started': str(datetime.utcnow()),
            'finished': str(datetime.utcnow()),
        }

    def test_match_populates_ppe(self):
        results = self.make_match_results(self.list_item_two.id,
                                          self.facility.id, 100)
        save_match_details(results)
        self.facility.refresh_from_db()

        self.assertEqual(self.list_item_two.ppe_product_types,
                         self.facility.ppe_product_types)
        self.assertEqual(self.list_item_two.ppe_contact_phone,
                         self.facility.ppe_contact_phone)
        self.assertEqual(self.list_item_two.ppe_contact_email,
                         self.facility.ppe_contact_email)
        self.assertEqual(self.list_item_two.ppe_website,
                         self.facility.ppe_website)

    def test_match_does_not_overwrite_ppe_product_types(self):
        self.facility.ppe_product_types = ['TEST']
        self.facility.save()

        results = self.make_match_results(self.list_item_two.id,
                                          self.facility.id, 100)
        save_match_details(results)
        self.facility.refresh_from_db()

        self.assertEqual(['TEST'],
                         self.facility.ppe_product_types)
        self.assertEqual(self.list_item_two.ppe_contact_phone,
                         self.facility.ppe_contact_phone)
        self.assertEqual(self.list_item_two.ppe_contact_email,
                         self.facility.ppe_contact_email)
        self.assertEqual(self.list_item_two.ppe_website,
                         self.facility.ppe_website)

    def test_match_does_not_overwrite_ppe_contact_phone(self):
        self.facility.ppe_contact_phone = 'ttt-ttt-tttt'
        self.facility.save()

        results = self.make_match_results(self.list_item_two.id,
                                          self.facility.id, 100)
        save_match_details(results)
        self.facility.refresh_from_db()

        self.assertEqual(self.list_item_two.ppe_product_types,
                         self.facility.ppe_product_types)
        self.assertEqual('ttt-ttt-tttt',
                         self.facility.ppe_contact_phone)
        self.assertEqual(self.list_item_two.ppe_contact_email,
                         self.facility.ppe_contact_email)
        self.assertEqual(self.list_item_two.ppe_website,
                         self.facility.ppe_website)

    def test_match_does_not_overwrite_ppe_contact_email(self):
        self.facility.ppe_contact_email = 'TTT@TT.COM'
        self.facility.save()

        results = self.make_match_results(self.list_item_two.id,
                                          self.facility.id, 100)
        save_match_details(results)
        self.facility.refresh_from_db()

        self.assertEqual(self.list_item_two.ppe_product_types,
                         self.facility.ppe_product_types)
        self.assertEqual(self.list_item_two.ppe_contact_phone,
                         self.facility.ppe_contact_phone)
        self.assertEqual('TTT@TT.COM',
                         self.facility.ppe_contact_email)
        self.assertEqual(self.list_item_two.ppe_website,
                         self.facility.ppe_website)

    def test_match_does_not_overwrite_ppe_website(self):
        self.facility.ppe_website = 'HTTP://TEST.COM'
        self.facility.save()

        results = self.make_match_results(self.list_item_two.id,
                                          self.facility.id, 100)
        save_match_details(results)
        self.facility.refresh_from_db()

        self.assertEqual(self.list_item_two.ppe_product_types,
                         self.facility.ppe_product_types)
        self.assertEqual(self.list_item_two.ppe_contact_phone,
                         self.facility.ppe_contact_phone)
        self.assertEqual(self.list_item_two.ppe_contact_email,
                         self.facility.ppe_contact_email)
        self.assertEqual('HTTP://TEST.COM',
                         self.facility.ppe_website)

    def match_url(self, match, action='detail'):
        return reverse('facility-match-{}'.format(action),
                       kwargs={'pk': match.pk})

    def test_confirm_match_populates_ppe(self):
        results = self.make_match_results(self.list_item_two.id,
                                          self.facility.id, 70)
        save_match_details(results)

        pending_qs = self.facility.facilitymatch_set.filter(
            status=FacilityMatch.PENDING)
        self.assertEqual(1, pending_qs.count())
        match = pending_qs[0]
        self.assertEqual(FacilityMatch.PENDING, match.status)

        self.client.login(email=self.email_two,
                          password='password')
        response = self.client.post(
            self.match_url(match, action='confirm')
        )
        self.assertEqual(200, response.status_code)

        self.facility.refresh_from_db()

        self.assertEqual(self.list_item_two.ppe_product_types,
                         self.facility.ppe_product_types)
        self.assertEqual(self.list_item_two.ppe_contact_phone,
                         self.facility.ppe_contact_phone)
        self.assertEqual(self.list_item_two.ppe_contact_email,
                         self.facility.ppe_contact_email)
        self.assertEqual(self.list_item_two.ppe_website,
                         self.facility.ppe_website)

    def reject_match_and_assert(self):
        """
        This helper creates a potential match for line_item_two, submits a
        request to reject it, and asserts that a new facility is created from
        line_item_two. The newly created `Facility` object is returned.
        """
        results = self.make_match_results(self.list_item_two.id,
                                          self.facility.id, 70)
        save_match_details(results)

        pending_qs = self.facility.facilitymatch_set.filter(
            status=FacilityMatch.PENDING)
        self.assertEqual(1, pending_qs.count())
        match = pending_qs[0]
        self.assertEqual(FacilityMatch.PENDING, match.status)

        self.client.login(email=self.email_two,
                          password='password')
        response = self.client.post(
            self.match_url(match, action='reject')
        )
        self.assertEqual(200, response.status_code)

        facility = Facility.objects.get(created_from=self.list_item_two)

        self.assertEqual(self.list_item_two.ppe_product_types,
                         facility.ppe_product_types)
        self.assertEqual(self.list_item_two.ppe_contact_phone,
                         facility.ppe_contact_phone)
        self.assertEqual(self.list_item_two.ppe_contact_email,
                         facility.ppe_contact_email)
        self.assertEqual(self.list_item_two.ppe_website,
                         facility.ppe_website)

        return facility

    def test_reject_match_creates_facility_with_ppe(self):
        self.reject_match_and_assert()

    def test_deactivating_created_from_source_clears_ppe(self):
        facility = self.reject_match_and_assert()

        facility.created_from.source.is_active = False
        facility.created_from.source.save()
        facility.refresh_from_db()

        self.assertEqual([], facility.ppe_product_types)
        self.assertEqual('', facility.ppe_contact_phone)
        self.assertEqual('', facility.ppe_contact_email)
        self.assertEqual('', facility.ppe_website)

    def test_deactivating_created_from_match_clears_ppe(self):
        facility = self.reject_match_and_assert()

        for match in facility.created_from.facilitymatch_set.all():
            match.is_active = False
            match.save()
        facility.refresh_from_db()

        self.assertEqual([], facility.ppe_product_types)
        self.assertEqual('', facility.ppe_contact_phone)
        self.assertEqual('', facility.ppe_contact_email)
        self.assertEqual('', facility.ppe_website)


class ApiLimitTest(TestCase):
    def setUp(self):
        self.email_one = 'one@example.com'
        self.email_two = 'two@example.com'
        self.user_one = User.objects.create(email=self.email_one)
        self.user_two = User.objects.create(email=self.email_two)

        self.contrib_one = Contributor \
            .objects \
            .create(admin=self.user_one,
                    name='contributor one',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.contrib_two = Contributor \
            .objects \
            .create(admin=self.user_two,
                    name='contributor two',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)
        now = timezone.now()
        self.limit_one = ApiLimit.objects.create(contributor=self.contrib_one,
                                                 yearly_limit=10,
                                                 period_start_date=now)

        self.limit_two = ApiLimit.objects.create(contributor=self.contrib_two,
                                                 yearly_limit=10,
                                                 period_start_date=now)

        self.notification_time = timezone.now()
        self.notification = ContributorNotifications \
            .objects \
            .create(contributor=self.contrib_two,
                    api_limit_warning_sent_on=self.notification_time,
                    api_limit_exceeded_sent_on=self.notification_time,
                    api_grace_limit_exceeded_sent_on=self.notification_time)

    def test_under_limit_does_nothing(self):
        check_api_limits(timezone.now())
        self.assertEqual(ApiBlock.objects.filter(
                         contributor=self.contrib_one).count(), 0)

    def test_limit_only_applies_within_period(self):
        last_month = timezone.now() - relativedelta(months=1)
        for x in range(10):
            r = RequestLog.objects.create(user=self.user_one,
                                          response_code=200)
            r.created_at = last_month
            r.save()

        check_api_limits(timezone.now())

        self.assertEqual(ApiBlock.objects.filter(
                         contributor=self.contrib_one).count(), 0)

        warning = ContributorNotifications.objects.get(
                  contributor=self.contrib_one)
        self.assertIsNone(warning.api_limit_warning_sent_on)

    def test_limit_warning_sent_once(self):
        for x in range(10):
            RequestLog.objects.create(user=self.user_one, response_code=200)
            RequestLog.objects.create(user=self.user_two, response_code=200)

        check_api_limits(timezone.now())

        self.assertEqual(ApiBlock.objects.filter(
                         contributor=self.contrib_one).count(), 0)
        self.assertEqual(ApiBlock.objects.filter(
                         contributor=self.contrib_two).count(), 0)

        warning = ContributorNotifications.objects.get(
                  contributor=self.contrib_one)
        self.assertIsNotNone(warning.api_limit_warning_sent_on)

        warning_two = ContributorNotifications.objects.get(
                      contributor=self.contrib_two)
        self.assertEqual(warning_two.api_limit_warning_sent_on,
                         self.notification_time)

    def test_over_limit_block_set_once(self):
        ApiBlock.objects.create(contributor=self.contrib_two,
                                until=get_end_of_year(self.notification_time),
                                active=False, limit=10, actual=11)

        for x in range(11):
            RequestLog.objects.create(user=self.user_one, response_code=200)
            RequestLog.objects.create(user=self.user_two, response_code=200)

        check_api_limits(timezone.now())

        self.assertEqual(ApiBlock.objects.filter(
                         contributor=self.contrib_one).count(), 1)
        self.assertEqual(ApiBlock.objects.filter(
                         contributor=self.contrib_two).count(), 1)

        notice = ContributorNotifications.objects.get(
                 contributor=self.contrib_one)
        self.assertIsNotNone(notice.api_limit_exceeded_sent_on)

        notice_two = ContributorNotifications.objects.get(
                     contributor=self.contrib_two)
        self.assertEqual(notice_two.api_limit_exceeded_sent_on,
                         self.notification_time)


class CloseListTest(TestCase):
    def setUp(self):
        self.country_code = 'US'

        self.user = User.objects.create(email='one@example.com')

        self.contrib = Contributor \
            .objects \
            .create(admin=self.user,
                    name='contributor',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        self.list_one = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one",
                    name='list 1')

        self.source_one = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_one,
                    contributor=self.contrib)

        self.list_item_one = FacilityListItem \
            .objects \
            .create(row_index=0,
                    source=self.source_one,
                    status=FacilityListItem.MATCHED)

        self.facility_one = Facility.objects.create(
            country_code=self.country_code,
            created_from=self.list_item_one,
            facilitylistitem=self.list_item_one,
            location=Point(0, 0),
            is_closed=False,
        )
        self.list_item_one.facility = self.facility_one
        self.list_item_one.save()

        self.list_item_one_b = FacilityListItem \
            .objects \
            .create(row_index=0,
                    source=self.source_one,
                    status=FacilityListItem.MATCHED)

        self.facility_one_b = Facility.objects.create(
            country_code=self.country_code,
            created_from=self.list_item_one_b,
            facilitylistitem=self.list_item_one_b,
            location=Point(0, 0),
            is_closed=False,
        )
        self.list_item_one_b.facility = self.facility_one_b
        self.list_item_one_b.save()

        self.list_two = FacilityList \
            .objects \
            .create(header="header",
                    file_name="one-b",
                    name='list 2')

        self.source_two = Source \
            .objects \
            .create(source_type=Source.LIST,
                    facility_list=self.list_two,
                    contributor=self.contrib)

        self.list_item_two = FacilityListItem \
            .objects \
            .create(row_index=0,
                    source=self.source_two,
                    status=FacilityListItem.MATCHED)

        self.facility_two = Facility.objects.create(
            country_code=self.country_code,
            created_from=self.list_item_two,
            facilitylistitem=self.list_item_two,
            location=Point(0, 0),
            is_closed=False,
        )
        self.list_item_two.facility = self.facility_two
        self.list_item_two.save()

    def test_closes_list(self):
        close_list(self.list_one.id, self.user.id)

        f_one = Facility.objects.get(id=self.facility_one.id)
        f_one_b = Facility.objects.get(id=self.facility_one_b.id)
        f_two = Facility.objects.get(id=self.facility_two.id)

        self.assertTrue(f_one.is_closed)
        self.assertTrue(f_one_b.is_closed)
        self.assertFalse(f_two.is_closed)

        activity = FacilityActivityReport.objects.all().count()
        self.assertEquals(2, activity)


class NonstandardFieldsApiTest(APITestCase):
    def setUp(self):
        self.url = reverse('nonstandard-fields-list')
        self.user_email = 'test@example.com'
        self.user_password = 'example123'
        self.user = User.objects.create(email=self.user_email)
        self.user.set_password(self.user_password)
        self.user.save()

        self.contributor = Contributor \
            .objects \
            .create(admin=self.user,
                    name='test contributor 1',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

    def test_nonstandard_fields(self):
        NonstandardField.objects.create(
            column_name='extra_1',
            contributor=self.contributor)
        NonstandardField.objects.create(
            column_name='extra_2',
            contributor=self.contributor)
        self.client.login(email=self.user_email,
                          password=self.user_password)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = json.loads(response.content)
        self.assertEquals(7, len(content))
        self.assertIn('extra_1', content)
        self.assertIn('extra_2', content)
        self.assertIn('parent_company', content)

    def test_without_nonstandard_fields(self):
        self.client.login(email=self.user_email,
                          password=self.user_password)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = json.loads(response.content)
        self.assertEquals(5, len(content))
        self.assertNotIn('extra_1', content)
        self.assertNotIn('extra_2', content)
        self.assertIn('parent_company', content)


class ContributorFieldsApiTest(APITestCase):
    def setUp(self):
        self.url = '/api/facilities/'
        self.user_email = 'test@example.com'
        self.user_password = 'example123'
        self.user = User.objects.create(email=self.user_email)
        self.user.set_password(self.user_password)
        self.user.save()

        self.embed_config = EmbedConfig.objects.create()

        self.contributor = Contributor \
            .objects \
            .create(admin=self.user,
                    name='test contributor 1',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE,
                    embed_config=self.embed_config)
        self.embed_one = EmbedField.objects.create(
            embed_config=self.embed_config,
            order=0,
            column_name='extra_1',
            display_name='ExtraOne',
            visible=True,
            searchable=True
        )
        self.embed_two = EmbedField.objects.create(
            embed_config=self.embed_config,
            order=1,
            column_name='extra_2',
            display_name='ExtraTwo',
            visible=True,
            searchable=True
        )

        self.list = FacilityList \
            .objects \
            .create(header='country,name,address,extra_1',
                    file_name='one',
                    name='First List')

        self.list_source = Source \
            .objects \
            .create(facility_list=self.list,
                    source_type=Source.LIST,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        raw_data = '"US","Towel Factory 42","42 Dolphin St","data one"'
        self.list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.list_source,
                    raw_data=raw_data)

        self.facility = Facility.objects.create(
            country_code=self.list_item.country_code,
            created_from=self.list_item,
            location=Point(0, 0),
        )
        self.list_item.facility = self.facility
        self.list_item.save()

        self.match_one = FacilityMatch.objects.create(
            facility_list_item=self.list_item,
            facility=self.facility,
            is_active=True,
            results={}
        )

        self.api_source = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True,
                    contributor=self.contributor)

        self.api_list_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    raw_data=("{'country': 'US', 'name': 'Item'," +
                              "'address': 'Address', 'extra_2': 'data two'}"),
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.api_source)
        self.facility_two = Facility.objects.create(
            country_code=self.api_list_item.country_code,
            created_from=self.api_list_item,
            location=Point(0, 0),
        )
        self.api_list_item.facility = self.facility_two
        self.api_list_item.save()
        self.match_two = FacilityMatch.objects.create(
            facility_list_item=self.api_list_item,
            facility=self.facility_two,
            is_active=True,
            results={}
        )

    def test_list_fields(self):
        response = self.client.get(
            self.url + self.facility.id + '/?embed=1&contributor=' +
            str(self.contributor.id))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = json.loads(response.content)
        contributor_fields = content['properties']['contributor_fields']
        self.assertEquals(2, len(contributor_fields))
        field = contributor_fields[0]
        field_two = contributor_fields[1]
        self.assertEquals('ExtraOne', field['label'])
        self.assertEquals('data one', field['value'])
        self.assertEquals('ExtraTwo', field_two['label'])
        self.assertEquals(None, field_two['value'])

    def test_single_fields(self):
        response = self.client.get(
            self.url + self.facility_two.id + '/?embed=1&contributor=' +
            str(self.contributor.id))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = json.loads(response.content)
        contributor_fields = content['properties']['contributor_fields']
        self.assertEquals(2, len(contributor_fields))
        field = contributor_fields[0]
        field_two = contributor_fields[1]
        self.assertEquals('ExtraOne', field['label'])
        self.assertEquals(None, field['value'])
        self.assertEquals('ExtraTwo', field_two['label'])
        self.assertEquals('data two', field_two['value'])

    def test_without_embed(self):
        response = self.client.get(
            self.url + self.facility.id + '/?contributor=' +
            str(self.contributor.id))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = json.loads(response.content)
        contributor_fields = content['properties']['contributor_fields']
        self.assertEquals(0, len(contributor_fields))

    def test_without_contributor(self):
        response = self.client.get(
            self.url + self.facility.id + '/?embed=1')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = json.loads(response.content)
        contributor_fields = content['properties']['contributor_fields']
        self.assertEquals(0, len(contributor_fields))

    def test_inactive_match(self):
        self.match_one.is_active = False
        self.match_one.save()
        response = self.client.get(
            self.url + self.facility.id + '/?embed=1&contributor=' +
            str(self.contributor.id))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = json.loads(response.content)
        contributor_fields = content['properties']['contributor_fields']
        self.assertEquals(2, len(contributor_fields))
        field = contributor_fields[0]
        field_two = contributor_fields[1]
        self.assertEquals('ExtraOne', field['label'])
        self.assertEquals(None, field['value'])
        self.assertEquals('ExtraTwo', field_two['label'])
        self.assertEquals(None, field_two['value'])

    def test_custom_text(self):
        indexes = FacilityIndex.objects.count()
        self.assertEquals(2, indexes)
        index_one = FacilityIndex.objects.get(id=self.facility.id)
        index_one_data = '{}|data one'.format(self.contributor.id)
        self.assertIn(index_one_data, index_one.custom_text)
        index_two = FacilityIndex.objects.get(id=self.facility_two.id)
        index_two_data = '{}|data two'.format(self.contributor.id)
        self.assertIn(index_two_data, index_two.custom_text)

    def test_custom_text_excludes_unsearchable(self):
        self.embed_two.searchable = False
        self.embed_two.save()

        # Run indexing manually because we don't use a post_save signal for
        # embed fields (since they are deleted and recreated each time the
        # config is updated, it creates a lot of overhead) and instead update
        # the index directly in the embed config update views
        f_ids = Facility.objects \
            .filter(facilitylistitem__source__contributor=self.contributor)\
            .values_list('id', flat=True)
        if len(f_ids) > 0:
            index_custom_text(f_ids)

        index_one = FacilityIndex.objects.get(id=self.facility.id)
        index_one_data = '{}|data one'.format(self.contributor.id)
        self.assertIn(index_one_data, index_one.custom_text)
        index_two = FacilityIndex.objects.get(id=self.facility_two.id)
        index_two_data = '{}|data two'.format(self.contributor.id)
        self.assertNotIn(index_two_data, index_two.custom_text)

    def test_custom_text_excludes_inactive(self):
        self.match_one.is_active = False
        self.match_one.save()

        index_one = FacilityIndex.objects.get(id=self.facility.id)
        index_one_data = '{}|data one'.format(self.contributor.id)
        self.assertNotIn(index_one_data, index_one.custom_text)
        index_two = FacilityIndex.objects.get(id=self.facility_two.id)
        index_two_data = '{}|data two'.format(self.contributor.id)
        self.assertIn(index_two_data, index_two.custom_text)

    def test_custom_text_uses_most_recent(self):
        new_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    raw_data=("{'country': 'US', 'name': 'Item'," +
                              "'address': 'Address'," +
                              " 'extra_2': 'data three'}"),
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=self.api_source)
        new_item.facility = self.facility_two
        new_item.save()
        FacilityMatch.objects.create(
            facility_list_item=new_item,
            facility=self.facility_two,
            is_active=True,
            results={}
        )

        index_one = FacilityIndex.objects.get(id=self.facility.id)
        index_one_data = '{}|data one'.format(self.contributor.id)
        self.assertIn(index_one_data, index_one.custom_text)
        index_two = FacilityIndex.objects.get(id=self.facility_two.id)
        index_two_data = '{}|data two'.format(self.contributor.id)
        self.assertNotIn(index_two_data, index_two.custom_text)
        index_three_data = '{}|data three'.format(self.contributor.id)
        self.assertIn(index_three_data, index_two.custom_text)

    def test_custom_text_uses_field_contributor(self):
        user_two = User.objects.create(email='test2@example.com')
        contributor_two = Contributor \
            .objects \
            .create(admin=user_two,
                    name='test contributor 2',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        api_source_two = Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True,
                    contributor=contributor_two)

        new_item = FacilityListItem \
            .objects \
            .create(name='Item',
                    address='Address',
                    country_code='US',
                    raw_data=("{'country': 'US', 'name': 'Item'," +
                              "'address': 'Address'," +
                              " 'extra_2': 'data three'}"),
                    row_index=1,
                    geocoded_point=Point(0, 0),
                    status=FacilityListItem.CONFIRMED_MATCH,
                    source=api_source_two)
        new_item.facility = self.facility_two
        new_item.save()
        FacilityMatch.objects.create(
            facility_list_item=new_item,
            facility=self.facility_two,
            is_active=True,
            results={}
        )

        index_one = FacilityIndex.objects.get(id=self.facility.id)
        index_one_data = '{}|data one'.format(self.contributor.id)
        self.assertIn(index_one_data, index_one.custom_text)
        index_two = FacilityIndex.objects.get(id=self.facility_two.id)
        index_two_data = '{}|data two'.format(self.contributor.id)
        self.assertIn(index_two_data, index_two.custom_text)
        index_three_data = '{}|data three'.format(contributor_two.id)
        self.assertNotIn(index_three_data, index_two.custom_text)


class ContributorManagerTest(TestCase):
    fixtures = ['users', 'contributors']

    def test_filter_by_name(self):
        matches = Contributor.objects.filter_by_name('factory a')
        self.assertGreater(matches.count(), 0)

        # No result should be returned for exact match
        matches = Contributor.objects.filter_by_name('factory')
        self.assertEqual(matches.count(), 0)

    def test_filter_by_name_verified(self):
        user1 = User.objects.create(email='test1@test.com')
        user2 = User.objects.create(email='test2@test.com')
        c1 = Contributor \
            .objects \
            .create(admin=user1,
                    name='TESTING',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)
        c2 = Contributor \
            .objects \
            .create(admin=user2,
                    name='TESTING',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        matches = Contributor.objects.filter_by_name('TESTING')
        self.assertEqual(2, matches.count())
        # When the names are the same and neither is verified then the second
        # contributor happens to sort first
        self.assertEqual(c2, matches[0])

        c1.is_verified = True
        c1.save()
        matches = Contributor.objects.filter_by_name('TESTING')
        self.assertEqual(2, matches.count())
        # Marking c1 as verified forces it to sort first
        self.assertEqual(c1, matches[0])

    def test_filter_by_name_source(self):
        user1 = User.objects.create(email='test1@test.com')
        user2 = User.objects.create(email='test2@test.com')
        c1 = Contributor \
            .objects \
            .create(admin=user1,
                    name='TESTING',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)
        c2 = Contributor \
            .objects \
            .create(admin=user2,
                    name='TESTING',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        matches = Contributor.objects.filter_by_name('TESTING')
        self.assertEqual(2, matches.count())
        # When the names are the same and neither is verified than the second
        # contributor happens to sort first
        self.assertEqual(c2, matches[0])

        Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True,
                    contributor=c1)

        matches = Contributor.objects.filter_by_name('TESTING')
        self.assertEqual(2, matches.count())
        # An active source forces it to sort first
        self.assertEqual(c1, matches[0])

    def test_filter_by_name_verified_and_source(self):
        user1 = User.objects.create(email='test1@test.com')
        user2 = User.objects.create(email='test2@test.com')
        c1 = Contributor \
            .objects \
            .create(admin=user1,
                    name='TESTING',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)
        c2 = Contributor \
            .objects \
            .create(admin=user2,
                    name='TESTING',
                    contrib_type=Contributor.OTHER_CONTRIB_TYPE)

        Source \
            .objects \
            .create(source_type=Source.SINGLE,
                    is_active=True,
                    is_public=True,
                    contributor=c1)

        c2.is_verified = True
        c2.save()

        matches = Contributor.objects.filter_by_name('TESTING')
        self.assertEqual(2, matches.count())
        # A verified contributor sorts before one with a source
        self.assertEqual(c2, matches[0])


class ParentCompanyTestCase(FacilityAPITestCaseBase):
    def setUp(self):
        super(ParentCompanyTestCase, self).setUp()
        self.url = reverse('facility-list')

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_submit_parent_company_no_match(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, {
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'parent_company': 'A random value'
        })
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.PARENT_COMPANY, ef.field_name)
        self.assertEqual({
            'raw_value': 'A random value',
            'name': 'A random value'
        }, ef.value)
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertIn('A random value', facility_index.parent_company_name)
        self.assertEquals(0, len(facility_index.parent_company_id))

    @skip('Skip fuzzy matching. Will revisit at #1805')
    @patch('api.geocoding.requests.get')
    def test_submit_parent_company_fuzzy_match(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, {
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'parent_company': 'TEST CNTRIBUTOR'
        })
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.PARENT_COMPANY, ef.field_name)
        self.assertEqual({
            'raw_value': 'TEST CNTRIBUTOR',
            'contributor_name': self.contributor.name,
            'contributor_id': self.contributor.id
        }, ef.value)
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertIn(self.contributor.name,
                      facility_index.parent_company_name)
        self.assertIn(self.contributor.id, facility_index.parent_company_id)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_submit_parent_company_duplicate(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, {
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'parent_company': 'test contributor 1'
        })
        self.client.post(self.url, {
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'parent_company': 'test contributor 1'
        })
        self.assertEqual(2, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertIn(self.contributor.name,
                      facility_index.parent_company_name)
        self.assertEqual(1, len(facility_index.parent_company_name))
        self.assertIn(self.contributor.id, facility_index.parent_company_id)
        self.assertEqual(1, len(facility_index.parent_company_id))

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_search_by_name(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        facility_response = self.client.post(self.url, {
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'parent_company': 'test contributor 1'
        })
        facility_data = json.loads(facility_response.content)
        facility_id = facility_data['oar_id']

        response = self.client.get(
            self.url + '?parent_company={}'.format(self.contributor.id)
        )
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], facility_id)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_search_by_id(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        facility_response = self.client.post(self.url, {
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'parent_company': 'test contributor 1'
        })
        facility_data = json.loads(facility_response.content)
        facility_id = facility_data['oar_id']

        response = self.client.get(
            self.url + '?parent_company={}'.format(self.contributor.name)
        )
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], facility_id)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_search_by_multiple(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        facility_response = self.client.post(self.url, {
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'parent_company': 'test contributor 1'
        })
        facility_data = json.loads(facility_response.content)
        facility_id = facility_data['oar_id']

        response = self.client.get(
            self.url + '?parent_company={}?parent_company={}'.format(
                self.contributor.name, self.contributor.id
            )
        )
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], facility_id)


class ProductTypeTestCase(FacilityAPITestCaseBase):
    def setUp(self):
        super(ProductTypeTestCase, self).setUp()
        self.url = reverse('facility-list')

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_array(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'product_type': ['a', 'b']
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.PRODUCT_TYPE, ef.field_name)
        self.assertEqual({
            'raw_values': ['a', 'b']
        }, ef.value)
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertIn('a', facility_index.product_type)
        self.assertIn('b', facility_index.product_type)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_string(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'product_type': 'a|b'
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.PRODUCT_TYPE, ef.field_name)
        self.assertEqual({
            'raw_values': ['a', 'b']
        }, ef.value)
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertIn('a', facility_index.product_type)
        self.assertIn('b', facility_index.product_type)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_list_validation(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'product_type': {}
        }), content_type='application/json')
        self.assertEqual(0, ExtendedField.objects.all().count())
        self.assertEqual(response.status_code, 400)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_max_count(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'product_type': [str(a) for a in range(MAX_PRODUCT_TYPE_COUNT)]
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertEqual(MAX_PRODUCT_TYPE_COUNT,
                         len(facility_index.product_type))

        response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'product_type': [str(a) for a in range(MAX_PRODUCT_TYPE_COUNT + 1)]
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.all().count())
        self.assertEqual(response.status_code, 400)
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertEqual(MAX_PRODUCT_TYPE_COUNT,
                         len(facility_index.product_type))

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_search_by_product_type(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        facility_response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'product_type': ['a', 'b']
        }), content_type='application/json')

        facility_data = json.loads(facility_response.content)
        facility_id = facility_data['oar_id']

        response = self.client.get(self.url + '?product_type=A')
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], facility_id)


class FacilityAndProcessingTypeAPITest(FacilityAPITestCaseBase):
    def setUp(self):
        super(FacilityAndProcessingTypeAPITest, self).setUp()
        self.url = reverse('facility-list')

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_single_processing_value(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'processing_type': ['cutting']
        }), content_type='application/json')
        self.assertEqual(2, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.FACILITY_TYPE, ef.field_name)
        self.assertEqual(
            {'matched_values': [
                [
                    'PROCESSING_TYPE',
                    'EXACT',
                    'Final Product Assembly',
                    'Cutting'
                ]
            ], 'raw_values': ['cutting']}, ef.value)
        ef = ExtendedField.objects.last()
        self.assertEqual(ExtendedField.PROCESSING_TYPE, ef.field_name)
        self.assertEqual(
            {'matched_values': [
                [
                    'PROCESSING_TYPE',
                    'EXACT',
                    'Final Product Assembly',
                    'Cutting'
                ]
            ], 'raw_values': ['cutting']}, ef.value)
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertNotIn('Cutting', facility_index.facility_type)
        self.assertIn('Final Product Assembly', facility_index.facility_type)
        self.assertIn('Cutting', facility_index.processing_type)
        self.assertNotIn('Final Product Assembly',
                         facility_index.processing_type)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_multiple_facility_values(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'facility_type': ['office hq', 'final product assembly']
        }), content_type='application/json')
        self.assertEqual(2, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.FACILITY_TYPE, ef.field_name)
        self.assertEqual(
            {'matched_values': [
                [
                    'FACILITY_TYPE',
                    'EXACT',
                    'Office / HQ',
                    'Office / HQ'
                ],
                [
                    'FACILITY_TYPE',
                    'EXACT',
                    'Final Product Assembly',
                    'Final Product Assembly'
                ]
            ], 'raw_values': ['office hq', 'final product assembly']},
            ef.value)
        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertIn('Final Product Assembly', facility_index.facility_type)
        self.assertIn('Office / HQ', facility_index.facility_type)
        self.assertEqual(0, len(facility_index.processing_type))

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_non_taxonomy_value(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'facility_type_processing_type': 'sewing|not a taxonomy value'
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.filter(
            field_name='facility_type').count())
        self.assertEqual(1, ExtendedField.objects.filter(
            field_name='processing_type').count())
        self.assertEqual(response.status_code, 201)

        data = json.loads(response.content)
        index_row = FacilityIndex.objects.filter(id=data['oar_id']).first()
        self.assertEqual(['Final Product Assembly'], index_row.facility_type)
        self.assertEqual(['Sewing'], index_row.processing_type)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_search_by_processing_type(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        facility_response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'processing_type': ['cutting']
        }), content_type='application/json')
        facility_data = json.loads(facility_response.content)
        facility_id = facility_data['oar_id']

        response = self.client.get(self.url + '?processing_type=cutting')
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], facility_id)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_search_by_facility_type(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        facility_response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'facility_type': ['office hq', 'final product assembly']
        }), content_type='application/json')
        facility_data = json.loads(facility_response.content)
        facility_id = facility_data['oar_id']

        response = self.client.get(
            self.url + '?facility_type=final%20product%20assembly'
        )
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], facility_id)


class NumberOfWorkersAPITest(FacilityAPITestCaseBase):
    def setUp(self):
        super(NumberOfWorkersAPITest, self).setUp()
        self.url = reverse('facility-list')

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_single_value(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'number_of_workers': '2000'
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.NUMBER_OF_WORKERS, ef.field_name)
        self.assertEqual({'min': 2000, 'max': 2000}, ef.value)

        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertEquals(1, len(facility_index.number_of_workers))
        self.assertIn('1001-5000', facility_index.number_of_workers)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_range_value(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'number_of_workers': '0-500'
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.NUMBER_OF_WORKERS, ef.field_name)
        self.assertEqual({'min': 0, 'max': 500}, ef.value)

        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertEquals(1, len(facility_index.number_of_workers))
        self.assertIn('Less than 1000', facility_index.number_of_workers)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_crossrange_value(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'number_of_workers': '0 to 10000'
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.NUMBER_OF_WORKERS, ef.field_name)
        self.assertEqual({'min': 0, 'max': 10000}, ef.value)

        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertEquals(3, len(facility_index.number_of_workers))
        self.assertIn('Less than 1000', facility_index.number_of_workers)
        self.assertIn('1001-5000', facility_index.number_of_workers)
        self.assertIn('5001-10000', facility_index.number_of_workers)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_maxrange_value(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'number_of_workers': '20,000-100,000'
        }), content_type='application/json')
        self.assertEqual(1, ExtendedField.objects.all().count())
        ef = ExtendedField.objects.first()
        self.assertEqual(ExtendedField.NUMBER_OF_WORKERS, ef.field_name)
        self.assertEqual({'min': 20000, 'max': 100000}, ef.value)

        facility_index = FacilityIndex.objects.get(id=ef.facility.id)
        self.assertEquals(1, len(facility_index.number_of_workers))
        self.assertIn('More than 10000', facility_index.number_of_workers)

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_search_by_range(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        facility_response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'number_of_workers': '1500 to 2000'
        }), content_type='application/json')

        facility_data = json.loads(facility_response.content)
        facility_id = facility_data['oar_id']

        response = self.client.get(
            self.url + '?number_of_workers=1001-5000'
        )
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], facility_id)

    @patch('api.geocoding.requests.get')
    def test_search_without_matches(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'number_of_workers': '1000 to 9000'
        }), content_type='application/json')

        response = self.client.get(
            self.url + '?number_of_workers=More%20than%2010000'
        )
        data = json.loads(response.content)
        self.assertEquals(data['count'], 0)


class NativeLanguageNameAPITest(FacilityAPITestCaseBase):
    def setUp(self):
        super(NativeLanguageNameAPITest, self).setUp()
        self.url = reverse('facility-list')
        self.long_name = '杭州湾开发区兴慈二路与滨海二路叉口恒元工业园区A3'

    @patch('api.geocoding.requests.get')
    @skip('DB is read-only')
    def test_search(self, mock_get):
        mock_get.return_value = Mock(ok=True, status_code=200)
        mock_get.return_value.json.return_value = geocoding_data
        self.join_group_and_login()
        facility_response = self.client.post(self.url, json.dumps({
            'country': "US",
            'name': "Azavea",
            'address': "990 Spring Garden St., Philadelphia PA 19123",
            'native_language_name': self.long_name
        }), content_type='application/json')

        facility_data = json.loads(facility_response.content)
        facility_id = facility_data['oar_id']

        response = self.client.get(
            self.url + '?native_language_name={}'.format(self.long_name)
        )
        data = json.loads(response.content)
        self.assertEquals(data['count'], 1)
        self.assertEquals(data['features'][0]['id'], facility_id)


class ExactMatchTest(FacilityAPITestCaseBase):
    def setUp(self):
        super(ExactMatchTest, self).setUp()
        self.contributor = Contributor.objects.first()

        self.email = 'test2@example.com'
        self.password = 'password'
        self.name = 'Test User 2'
        self.user = User(email=self.email)
        self.user.set_password(self.password)
        self.user.save()
        self.contributor_2 = Contributor.objects.create(
            name=self.name, admin=self.user)

        self.active_item_ids = [1, 2, 4]
        self.contributor.save()

    def test_sorting_match_contrib(self):
        matches = [{
            'id': 1,
            'facility_id': 1,
            'source__contributor_id': self.contributor_2.id,
            'updated_at': datetime.now(),
        }, {
            'id': 2,
            'facility_id': 2,
            'source__contributor_id': self.contributor.id,
            'updated_at': datetime.now(),
        }]
        results = sort_exact_matches(matches, self.active_item_ids,
                                     self.contributor)
        self.assertEquals(results[0]['facility_id'], 2)
        self.assertEquals(results[1]['facility_id'], 1)

    def test_sorting_active(self):
        matches = [{
            'id': 1,
            'facility_id': 3,
            'source__contributor_id': self.contributor.id,
            'updated_at': datetime.now(),
        }, {
            'id': 1,
            'facility_id': 1,
            'source__contributor_id': self.contributor.id,
            'updated_at': datetime.now(),
        }]
        results = sort_exact_matches(matches, self.active_item_ids,
                                     self.contributor)
        self.assertEquals(results[0]['facility_id'], 1)
        self.assertEquals(results[1]['facility_id'], 3)

    def test_sorting_newest(self):
        matches = [{
            'id': 1,
            'facility_id': 2,
            'source__contributor_id': self.contributor.id,
            'updated_at': datetime.now().replace(year=2020),
        }, {
            'id': 1,
            'facility_id': 1,
            'source__contributor_id': self.contributor.id,
            'updated_at': datetime.now(),
        }]
        results = sort_exact_matches(matches, self.active_item_ids,
                                     self.contributor)
        self.assertEquals(results[0]['facility_id'], 1)
        self.assertEquals(results[1]['facility_id'], 2)
